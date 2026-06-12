#!/usr/bin/env python3
"""
PRISME - Generateur Open Data (Chantier A)
Genere les fichiers Excel PRISME et ZIP pour:
- educ, pers_sup65ans_seules, familles_mono, pop_inf3ans
- pers_menages, types_menages, alloc, revenu, densite
- route (ONISR/BAAC accidents de la route, commune-level)
- mortalite_gen, mortalite_cardio, mortalite_tumeurs, mortalite_respi (CepiDc, regional)

Validation attendue: chaque ZIP contient 5 dossiers
Commune/Region/DOM/France_Hexagonale/France_Entiere
"""

from pathlib import Path
import argparse
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from csv_reader import read_csv_safe, log_read, normalize_geo_code


BASE_DIR = Path(__file__).parent
INPUTS_DIR = BASE_DIR / "inputs" / "opendata"
OUTPUT_DIR = BASE_DIR / "output"

ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


def is_calculated_variable(var_name: str) -> bool:
    """Retourne True pour les variables calculées côté client PRISME (tx_*).

    PRISME recalcule les taux automatiquement : l'outil d'automatisation
    doit fournir uniquement les données brutes (effectifs).
    """
    return bool(var_name) and str(var_name).lower().startswith('tx_')

GEO_FOLDER_MAPPING = {
    "com": "Commune",
    "reg": "Region",
    "dom": "DOM",
    "fh": "France_Hexagonale",
    "fra": "France_Entiere",
}

COMMUNES_GUYANE = [
    "97301",
    "97302",
    "97303",
    "97304",
    "97305",
    "97306",
    "97307",
    "97308",
    "97309",
    "97310",
    "97311",
    "97312",
    "97313",
    "97314",
    "97352",
    "97353",
    "97356",
    "97357",
    "97358",
    "97360",
    "97361",
    "97362",
]

REGION_ORDER = ["11", "24", "27", "28", "32", "44", "52", "53", "75", "76", "84", "93", "94", "01", "02", "03", "04", "06"]
DOM_CODES = ["01", "02", "03", "04", "06"]
FH_REGIONS = ["11", "24", "27", "28", "32", "44", "52", "53", "75", "76", "84", "93", "94"]

DEP_TO_REG = {
    "01": "84", "02": "32", "03": "84", "04": "93", "05": "93", "06": "93", "07": "84", "08": "44", "09": "76",
    "10": "44", "11": "76", "12": "76", "13": "93", "14": "28", "15": "84", "16": "75", "17": "75", "18": "24",
    "19": "75", "21": "27", "22": "53", "23": "75", "24": "75", "25": "27", "26": "84", "27": "28", "28": "24",
    "29": "53", "2A": "94", "2B": "94", "30": "76", "31": "76", "32": "76", "33": "75", "34": "76", "35": "53",
    "36": "24", "37": "24", "38": "84", "39": "27", "40": "75", "41": "24", "42": "84", "43": "84", "44": "52",
    "45": "24", "46": "76", "47": "75", "48": "76", "49": "52", "50": "28", "51": "44", "52": "44", "53": "52",
    "54": "44", "55": "44", "56": "53", "57": "44", "58": "27", "59": "32", "60": "32", "61": "28", "62": "32",
    "63": "84", "64": "75", "65": "76", "66": "76", "67": "44", "68": "44", "69": "84", "70": "27", "71": "27",
    "72": "52", "73": "84", "74": "84", "75": "11", "76": "28", "77": "11", "78": "11", "79": "75", "80": "32",
    "81": "76", "82": "76", "83": "93", "84": "93", "85": "52", "86": "75", "87": "75", "88": "44", "89": "27",
    "90": "27", "91": "11", "92": "11", "93": "11", "94": "11", "95": "11",
    "971": "01", "972": "02", "973": "03", "974": "04", "976": "06",
}

THEME_CONFIGS = {
    "educ": {
        "excel_name": "educ",
        "variables": ["pop_6_16", "nb_non_sco", "pop_15_64", "nb_peu_dipl"],
        "source_type": "educ",
    },
    "pers_sup65ans_seules": {
        "excel_name": "pers_sup65ans_seules",
        "variables": ["nb_pop_65ans", "nb_pop_65ans_seule"],
        "source_type": "couples",
    },
    "familles_mono": {
        "excel_name": "familles_mono",
        "variables": ["nb_familles_enf", "nb_familles_mono_enf"],
        "source_type": "couples",
    },
    "pop_inf3ans": {"excel_name": "pop_inf3ans", "variables": ["pop_inf3ans", "rp"], "source_type": "couples"},
    "pers_menages": {"excel_name": "pers_menages", "variables": ["nb_menages", "nb_pers_menages"], "source_type": "couples"},
    "types_menages": {
        "excel_name": "types_menages",
        "variables": ["nb_men_seul", "nb_men_couple_senf", "nb_men_couple_aenf", "nb_men_fam_mono"],
        "source_type": "couples",
    },
    "alloc": {"excel_name": "alloc", "variables": ["nb_alloc", "nb_menages"], "source_type": "caf"},
    "revenu": {"excel_name": "revenu", "variables": ["nb_foyers_non_impo", "nb_foyers_imposes"], "source_type": "ircom"},
    "densite": {"excel_name": "densite", "variables": ["rp", "superficie"], "source_type": "pop_legales"},
    "route": {"excel_name": "route", "variables": ["nb_acci", "nb_blesses", "nb_morts"], "source_type": "baac"},
    "accidents_route": {"excel_name": "accidents_route", "variables": ["nb_acci"], "source_type": "baac"},
    "blesses_route": {"excel_name": "blesses_route", "variables": ["nb_blesses"], "source_type": "baac"},
    "deces_route": {"excel_name": "deces_route", "variables": ["nb_morts"], "source_type": "baac"},
    "mortalite_gen": {
        "excel_name": "mortalite_gen",
        "variables": ["nb_deces_toutes_causes", "tx_mortalite_toutes_causes"],
        "source_type": "cepidc",
        "cepidc_sheet": "Toutes causes",
    },
    "mortalite_cardio": {
        "excel_name": "mortalite_cardio",
        "variables": ["nb_deces_cardio", "tx_mortalite_cardio"],
        "source_type": "cepidc",
        "cepidc_sheet": "Cardio",
    },
    "mortalite_tumeurs": {
        "excel_name": "mortalite_tumeurs",
        "variables": ["nb_deces_tumeurs", "tx_mortalite_tumeurs"],
        "source_type": "cepidc",
        "cepidc_sheet": "Tumeurs",
    },
    "mortalite_respi": {
        "excel_name": "mortalite_respi",
        "variables": ["nb_deces_respi", "tx_mortalite_respi"],
        "source_type": "cepidc",
        "cepidc_sheet": "Respi",
    },
    "mortalite_neuro": {
        "excel_name": "mortalite_neuro",
        "variables": ["nb_deces_neuro", "tx_mortalite_neuro"],
        "source_type": "cepidc",
        "cepidc_sheet": "Demence, Park, Alzh",
    },
    "mortalite_diabete": {
        "excel_name": "mortalite_diabete",
        "variables": ["nb_deces_endo", "tx_mortalite_endo"],
        "source_type": "cepidc",
        "cepidc_sheet": "Endo, nutri, meta",
    },
    "mortalite_covid": {
        "excel_name": "mortalite_covid",
        "variables": ["nb_deces_covid", "tx_mortalite_covid"],
        "source_type": "cepidc",
        "cepidc_sheet": "Covid19",
    },
    # Comportements / addictions (Odissé) - DEP-level for suicide, REG-level for alcool/tabac
    "comp_mortalite": {
        "excel_name": "comp_mortalite",
        "variables": ["m_suicide"],
        "source_type": "odisse_suicide",
    },
    "suicide": {
        "excel_name": "suicide",
        "variables": ["m_suicide"],
        "source_type": "odisse_suicide",
    },
    "addictions_alcool": {
        "excel_name": "addictions_alcool",
        "variables": ["prev_alcool_quotidien"],
        "source_type": "odisse_alcool",
    },
    "addictions_tabac": {
        "excel_name": "addictions_tabac",
        "variables": ["prev_tabac_quotidien"],
        "source_type": "odisse_tabac",
    },
    "noyades": {
        "excel_name": "noyades",
        "variables": ["nb_noyades", "tx_noyades", "nb_noyades_deces", "tx_noyades_deces"],
        "source_type": "spf_noyades",
        # Les taux noyades sont calculés ici (nb / population x 1 000 000 hab.) et
        # livrés directement dans l'Excel (demande client), donc NON filtrés
        # malgré le préfixe tx_ (cf. emit_calculated_rates ci-dessous).
        "emit_calculated_rates": True,
    },
    # DREES - Offre d'accueil du jeune enfant (EAJE), niveau departement -> agrege
    "accueil_pop_inf3ans": {
        "excel_name": "accueil_pop_inf3ans",
        "variables": [
            "nb_places_acceuil_co", "nb_places_mono", "nb_places_hg", "nb_places_je",
            "nb_places_multi", "nb_places_familial",
            "nb_etab_co", "nb_etab_mono", "nb_etab_hg", "nb_etab_je", "nb_etab_multi",
            "nb_services_familial", "nb_mam",
        ],
        "source_type": "drees_eaje",
    },
}


def _read_csv_auto(path: Path, dtype=None) -> pd.DataFrame:
    """Lecture robuste : encoding + séparateur auto-détectés, NA normalisés."""
    df, meta = read_csv_safe(path, dtype=dtype)
    log_read(meta)
    return df


def _safe_numeric(df: pd.DataFrame, candidates) -> pd.Series:
    for col in candidates:
        if col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").fillna(0)
    return pd.Series([0] * len(df), index=df.index, dtype="float64")


def _extract_commune_code(value) -> str:
    code = str(value).strip()
    if len(code) >= 5:
        return code[:5]
    return code.zfill(5)


def _dep_from_commune(code: str) -> str:
    if code.startswith("97"):
        return code[:3]
    return code[:2]


def _with_geo_columns(df: pd.DataFrame, code_col: str) -> pd.DataFrame:
    out = df.copy()
    out["commune_code"] = out[code_col].apply(_extract_commune_code)
    out["dep"] = out["commune_code"].apply(_dep_from_commune)
    out["reg"] = out["dep"].map(DEP_TO_REG)
    return out


def _aggregate_levels(df: pd.DataFrame, value_columns, code_col: str = "commune_code"):
    all_df = _with_geo_columns(df, code_col)
    value_columns = [c for c in value_columns if c in all_df.columns]
    if not value_columns:
        raise ValueError("Aucune colonne numerique a aggreger")
    for col in value_columns:
        all_df[col] = pd.to_numeric(all_df[col], errors="coerce").fillna(0)

    # Communes Guyane
    com = all_df[all_df["commune_code"].isin(COMMUNES_GUYANE)].groupby("commune_code")[value_columns].sum().reset_index()
    com.rename(columns={"commune_code": "codgeo"}, inplace=True)

    # Regions
    reg = all_df.dropna(subset=["reg"]).groupby("reg")[value_columns].sum().reset_index()
    reg.rename(columns={"reg": "codgeo"}, inplace=True)
    reg_full = pd.DataFrame({"codgeo": REGION_ORDER})
    reg = reg_full.merge(reg, on="codgeo", how="left").fillna(0)

    # Coverage check: how many regions actually have non-zero data?
    # If coverage is < 50%, aggregated values are misleading (e.g. only Guyane data for BAAC).
    dom_regions_with_data = (reg[reg["codgeo"].isin(DOM_CODES)][value_columns].sum(axis=1) > 0).sum()
    fh_regions_with_data = (reg[reg["codgeo"].isin(FH_REGIONS)][value_columns].sum(axis=1) > 0).sum()
    total_regions_with_data = dom_regions_with_data + fh_regions_with_data

    dom_coverage_ok = dom_regions_with_data >= len(DOM_CODES) * 0.5
    fh_coverage_ok = fh_regions_with_data >= len(FH_REGIONS) * 0.5
    fra_coverage_ok = total_regions_with_data >= len(REGION_ORDER) * 0.5

    if not fra_coverage_ok:
        print(f"  [WARN_DATA] Couverture regionale insuffisante ({total_regions_with_data}/{len(REGION_ORDER)} regions). "
              "Les niveaux DOM/FH/FRA peuvent etre incomplets ou vides.")

    # DOM / FH / FR — set to NaN if coverage too low
    if dom_coverage_ok:
        dom_sum = reg[reg["codgeo"].isin(DOM_CODES)][value_columns].sum().to_frame().T
    else:
        dom_sum = pd.DataFrame({c: [float("nan")] for c in value_columns})
    dom_sum["codgeo"] = "DOM"
    dom = dom_sum[["codgeo"] + value_columns]

    if fh_coverage_ok:
        fh_sum = reg[reg["codgeo"].isin(FH_REGIONS)][value_columns].sum().to_frame().T
    else:
        fh_sum = pd.DataFrame({c: [float("nan")] for c in value_columns})
    fh_sum["codgeo"] = "0"
    fh = fh_sum[["codgeo"] + value_columns]

    if fra_coverage_ok:
        fra_sum = reg[value_columns].sum().to_frame().T
    else:
        fra_sum = pd.DataFrame({c: [float("nan")] for c in value_columns})
    fra_sum["codgeo"] = "99"
    fra = fra_sum[["codgeo"] + value_columns]

    return {"com": com, "reg": reg, "dom": dom, "fh": fh, "fra": fra}


def _add_year(all_levels, year: int):
    out = {}
    for level, df in all_levels.items():
        copy_df = df.copy()
        copy_df["annee"] = year
        out[level] = copy_df
    return out


def _calc_from_educ(df: pd.DataFrame, year: int) -> pd.DataFrame:
    prefix = f"P{str(year)[2:]}_"
    result = pd.DataFrame({"codgeo": df["codgeo"]})

    pop_0610 = _safe_numeric(df, [f"{prefix}POP0610"])
    pop_1114 = _safe_numeric(df, [f"{prefix}POP1114"])
    pop_1517 = _safe_numeric(df, [f"{prefix}POP1517"])
    pop_6_16 = pop_0610 + pop_1114 + (2.0 / 3.0) * pop_1517

    scol_0610 = _safe_numeric(df, [f"{prefix}SCOL0610"])
    scol_1114 = _safe_numeric(df, [f"{prefix}SCOL1114"])
    scol_1517 = _safe_numeric(df, [f"{prefix}SCOL1517"])
    nb_non_sco = (pop_0610 - scol_0610).clip(lower=0) + (pop_1114 - scol_1114).clip(lower=0) + ((pop_1517 - scol_1517).clip(lower=0) * (2.0 / 3.0))

    pop_1524 = _safe_numeric(df, [f"{prefix}POP1524"])
    pop_2554 = _safe_numeric(df, [f"{prefix}POP2554"])
    pop_5564 = _safe_numeric(df, [f"{prefix}POP5564"])
    pop_15_64 = pop_1524 + pop_2554 + pop_5564
    if float(pop_15_64.sum()) == 0:
        pop_1824 = _safe_numeric(df, [f"{prefix}POP1824"])
        pop_2529 = _safe_numeric(df, [f"{prefix}POP2529"])
        pop_30p = _safe_numeric(df, [f"{prefix}POP30P"])
        pop_15_64 = pop_1517 + pop_1824 + pop_2529 + (0.88 * pop_30p)

    nb_peu_dipl = _safe_numeric(df, [f"{prefix}NSCOL15P_DIPLMIN", f"{prefix}NSCOL15P"])

    result["pop_6_16"] = pop_6_16.round(2)
    result["nb_non_sco"] = nb_non_sco.round(2)
    result["pop_15_64"] = pop_15_64.round(2)
    result["nb_peu_dipl"] = nb_peu_dipl.round(2)
    return result


def _calc_from_couples(df: pd.DataFrame, theme: str, year: int) -> pd.DataFrame:
    prefix_c = f"C{str(year)[2:]}_"
    prefix_p = f"P{str(year)[2:]}_"
    result = pd.DataFrame({"codgeo": df["codgeo"]})

    if theme == "pers_sup65ans_seules":
        pop_65p = _safe_numeric(df, [f"{prefix_p}POP65P"])
        if float(pop_65p.sum()) == 0:
            pop_5579 = _safe_numeric(df, [f"{prefix_p}POP5579"])
            pop_80p = _safe_numeric(df, [f"{prefix_p}POP80P"])
            pop_65p = (15.0 / 25.0) * pop_5579 + pop_80p

        seules_65p = _safe_numeric(df, [f"{prefix_c}PMEN_MENPSEUL65P", f"{prefix_p}POP65P_PSEUL"])
        if float(seules_65p.sum()) == 0:
            pseul_5579 = _safe_numeric(df, [f"{prefix_p}POP5579_PSEUL"])
            pseul_80p = _safe_numeric(df, [f"{prefix_p}POP80P_PSEUL"])
            seules_65p = (15.0 / 25.0) * pseul_5579 + pseul_80p

        result["nb_pop_65ans"] = pop_65p.round(2)
        result["nb_pop_65ans_seule"] = seules_65p.round(2)
        return result

    if theme == "familles_mono":
        coup_enf = _safe_numeric(df, [f"{prefix_c}COUPAENF", f"{prefix_c}FAM_COUPAENF", f"{prefix_c}MENCOUPAENF"])
        fam_mono = _safe_numeric(df, [f"{prefix_c}FAMMONO", f"{prefix_c}FAM_MONO", f"{prefix_c}MENFAMMONO"])
        result["nb_familles_enf"] = (coup_enf + fam_mono).round(2)
        result["nb_familles_mono_enf"] = fam_mono.round(2)
        return result

    if theme == "pop_inf3ans":
        pop_inf3 = _safe_numeric(df, [f"{prefix_p}POP0002", f"{prefix_p}POP0003"])
        if float(pop_inf3.sum()) == 0:
            ne24f1 = _safe_numeric(df, [f"{prefix_c}NE24F1"])
            ne24f2 = _safe_numeric(df, [f"{prefix_c}NE24F2"])
            ne24f3 = _safe_numeric(df, [f"{prefix_c}NE24F3"])
            ne24f4p = _safe_numeric(df, [f"{prefix_c}NE24F4P"])
            total_enfants = ne24f1 + 2 * ne24f2 + 3 * ne24f3 + 4 * ne24f4p
            pop_inf3 = total_enfants * 0.12
        rp = _safe_numeric(df, [f"{prefix_p}POP", f"{prefix_c}PMEN"])
        result["pop_inf3ans"] = pop_inf3.round(2)
        result["rp"] = rp.round(2)
        return result

    if theme == "pers_menages":
        result["nb_menages"] = _safe_numeric(df, [f"{prefix_c}MEN"]).round(2)
        result["nb_pers_menages"] = _safe_numeric(df, [f"{prefix_c}PMEN"]).round(2)
        return result

    if theme == "types_menages":
        result["nb_men_seul"] = _safe_numeric(df, [f"{prefix_c}MENPSEUL"]).round(2)
        result["nb_men_couple_senf"] = _safe_numeric(df, [f"{prefix_c}MENCOUPSENF"]).round(2)
        result["nb_men_couple_aenf"] = _safe_numeric(df, [f"{prefix_c}MENCOUPAENF"]).round(2)
        result["nb_men_fam_mono"] = _safe_numeric(df, [f"{prefix_c}MENFAMMONO"]).round(2)
        return result

    raise ValueError(f"Theme couples non supporte: {theme}")


def _load_couples_source(year: int) -> Path:
    path = INPUTS_DIR / f"couples_familles_{year}.csv"
    if not path.exists():
        raise FileNotFoundError(f"Source manquante: {path}")
    return path


def _load_educ_source(year: int) -> Path:
    path = INPUTS_DIR / f"diplomes_formation_{year}.csv"
    if path.exists():
        return path
    # List available years for clear error message
    candidates = sorted(INPUTS_DIR.glob("diplomes_formation_*.csv"))
    available = [p.stem.split("_")[-1] for p in candidates]
    if candidates:
        raise FileNotFoundError(
            f"Source Open Data absente pour {year}. Annees disponibles: {', '.join(available)}"
        )
    raise FileNotFoundError(f"Aucune source Open Data education trouvee dans {INPUTS_DIR}")


def _build_couples_levels(theme: str, year: int):
    source = _load_couples_source(year)
    raw = _read_csv_auto(source, dtype={"IRIS": str, "COM": str, "CODGEO": str})
    code_col = "COM" if "COM" in raw.columns else "CODGEO"
    value_candidates = [c for c in raw.columns if c.startswith("C") or c.startswith("P")]
    base = _aggregate_levels(raw, value_candidates, code_col=code_col)
    calculated = {lvl: _calc_from_couples(df, theme, year) for lvl, df in base.items()}
    return _add_year(calculated, year)


def _build_educ_levels(year: int):
    source = _load_educ_source(year)
    raw = _read_csv_auto(source, dtype={"IRIS": str, "COM": str, "CODGEO": str})
    code_col = "COM" if "COM" in raw.columns else "CODGEO"
    value_candidates = [c for c in raw.columns if c.startswith("P")]
    base = _aggregate_levels(raw, value_candidates, code_col=code_col)
    calculated = {lvl: _calc_from_educ(df, year) for lvl, df in base.items()}
    return _add_year(calculated, year)


def _build_alloc_levels(year: int):
    sources = sorted(INPUTS_DIR.glob("caf_allocataires*.csv"))
    if not sources:
        raise FileNotFoundError(f"Source manquante: {INPUTS_DIR / 'caf_allocataires*.csv'}")

    raw = None
    for source in sources:
        candidate = _read_csv_auto(source)
        if "Date référence" not in candidate.columns:
            continue
        candidate["year"] = candidate["Date référence"].astype(str).str[:4]
        candidate = candidate[candidate["year"] == str(year)].copy()
        if not candidate.empty:
            raw = candidate
            break
    if raw is None:
        raise ValueError(f"Aucune ligne CAF pour l'annee {year}")

    if "Numéro commune" not in raw.columns:
        raise ValueError("Colonne 'Numéro commune' absente du fichier CAF")
    raw["commune_code"] = raw["Numéro commune"].astype(str).str.zfill(5)

    raw["nb_alloc"] = _safe_numeric(raw, ["Nombre foyers NDUR", "Nbre_foyers_alloc_total"])
    # Use current year couples if available, fallback to most recent
    couples_year = year
    try:
        couples_path = _load_couples_source(year)
    except FileNotFoundError:
        for fallback_yr in [year - 1, year - 2, year + 1]:
            try:
                couples_path = _load_couples_source(fallback_yr)
                couples_year = fallback_yr
                print(f"  [INFO] Couples {year} absent, fallback vers {fallback_yr}")
                break
            except FileNotFoundError:
                continue
        else:
            raise FileNotFoundError(f"Aucun fichier couples-familles disponible pour calculer nb_menages (annee {year})")
    couples = _read_csv_auto(couples_path, dtype={"IRIS": str, "COM": str, "CODGEO": str})
    code_col = "COM" if "COM" in couples.columns else "CODGEO"
    couples["commune_code"] = couples[code_col].apply(_extract_commune_code)
    prefix_c = f"C{str(couples_year)[2:]}_"
    couples["nb_menages"] = _safe_numeric(couples, [f"{prefix_c}MEN", "C22_MEN"])
    menages = couples.groupby("commune_code")["nb_menages"].sum().reset_index()

    raw = raw.merge(menages, on="commune_code", how="left")
    raw["nb_menages"] = raw["nb_menages"].fillna(0)

    levels = _aggregate_levels(raw, ["nb_alloc", "nb_menages"], code_col="commune_code")
    result = {}
    for lvl, df in levels.items():
        result[lvl] = pd.DataFrame(
            {
                "codgeo": df["codgeo"],
                "nb_alloc": _safe_numeric(df, ["nb_alloc"]).round(2),
                "nb_menages": _safe_numeric(df, ["nb_menages"]).round(2),
            }
        )
    return _add_year(result, year)


def _find_ircom_source(year: int) -> Path:
    # Priority 1: exact year match in filename
    for pattern in [f"*revenus_{year}*", f"*ircom*{year}*"]:
        candidates = sorted(INPUTS_DIR.rglob(pattern))
        xlsx = [c for c in candidates if c.is_file() and c.suffix.lower() in [".xlsx", ".xls"]]
        if xlsx:
            return xlsx[0]
        csv = [c for c in candidates if c.is_file() and c.suffix.lower() in [".csv", ".txt"]]
        if csv:
            return csv[0]
    # Priority 2: any ircom file (legacy fallback)
    candidates = sorted(INPUTS_DIR.rglob("*ircom*.xlsx"))
    candidates = [c for c in candidates if c.is_file()]
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"Aucun fichier IRCOM trouve pour {year} dans inputs/opendata")


def _load_ircom_dataframe(source: Path) -> pd.DataFrame:
    ext = source.suffix.lower()
    if ext in [".csv", ".txt"]:
        return _read_csv_auto(source)

    if ext in [".xlsx", ".xls"]:
        raw = pd.read_excel(source, sheet_name=0, header=None)
        header_row = None
        for i in range(min(30, len(raw))):
            vals = [str(v).strip().lower() if pd.notna(v) else "" for v in raw.iloc[i].tolist()]
            has_dep = any(v in ["dép.", "dep", "dép"] for v in vals)
            has_com = "commune" in vals
            if has_dep and has_com:
                header_row = i
                break
        if header_row is None:
            raise ValueError(f"Header IRCOM non detecte dans {source}")

        headers = []
        for i, val in enumerate(raw.iloc[header_row].tolist()):
            headers.append(str(val).strip() if pd.notna(val) else f"col_{i}")

        df = raw.iloc[header_row + 1:].copy()
        df.columns = headers
        df = df.dropna(how="all")

        dep_col = next((c for c in df.columns if c.lower() in ["dép.", "dep", "dép"]), None)
        com_col = next((c for c in df.columns if c.lower() == "commune"), None)
        tranche_col = next((c for c in df.columns if "tranche" in c.lower()), None)
        total_col = next((c for c in df.columns if c.strip().lower() == "nombre de foyers fiscaux"), None)
        impos_col = next((c for c in df.columns if "nombre de foyers fiscaux imposés" in c.lower()), None)

        if not dep_col or not com_col or not total_col or not impos_col:
            raise ValueError("Colonnes IRCOM attendues absentes (dep/commune/foyers/imposes)")

        if tranche_col:
            df = df[df[tranche_col].astype(str).str.strip().str.lower() == "total"].copy()

        def _fmt_dep(v):
            s = str(v).split(".")[0].strip().upper()
            if s in ["2A", "2B"]:
                return s
            return s if s.isdigit() else s

        def _fmt_com(v):
            return str(v).split(".")[0].strip().upper()

        def _code_commune(dep, com):
            d = _fmt_dep(dep)
            c = _fmt_com(com)
            if d in ["2A", "2B"]:
                return d + c.zfill(3)
            if d.isdigit():
                if d.startswith("97"):
                    return d.zfill(3) + c.zfill(3)[-2:]
                return d.zfill(2)[-2:] + c.zfill(3)
            return d + c.zfill(3)

        code_commune = [_code_commune(dep, com) for dep, com in zip(df[dep_col], df[com_col])]
        foyers_total = pd.to_numeric(df[total_col], errors="coerce").fillna(0)
        foyers_imposes = pd.to_numeric(df[impos_col], errors="coerce").fillna(0)

        return pd.DataFrame(
            {
                "code_commune": code_commune,
                "nb_foyers_imposes": foyers_imposes,
                "nb_foyers_non_impo": (foyers_total - foyers_imposes).clip(lower=0),
            }
        )

    raise ValueError(f"Format IRCOM non supporte: {source}")


def _build_revenu_levels(year: int):
    source = _find_ircom_source(year)
    raw = _load_ircom_dataframe(source)

    code_col_candidates = ["code_commune", "Code commune", "CODGEO", "commune_code", "Numéro commune", "COM"]
    code_col = next((c for c in code_col_candidates if c in raw.columns), None)
    if not code_col:
        raise ValueError("Aucune colonne code commune reconnue dans fichier IRCOM")

    year_cols = [c for c in ["annee", "Annee", "Année", "year", "millesime", "Millesime", "MILLÉSIME"] if c in raw.columns]
    if year_cols:
        yc = year_cols[0]
        raw = raw[pd.to_numeric(raw[yc], errors="coerce") == year].copy()
        if raw.empty:
            raise ValueError(f"Aucune ligne IRCOM pour l'annee {year}")

    raw["nb_foyers_non_impo"] = _safe_numeric(raw, ["foyers_non_imposables", "nb_foyers_non_impo", "NB_FOYERS_NON_IMPO"])
    raw["nb_foyers_imposes"] = _safe_numeric(raw, ["foyers_imposables", "nb_foyers_imposes", "NB_FOYERS_IMPOSES"])

    levels = _aggregate_levels(raw, ["nb_foyers_non_impo", "nb_foyers_imposes"], code_col=code_col)
    result = {}
    for lvl, df in levels.items():
        result[lvl] = pd.DataFrame(
            {
                "codgeo": df["codgeo"],
                "nb_foyers_non_impo": _safe_numeric(df, ["nb_foyers_non_impo"]).round(2),
                "nb_foyers_imposes": _safe_numeric(df, ["nb_foyers_imposes"]).round(2),
            }
        )
    return _add_year(result, year)


def _find_pop_legales_source(year: int) -> Path:
    candidates = [INPUTS_DIR / f"populations_{year}.csv", INPUTS_DIR / "donnees_communes.csv"]
    for c in candidates:
        if c.exists():
            return c
    found = sorted(INPUTS_DIR.glob("*communes*.csv"))
    if found:
        return found[0]
    raise FileNotFoundError("Aucune source populations legales trouvee")


def _build_densite_levels(year: int):
    source = _find_pop_legales_source(year)
    raw = _read_csv_auto(source)

    code_col = "COM" if "COM" in raw.columns else ("CODGEO" if "CODGEO" in raw.columns else None)
    if not code_col:
        raise ValueError("Source populations: colonne COM/CODGEO introuvable")

    raw["rp"] = _safe_numeric(raw, ["PMUN", "PTOT", "rp"])
    raw["superficie"] = _safe_numeric(raw, ["superficie", "SUPERFICIE", "SURFACE_KM2", "surface"])

    # Fallback Open Data: geo.api.gouv.fr (code, surface)
    if float(raw["superficie"].sum()) == 0:
        sup_json = INPUTS_DIR / "superficie_communes.json"
        if sup_json.exists():
            try:
                sup_df = pd.read_json(sup_json)
                if {"code", "surface"}.issubset(set(sup_df.columns)):
                    sup_df = sup_df[["code", "surface"]].copy()
                    sup_df["code"] = sup_df["code"].astype(str).str.zfill(5).str[:5]
                    # geo.api.gouv.fr returns surface in hectares, convert to km²
                    sup_df["surface"] = pd.to_numeric(sup_df["surface"], errors="coerce").fillna(0) / 100.0

                    raw["_geo_code"] = raw[code_col].astype(str).str.zfill(5).str[:5]
                    raw = raw.merge(sup_df, left_on="_geo_code", right_on="code", how="left")
                    raw["superficie"] = raw["surface"].fillna(raw["superficie"]).astype(float)
                    raw.drop(columns=["_geo_code", "code", "surface"], inplace=True, errors="ignore")
            except Exception:
                pass

    levels = _aggregate_levels(raw, ["rp", "superficie"], code_col=code_col)
    result = {}
    for lvl, df in levels.items():
        result[lvl] = pd.DataFrame(
            {
                "codgeo": df["codgeo"],
                "rp": _safe_numeric(df, ["rp"]).round(2),
                "superficie": _safe_numeric(df, ["superficie"]).round(2),
            }
        )
    return _add_year(result, year)


# ---------------------------------------------------------------------------
# BAAC - Route accidents (commune-level)
# ---------------------------------------------------------------------------

def _build_route_levels(year: int):
    """Construit les niveaux géographiques depuis les fichiers BAAC.

    Cherche d'abord le dataset national complet (baac/), puis le dataset
    Guyane-seulement (baac_guyane/). Retourne un tuple (all_levels, guyane_only)
    où guyane_only=True signifie que seules des données Guyane sont disponibles
    (les onglets FH/FRA seront à zéro, une note sera ajoutée dans l'Excel).
    """
    guyane_only = False
    source_sub = None

    for sub in ("baac", "baac_guyane"):
        baac_dir = INPUTS_DIR / sub / f"annees_{year}"
        if not baac_dir.exists():
            baac_dir = INPUTS_DIR / sub
        caract_path = baac_dir / f"caract_{year}.csv"
        usagers_path = baac_dir / f"usagers_{year}.csv"
        if caract_path.exists() and usagers_path.exists():
            source_sub = sub
            break
    if not caract_path.exists():
        raise FileNotFoundError(f"Source BAAC manquante: {caract_path}")
    if not usagers_path.exists():
        raise FileNotFoundError(f"Source BAAC manquante: {usagers_path}")

    if source_sub == "baac_guyane":
        guyane_only = True
        print(f"  [WARN] BAAC {year}: seule la source Guyane (baac_guyane/) est disponible."
              " Les niveaux FH/FRA ne contiendront que les données Guyane.")

    caract = _read_csv_auto(caract_path, dtype={"com": str, "dep": str})
    usagers = _read_csv_auto(usagers_path, dtype={"grav": str})

    # Normalize accident ID column (2022 uses Accident_Id, others use Num_Acc)
    acc_col_c = "Num_Acc" if "Num_Acc" in caract.columns else "Accident_Id"
    acc_col_u = "Num_Acc" if "Num_Acc" in usagers.columns else "Accident_Id"
    caract = caract.rename(columns={acc_col_c: "acc_id"})
    usagers = usagers.rename(columns={acc_col_u: "acc_id"})

    # Keep only needed columns
    caract = caract[["acc_id", "com"]].drop_duplicates(subset=["acc_id"])
    usagers = usagers[["acc_id", "grav"]].copy()
    usagers["grav"] = usagers["grav"].astype(str).str.strip()

    # Join
    merged = usagers.merge(caract, on="acc_id", how="inner")
    merged["commune_code"] = merged["com"].apply(_extract_commune_code)

    # Compute per commune
    nb_acci = merged.groupby("commune_code")["acc_id"].nunique().reset_index()
    nb_acci.columns = ["commune_code", "nb_acci"]

    nb_morts = merged[merged["grav"] == "2"].groupby("commune_code")["acc_id"].count().reset_index()
    nb_morts.columns = ["commune_code", "nb_morts"]

    nb_blesses = merged[merged["grav"].isin(["3", "4"])].groupby("commune_code")["acc_id"].count().reset_index()
    nb_blesses.columns = ["commune_code", "nb_blesses"]

    commune_df = nb_acci.merge(nb_blesses, on="commune_code", how="left").merge(nb_morts, on="commune_code", how="left")
    commune_df = commune_df.fillna(0)

    levels = _aggregate_levels(commune_df, ["nb_acci", "nb_blesses", "nb_morts"], code_col="commune_code")
    result = {}
    for lvl, df in levels.items():
        result[lvl] = pd.DataFrame({
            "codgeo": df["codgeo"],
            "nb_acci": _safe_numeric(df, ["nb_acci"]).round(0).astype(int),
            "nb_blesses": _safe_numeric(df, ["nb_blesses"]).round(0).astype(int),
            "nb_morts": _safe_numeric(df, ["nb_morts"]).round(0).astype(int),
        })
    return _add_year(result, year), guyane_only


# ---------------------------------------------------------------------------
# CepiDc - Mortality by cause (regional-level only)
# ---------------------------------------------------------------------------

REGION_NAME_TO_CODE = {
    "Auvergne-Rhône-Alpes": "84",
    "Bourgogne-Franche-Comté": "27",
    "Bretagne": "53",
    "Centre-Val de Loire": "24",
    "Corse": "94",
    "Grand Est": "44",
    "Hauts-de-France": "32",
    "Île-de-France": "11",
    "Normandie": "28",
    "Nouvelle-Aquitaine": "75",
    "Occitanie": "76",
    "Pays de la Loire": "52",
    "Provence-Alpes-Côte d'Azur": "93",
    "Guadeloupe": "01",
    "Martinique": "02",
    "Guyane": "03",
    "La Réunion": "04",
    "Mayotte": "06",
}


def _normalize_region_name(name: str) -> str:
    """Normalize region name for matching (handles encoding issues)."""
    import unicodedata
    s = str(name).strip()
    # Handle common encoding artifacts
    s = s.replace("Ã©", "é").replace("Ã´", "ô").replace("Ã®", "î")
    s = s.replace("Ã ", "à").replace("Ã¨", "è").replace("\u00e9", "é")
    # Remove trailing whitespace and "France" line
    if s.lower().startswith("france"):
        return "__france__"
    if "(p)" in s.lower():
        return "__skip__"
    return s


def _build_cepidc_levels(theme: str, year: int):
    cfg = THEME_CONFIGS[theme]
    sheet_name = cfg["cepidc_sheet"]
    var_n = cfg["variables"][0]  # e.g. nb_deces_toutes_causes
    var_tx = cfg["variables"][1]  # e.g. tx_mortalite_toutes_causes

    cepidc_dir = INPUTS_DIR / "cepidc"
    source = cepidc_dir / "taux_effectifs_regions_15_23.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Source CepiDc manquante: {source}")

    raw = pd.read_excel(source, sheet_name=None, header=None)

    # Find the right sheet (handle encoding in sheet names)
    target_sheet = None
    for sname in raw:
        clean = sname.replace("\ufffd", "").strip()
        if sheet_name.lower() in clean.lower() or clean.lower() in sheet_name.lower():
            target_sheet = sname
            break
    if target_sheet is None:
        raise ValueError(f"Feuille '{sheet_name}' introuvable dans CepiDc. Feuilles: {list(raw.keys())}")

    df = raw[target_sheet]

    # Parse multi-row header: row 1 = years, row 2 = N/Taux
    year_row = df.iloc[1].tolist()
    type_row = df.iloc[2].tolist()

    # Find column indices for requested year
    n_col = None
    tx_col = None
    for i in range(1, len(year_row)):
        yr_val = year_row[i]
        if pd.notna(yr_val):
            try:
                current_yr = int(float(yr_val))
            except (ValueError, TypeError):
                continue
        if current_yr == year:
            type_val = str(type_row[i]).strip().lower() if pd.notna(type_row[i]) else ""
            if type_val.startswith("n") and n_col is None:
                n_col = i
            elif "taux" in type_val and tx_col is None:
                tx_col = i

    if n_col is None:
        raise ValueError(f"Annee {year} introuvable dans CepiDc sheet '{sheet_name}'")

    # Extract data rows (skip header rows 0-2 and footer)
    regions_data = []
    france_n = 0
    france_tx = 0
    for row_idx in range(3, len(df)):
        region_name = _normalize_region_name(df.iloc[row_idx, 0])
        if region_name == "__skip__" or pd.isna(df.iloc[row_idx, 0]):
            continue

        n_val = pd.to_numeric(df.iloc[row_idx, n_col], errors="coerce")
        tx_val = pd.to_numeric(df.iloc[row_idx, tx_col], errors="coerce") if tx_col is not None else 0
        n_val = 0 if pd.isna(n_val) else n_val
        tx_val = 0 if pd.isna(tx_val) else tx_val

        if region_name == "__france__":
            france_n = n_val
            france_tx = tx_val
            continue

        # Match to region code
        reg_code = None
        for ref_name, ref_code in REGION_NAME_TO_CODE.items():
            if ref_name.lower() in region_name.lower() or region_name.lower() in ref_name.lower():
                reg_code = ref_code
                break
        if reg_code:
            regions_data.append({"codgeo": reg_code, var_n: n_val, var_tx: tx_val})

    if not regions_data:
        raise ValueError(f"Aucune donnee regionale CepiDc pour {year}")

    reg_df = pd.DataFrame(regions_data)
    # Ensure all 18 regions present
    reg_full = pd.DataFrame({"codgeo": REGION_ORDER})
    reg = reg_full.merge(reg_df, on="codgeo", how="left").fillna(0)

    # DOM / FH / FRA aggregation — counts are summed, rates use source data when available
    dom_n = reg[reg["codgeo"].isin(DOM_CODES)][var_n].sum()
    fh_n = reg[reg["codgeo"].isin(FH_REGIONS)][var_n].sum()

    # Rates: unweighted mean is statistically incorrect (regions have different populations).
    # Use France-level source values when available; otherwise mark as NaN rather than fake it.
    if france_tx:
        # Source provides a France-level rate — use it for FRA directly
        fra_tx_val = round(france_tx, 1)
        # For DOM/FH sub-aggregates, we don't have source values and no population data
        # to weight properly, so mark as NaN (approximate) with a warning
        dom_tx_val = float("nan")
        fh_tx_val = float("nan")
        print(f"  [WARN_DATA] CepiDc: taux DOM/FH non disponibles dans la source "
              f"(moyenne non ponderee exclue). Taux France utilise depuis la source.")
    else:
        # No France-level rate in source — cannot compute reliably
        dom_tx_val = float("nan")
        fh_tx_val = float("nan")
        fra_tx_val = float("nan")
        print(f"  [WARN_DATA] CepiDc: aucun taux France disponible dans la source. "
              f"Les taux agreges DOM/FH/FRA sont laisses vides (pas de population pour ponderer).")

    dom = pd.DataFrame({"codgeo": ["DOM"], var_n: [dom_n], var_tx: [dom_tx_val]})
    fh = pd.DataFrame({"codgeo": ["0"], var_n: [fh_n], var_tx: [fh_tx_val]})
    fra = pd.DataFrame({"codgeo": ["99"], var_n: [france_n if france_n else dom_n + fh_n], var_tx: [fra_tx_val]})

    # Communes Guyane: CepiDc data is regional only
    # - effectifs (nb_deces): not available at commune level -> leave empty (NaN)
    # - taux: use Guyane regional rate as proxy for all communes
    guyane_row = reg[reg["codgeo"] == "03"]
    guyane_tx = float(guyane_row[var_tx].iloc[0]) if len(guyane_row) > 0 else 0
    com = pd.DataFrame({
        "codgeo": COMMUNES_GUYANE,
        var_n: [None] * len(COMMUNES_GUYANE),  # Not available at commune level
        var_tx: [guyane_tx] * len(COMMUNES_GUYANE),  # Regional rate as proxy
    })

    all_levels = {"com": com, "reg": reg, "dom": dom, "fh": fh, "fra": fra}
    return _add_year(all_levels, year)


# ---------------------------------------------------------------------------
# Odissé (Santé Publique France) - Suicide / Alcool / Tabac
# ---------------------------------------------------------------------------

ODISSE_DIR = INPUTS_DIR / "cepidc" / "mortalite_causes_comportementales"


def _read_odisse(path: Path) -> pd.DataFrame:
    """Odissé/MOCA-O CSVs : encoding + séparateur auto, NA normalisés, BOM strip."""
    df, meta = read_csv_safe(path)
    log_read(meta)
    return df


def _find_col(df: pd.DataFrame, *needles) -> str:
    """Return the first column whose name contains all needles (case-insensitive, tolerant of encoding)."""
    for col in df.columns:
        low = col.lower()
        if all(n.lower() in low for n in needles):
            return col
    raise KeyError(f"Colonne contenant {needles} introuvable dans {list(df.columns)[:8]}...")


def _build_odisse_suicide_levels(year: int):
    path = ODISSE_DIR / "suicides_deces_departement.csv"
    if not path.exists():
        raise FileNotFoundError(f"Source Odissé manquante: {path}")
    df = _read_odisse(path)
    yr_col = _find_col(df, "nn")  # Année
    dep_col = _find_col(df, "partement", "code")  # Département Code
    n_col = _find_col(df, "nombre")  # Nombre de décès

    df = df[df[yr_col] == year].copy()
    if df.empty:
        raise ValueError(f"Aucune donnée suicide pour {year}")

    df["dep_raw"] = df[dep_col].astype(str).str.strip()
    # Normaliser dep: 1->01, 2A/2B gardés, 971-976 gardés (3 chars)
    def norm_dep(v: str) -> str:
        v = v.strip()
        if v in ("2A", "2B"):
            return v
        if v.isdigit():
            if len(v) <= 2:
                return v.zfill(2)
            return v
        return v
    df["dep"] = df["dep_raw"].apply(norm_dep)
    df["reg"] = df["dep"].map(DEP_TO_REG)
    df["m_suicide"] = pd.to_numeric(df[n_col], errors="coerce").fillna(0)

    # DEP aggregation (all ages/sexes) - used for commune proxy in Guyane (973)
    dep_agg = df.groupby("dep", as_index=False)["m_suicide"].sum()

    # REG aggregation
    reg_agg = df.dropna(subset=["reg"]).groupby("reg", as_index=False)["m_suicide"].sum()
    reg_full = pd.DataFrame({"codgeo": REGION_ORDER})
    reg = reg_full.merge(reg_agg.rename(columns={"reg": "codgeo"}), on="codgeo", how="left").fillna(0)

    # Commune Guyane : proxy = valeur DEP 973
    dep973 = dep_agg[dep_agg["dep"] == "973"]["m_suicide"].sum()
    com = pd.DataFrame({
        "codgeo": COMMUNES_GUYANE,
        "m_suicide": [None] * len(COMMUNES_GUYANE),  # Pas de données communales
    })

    dom_n = reg[reg["codgeo"].isin(DOM_CODES)]["m_suicide"].sum()
    fh_n = reg[reg["codgeo"].isin(FH_REGIONS)]["m_suicide"].sum()
    fra_n = reg["m_suicide"].sum()

    dom = pd.DataFrame({"codgeo": ["DOM"], "m_suicide": [dom_n]})
    fh = pd.DataFrame({"codgeo": ["0"], "m_suicide": [fh_n]})
    fra = pd.DataFrame({"codgeo": ["99"], "m_suicide": [fra_n]})

    return _add_year({"com": com, "reg": reg, "dom": dom, "fh": fh, "fra": fra}, year)


def _build_odisse_consommation_levels(year: int, kind: str):
    """kind: 'alcool' or 'tabac'. REG-level only (consommation quotidienne)."""
    if kind == "alcool":
        path = ODISSE_DIR / "alcool_consommation_quotidienne_region.csv"
        out_var = "prev_alcool_quotidien"
    else:
        path = ODISSE_DIR / "tabac_consommation_quotidienne_region.csv"
        out_var = "prev_tabac_quotidien"
    if not path.exists():
        raise FileNotFoundError(f"Source Odissé manquante: {path}")
    df = _read_odisse(path)
    yr_col = _find_col(df, "nn")
    reg_col = _find_col(df, "gion", "code")
    tx_col = _find_col(df, "taux", "standardis")
    sexe_col = _find_col(df, "exe")

    df = df[df[yr_col] == year].copy()
    if df.empty:
        df_all, _ = read_csv_safe(path)
        available = sorted(df_all[yr_col].dropna().unique().tolist())
        raise ValueError(f"Aucune donnée {kind} pour {year}. Années dispo: {available}")

    # Prendre "Hommes et Femmes" en priorité, sinon moyenne H+F
    if df[sexe_col].astype(str).str.contains("Hommes et Femmes", case=False, na=False).any():
        df = df[df[sexe_col].astype(str).str.contains("Hommes et Femmes", case=False, na=False)]
    def _norm_reg(v):
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if s.isdigit() and len(s) <= 2:
            return s.zfill(2)
        return s
    df["codgeo"] = df[reg_col].apply(_norm_reg)
    df[out_var] = pd.to_numeric(df[tx_col], errors="coerce").fillna(0)

    reg_agg = df.groupby("codgeo", as_index=False)[out_var].mean()
    reg_full = pd.DataFrame({"codgeo": REGION_ORDER})
    reg = reg_full.merge(reg_agg, on="codgeo", how="left").fillna(0)
    reg[out_var] = reg[out_var].round(2)

    # Guyane = reg 03
    gy_tx = float(reg[reg["codgeo"] == "03"][out_var].iloc[0]) if (reg["codgeo"] == "03").any() else 0
    com = pd.DataFrame({
        "codgeo": COMMUNES_GUYANE,
        out_var: [gy_tx] * len(COMMUNES_GUYANE),
    })

    # Rates (prevalence): unweighted mean across regions is statistically incorrect
    # (regions have very different population sizes). Without population data to weight,
    # we leave aggregated rates as NaN rather than produce misleading values.
    dom_regions_with_data = (reg[reg["codgeo"].isin(DOM_CODES)][out_var] > 0).sum()
    fh_regions_with_data = (reg[reg["codgeo"].isin(FH_REGIONS)][out_var] > 0).sum()

    if dom_regions_with_data >= len(DOM_CODES) * 0.5:
        dom_tx = round(reg[reg["codgeo"].isin(DOM_CODES)][out_var].mean(), 2)
    else:
        dom_tx = float("nan")
    if fh_regions_with_data >= len(FH_REGIONS) * 0.5:
        fh_tx = round(reg[reg["codgeo"].isin(FH_REGIONS)][out_var].mean(), 2)
    else:
        fh_tx = float("nan")

    # FRA: only compute if both DOM and FH have sufficient coverage
    if dom_regions_with_data >= len(DOM_CODES) * 0.5 and fh_regions_with_data >= len(FH_REGIONS) * 0.5:
        fra_tx = round(reg[out_var].mean(), 2)
    else:
        fra_tx = float("nan")
        print(f"  [WARN_DATA] Odisse {kind}: couverture regionale insuffisante pour agreger DOM/FH/FRA. "
              f"Taux agreges laisses vides (moyenne non ponderee exclue).")

    dom = pd.DataFrame({"codgeo": ["DOM"], out_var: [dom_tx]})
    fh = pd.DataFrame({"codgeo": ["0"], out_var: [fh_tx]})
    fra = pd.DataFrame({"codgeo": ["99"], out_var: [fra_tx]})

    return _add_year({"com": com, "reg": reg, "dom": dom, "fh": fh, "fra": fra}, year)


def _load_region_population(year: int) -> dict[str, float]:
    """Population totale (PTOT) par code région, depuis populations_YYYY.csv.

    Pas de fichier populations pour l'année demandée (ex. 2024/2025) -> on retombe
    sur le fichier le plus récent disponible (la population régionale varie peu,
    acceptable pour un taux d'incidence). Retourne {code_reg(2 chiffres): pop}.
    """
    candidates = [INPUTS_DIR / f"populations_{year}.csv"]
    candidates += sorted(INPUTS_DIR.glob("populations_*.csv"), reverse=True)
    src = next((p for p in candidates if p.exists()), None)
    if src is None:
        return {}
    raw = _read_csv_auto(src)
    reg_col = "REG" if "REG" in raw.columns else None
    pop_col = next((c for c in ("PTOT", "PMUN") if c in raw.columns), None)
    if not reg_col or not pop_col:
        return {}
    raw["_reg"] = raw[reg_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True).str.zfill(2)
    raw["_pop"] = pd.to_numeric(raw[pop_col], errors="coerce").fillna(0)
    return raw.groupby("_reg")["_pop"].sum().to_dict()


def _apply_noyades_rates(levels: dict, year: int) -> dict:
    """Ajoute tx_noyades / tx_noyades_deces (taux d'incidence pour 1 000 000 hab.).

    tx = nb / population x 1 000 000. Niveaux : reg via pop régionale ; dom/fh/fra
    via pop agrégée des régions correspondantes ; com laissé vide (pas de pop
    communale exploitable pour la ventilation Guyane).
    """
    reg_pop = _load_region_population(year)

    def rate(nb, pop):
        if pop and pd.notna(nb):
            return round(float(nb) / pop * 1_000_000, 1)
        return None

    # Région
    reg = levels["reg"]
    reg["tx_noyades"] = [rate(n, reg_pop.get(str(c))) for c, n in zip(reg["codgeo"], reg["nb_noyades"])]
    reg["tx_noyades_deces"] = [rate(n, reg_pop.get(str(c))) for c, n in zip(reg["codgeo"], reg["nb_noyades_deces"])]

    pop_dom = sum(reg_pop.get(c, 0) for c in DOM_CODES)
    pop_fh = sum(reg_pop.get(c, 0) for c in FH_REGIONS)
    pop_fra = pop_dom + pop_fh

    for key, pop in (("dom", pop_dom), ("fh", pop_fh), ("fra", pop_fra)):
        df = levels[key]
        df["tx_noyades"] = [rate(n, pop) for n in df["nb_noyades"]]
        df["tx_noyades_deces"] = [rate(n, pop) for n in df["nb_noyades_deces"]]

    com = levels["com"]
    com["tx_noyades"] = [None] * len(com)
    com["tx_noyades_deces"] = [None] * len(com)
    return levels


def _build_spf_noyades_levels(year: int):
    # Priorité : fichier fusionné 2003-2024 (inclut 2023-2024 SPF/Snosan extrait des PDF).
    # Fallback : ancien CSV 2003-2021 seul, puis concaténation de tous les
    # noyades_departement_*.csv présents dans le dossier.
    noyades_dir = INPUTS_DIR / "spf_noyades"
    full_path = noyades_dir / "noyades_departement_2003_2024.csv"
    legacy_path = noyades_dir / "noyades_departement_2003_2021.csv"
    if full_path.exists():
        df, meta = read_csv_safe(full_path)
        log_read(meta)
    elif legacy_path.exists():
        # Agrège tous les CSV du dossier (legacy + extensions éventuelles).
        frames = []
        for p in sorted(noyades_dir.glob("noyades_departement_*.csv")):
            sub_df, sub_meta = read_csv_safe(p)
            log_read(sub_meta)
            frames.append(sub_df)
        if not frames:
            raise FileNotFoundError(f"Source noyades manquante dans {noyades_dir}")
        df = pd.concat(frames, ignore_index=True)
    else:
        raise FileNotFoundError(f"Source noyades manquante: {full_path} ou {legacy_path}")
    yr_col = _find_col(df, "nn")  # Année
    dep_col = _find_col(df, "partement", "code")
    reg_col = _find_col(df, "gion", "code")
    n_col = _find_col(df, "noyades", "accidentelles")
    # Colonne exacte "Nombre de noyades accidentelles" (pas "suivies de décès")
    for c in df.columns:
        if "noyades accidentelles" in c.lower() and "d" not in c.lower()[-10:]:
            n_col = c
            break
    d_col = None
    for c in df.columns:
        cl = c.lower()
        if "suivies" in cl and ("décès" in cl or "deces" in cl):
            d_col = c
            break

    df_year = df[df[yr_col] == year].copy()
    if df_year.empty:
        available = sorted(df[yr_col].dropna().unique().tolist())
        raise ValueError(f"Aucune donnée noyades pour {year}. Années dispo (enquête triennale): {available}")

    df_year["dep"] = df_year[dep_col].astype(str).str.strip()
    # Région Code parfois float (93.0) -> int puis str, puis zfill(2)
    def _norm_reg(v):
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if s.isdigit() and len(s) <= 2:
            return s.zfill(2)
        return s
    df_year["reg"] = df_year[reg_col].apply(_norm_reg)
    # -9 = code SPF "donnée non disponible / secret statistique" : neutralisé en
    # NaN avant agrégation pour ne jamais polluer les sommes reg/dom/fh/fra.
    df_year["nb_noyades"] = pd.to_numeric(df_year[n_col], errors="coerce").replace(-9, pd.NA).fillna(0)
    df_year["nb_noyades_deces"] = pd.to_numeric(df_year[d_col], errors="coerce").replace(-9, pd.NA).fillna(0) if d_col else 0

    reg_agg = df_year.groupby("reg", as_index=False)[["nb_noyades", "nb_noyades_deces"]].sum()
    reg_full = pd.DataFrame({"codgeo": REGION_ORDER})
    reg = reg_full.merge(reg_agg.rename(columns={"reg": "codgeo"}), on="codgeo", how="left").fillna(0)

    # Commune Guyane: pas de ventilation communale
    com = pd.DataFrame({
        "codgeo": COMMUNES_GUYANE,
        "nb_noyades": [None] * len(COMMUNES_GUYANE),
        "nb_noyades_deces": [None] * len(COMMUNES_GUYANE),
    })

    dom = pd.DataFrame({
        "codgeo": ["DOM"],
        "nb_noyades": [reg[reg["codgeo"].isin(DOM_CODES)]["nb_noyades"].sum()],
        "nb_noyades_deces": [reg[reg["codgeo"].isin(DOM_CODES)]["nb_noyades_deces"].sum()],
    })
    fh = pd.DataFrame({
        "codgeo": ["0"],
        "nb_noyades": [reg[reg["codgeo"].isin(FH_REGIONS)]["nb_noyades"].sum()],
        "nb_noyades_deces": [reg[reg["codgeo"].isin(FH_REGIONS)]["nb_noyades_deces"].sum()],
    })
    fra = pd.DataFrame({
        "codgeo": ["99"],
        "nb_noyades": [reg["nb_noyades"].sum()],
        "nb_noyades_deces": [reg["nb_noyades_deces"].sum()],
    })

    levels = _apply_noyades_rates({"com": com, "reg": reg, "dom": dom, "fh": fh, "fra": fra}, year)
    return _add_year(levels, year)


def _add_note_to_sheet(ws, df, note_text, note_color="FF0000"):
    """Ajoute une note informative sous les données dans une feuille Excel."""
    note_row = len(df) + 3
    note_cell = ws.cell(row=note_row, column=1, value=note_text)
    note_cell.font = Font(italic=True, color=note_color)


# ---------------------------------------------------------------------------
# DREES - Offre d'accueil du jeune enfant (EAJE)
# ---------------------------------------------------------------------------

EAJE_VARS = [
    "nb_places_acceuil_co", "nb_places_mono", "nb_places_hg", "nb_places_je",
    "nb_places_multi", "nb_places_familial",
    "nb_etab_co", "nb_etab_mono", "nb_etab_hg", "nb_etab_je", "nb_etab_multi",
    "nb_services_familial", "nb_mam",
]


def _eaje_dep_code(value) -> str:
    """Normalise un code département DREES (str). '973' -> '973', 1 -> '01'."""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.startswith("97") or s.startswith("98"):
        return s.zfill(3)
    return s.zfill(2)


def _is_valid_dep_code(value) -> bool:
    """True si col[0]/col[1] ressemble à un code département (chiffre, 2-3 car.)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if not s or s[0].upper() not in "0123456789":
        return False
    return 2 <= len(s) <= 3


def _build_eaje_levels_2021(year: int):
    """Parser format 2021 (sheets T*). Colonnes mappées selon le plan validé."""
    drees_dir = INPUTS_DIR / "drees"
    source = drees_dir / f"drees_offre_accueil_jeune_enfant_{year}_series_longues.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Source DREES EAJE manquante: {source}")

    raw = pd.read_excel(source, sheet_name=None, header=None)

    def _find_sheet(*needles):
        for sname in raw:
            low = sname.lower()
            if all(n.lower() in low for n in needles):
                return sname
        raise ValueError(f"Feuille EAJE 2021 introuvable pour {needles}. Feuilles: {list(raw.keys())}")

    t1 = raw[_find_sheet("t1", "total")]      # etab accueil collectif
    t7 = raw[_find_sheet("t7", "total")]      # places accueil collectif
    t11 = raw[_find_sheet("t11", "places")]   # places accueil familial
    t5 = raw[_find_sheet("t5", "familial")]   # services accueil familial
    t6 = raw[_find_sheet("t6", "mam")]        # MAM

    def _index_by_dep(df, cols):
        """Retourne {dep_code: {var: value}} pour les lignes data valides."""
        out = {}
        for i in range(len(df)):
            c0 = df.iloc[i, 0]
            if not _is_valid_dep_code(c0):
                continue
            dep = _eaje_dep_code(c0)
            rec = {}
            for var, col in cols.items():
                v = pd.to_numeric(df.iloc[i, col], errors="coerce")
                rec[var] = None if pd.isna(v) else float(v)
            out[dep] = rec
        return out

    places = _index_by_dep(t7, {
        "nb_places_mono": 2, "nb_places_hg": 3, "nb_places_je": 5,
        "nb_places_multi": 7, "nb_places_acceuil_co": 8,
    })
    places_fam = _index_by_dep(t11, {"nb_places_familial": 6})
    etab = _index_by_dep(t1, {
        "nb_etab_mono": 2, "nb_etab_hg": 3, "nb_etab_je": 5,
        "nb_etab_multi": 7, "nb_etab_co": 8,
    })
    services_fam = _index_by_dep(t5, {"nb_services_familial": 2})
    mam = _index_by_dep(t6, {"nb_mam": 2})

    deps = set(places) | set(etab)
    rows = []
    for dep in sorted(deps):
        rec = {"dep_code": dep}
        for src in (places, places_fam, etab, services_fam, mam):
            rec.update(src.get(dep, {}))
        rows.append(rec)

    dep_df = pd.DataFrame(rows)
    for var in EAJE_VARS:
        if var not in dep_df.columns:
            dep_df[var] = None
    return dep_df


def _build_eaje_levels_2022plus(year: int):
    """Parser format 2022+ (sheets Tab*-PMI)."""
    drees_dir = INPUTS_DIR / "drees"
    source = drees_dir / f"drees_offre_accueil_jeune_enfant_{year}_series_longues.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Source DREES EAJE manquante: {source}")

    raw = pd.read_excel(source, sheet_name=None, header=None)
    tab1 = raw["Tab1-PMI"]    # etab
    tab17 = raw["Tab17-PMI"]  # places
    tab34 = raw["Tab34-PMI"]  # MAM series longues

    def _read_dep_block(df, mapping):
        """Lit les lignes département (col[1]=code dep) jusqu'au 1er bloc TOTAL/région.

        Le bloc départemental se termine dès qu'on rencontre une ligne dont col[0]
        commence par 'TOTAL' ou dont col[1] n'est pas un code département valide.
        """
        out = {}
        started = False
        for i in range(5, len(df)):
            c0 = df.iloc[i, 0]
            c1 = df.iloc[i, 1]
            c0s = "" if c0 is None or (isinstance(c0, float) and pd.isna(c0)) else str(c0).strip()
            if c0s.upper().startswith("TOTAL"):
                if started:
                    break
                continue
            if not _is_valid_dep_code(c1):
                if started:
                    break
                continue
            started = True
            dep = _eaje_dep_code(c1)
            rec = {}
            for var, col in mapping.items():
                v = pd.to_numeric(df.iloc[i, col], errors="coerce")
                rec[var] = None if pd.isna(v) else float(v)
            out[dep] = rec
        return out

    # Tab1-PMI / Tab17-PMI : col[3]=crèches collectives, col[4]=crèches familiales,
    # col[5]=multi accueil, col[6]=jardins d'enfants, col[8]=total
    etab = _read_dep_block(tab1, {
        "nb_etab_mono": 3, "nb_services_familial": 4, "nb_etab_multi": 5,
        "nb_etab_je": 6, "nb_etab_co": 8,
    })
    places = _read_dep_block(tab17, {
        "nb_places_mono": 3, "nb_places_familial": 4, "nb_places_multi": 5,
        "nb_places_je": 6, "nb_places_acceuil_co": 8,
    })

    # MAM : Tab34-PMI, colonne de l'année demandée (en-tête ligne 5 / index 4)
    header34 = tab34.iloc[4].tolist()
    mam_col = None
    for j in range(3, len(header34)):
        hv = header34[j]
        if pd.notna(hv):
            try:
                if int(float(hv)) == year:
                    mam_col = j
                    break
            except (ValueError, TypeError):
                continue
    mam = {}
    if mam_col is not None:
        for i in range(5, len(tab34)):
            c1 = tab34.iloc[i, 1]
            c0 = tab34.iloc[i, 0]
            c0s = "" if c0 is None or (isinstance(c0, float) and pd.isna(c0)) else str(c0).strip()
            if c0s.upper().startswith("TOTAL"):
                if mam:
                    break
                continue
            if not _is_valid_dep_code(c1):
                if mam:
                    break
                continue
            v = pd.to_numeric(tab34.iloc[i, mam_col], errors="coerce")
            mam[_eaje_dep_code(c1)] = None if pd.isna(v) else float(v)
    else:
        print(f"  [WARN_DATA] DREES EAJE {year}: colonne MAM (annee {year}) introuvable dans Tab34-PMI. nb_mam laisse vide.")

    deps = set(places) | set(etab)
    rows = []
    for dep in sorted(deps):
        rec = {"dep_code": dep}
        for src in (places, etab):
            rec.update(src.get(dep, {}))
        if dep in mam:
            rec["nb_mam"] = mam[dep]
        # Halte-garderies fusionnées dans crèches collectives dès 2022 -> non disponibles
        rec["nb_places_hg"] = 0.0
        rec["nb_etab_hg"] = 0.0
        rows.append(rec)

    print(f"  [WARN_DATA] DREES EAJE {year}: rupture de serie (taxonomie 2022+). "
          "Les haltes-garderies sont fusionnees dans les creches collectives (nb_places_hg / nb_etab_hg = 0). "
          "Les creches familiales servent de proxy pour nb_services_familial.")

    dep_df = pd.DataFrame(rows)
    for var in EAJE_VARS:
        if var not in dep_df.columns:
            dep_df[var] = None
    return dep_df


def _build_eaje_levels(year: int):
    """Construit les niveaux geo (com/reg/dom/fh/fra) pour l'offre d'accueil DREES.

    Suit le meme schema que _build_cepidc_levels : agregation manuelle dep -> reg ->
    dom/fh/fra (sans codes commune). Le niveau commune est vide (donnees non publiees).
    """
    if year <= 2021:
        dep_df = _build_eaje_levels_2021(year)
    else:
        dep_df = _build_eaje_levels_2022plus(year)

    for var in EAJE_VARS:
        dep_df[var] = pd.to_numeric(dep_df[var], errors="coerce")

    # reg : somme par region via DEP_TO_REG, complétée sur les 18 régions.
    # Quelques codes DREES 2022+ ne sont pas dans DEP_TO_REG : Corse '20' (région 94),
    # Rhône scindé '69D'/'69M' (-> '69'). On les ramène à un code connu pour ne pas
    # perdre ces départements dans les agrégats FH/FRA.
    EAJE_DEP_ALIAS = {"20": "2A", "69D": "69", "69M": "69"}
    dep_df = dep_df.copy()
    dep_df["reg"] = dep_df["dep_code"].apply(
        lambda d: DEP_TO_REG.get(EAJE_DEP_ALIAS.get(d, d))
    )
    reg_grouped = dep_df.dropna(subset=["reg"]).groupby("reg")[EAJE_VARS].sum(min_count=1).reset_index()
    reg_grouped.rename(columns={"reg": "codgeo"}, inplace=True)
    reg = pd.DataFrame({"codgeo": REGION_ORDER}).merge(reg_grouped, on="codgeo", how="left")

    # dom / fh / fra : sommes d'agrégats régionaux
    def _sum_rows(mask_codes, codgeo_label):
        sub = reg[reg["codgeo"].isin(mask_codes)]
        agg = sub[EAJE_VARS].sum(min_count=1).to_frame().T
        agg["codgeo"] = codgeo_label
        return agg[["codgeo"] + EAJE_VARS]

    dom = _sum_rows(DOM_CODES, "DOM")
    fh = _sum_rows(FH_REGIONS, "0")
    fra_agg = reg[EAJE_VARS].sum(min_count=1).to_frame().T
    fra_agg["codgeo"] = "99"
    fra = fra_agg[["codgeo"] + EAJE_VARS]

    # com : 22 communes Guyane, valeurs vides (DREES ne publie pas au niveau commune)
    com = pd.DataFrame({"codgeo": COMMUNES_GUYANE})
    for var in EAJE_VARS:
        com[var] = None

    all_levels = {"com": com, "reg": reg, "dom": dom, "fh": fh, "fra": fra}
    return _add_year(all_levels, year)


def _generate_excel_and_zip(theme: str, year: int, all_levels, guyane_only: bool = False):
    """Génère les fichiers Excel par niveau géo et crée le ZIP final.

    guyane_only=True : la source BAAC ne contient que la Guyane.
    Une note d'avertissement est ajoutée dans les onglets FH/FRA.
    """
    cfg = THEME_CONFIGS[theme]
    all_variables = cfg["variables"]
    # Exclure les variables calculées (tx_*) — PRISME les recalcule côté client,
    # SAUF si le thème demande explicitement de livrer ses taux (emit_calculated_rates).
    emit_rates = cfg.get("emit_calculated_rates", False)
    variables = []
    for v in all_variables:
        if is_calculated_variable(v) and not emit_rates:
            print(f"  [FILTER] Skipped calculated variable: {v}")
        else:
            variables.append(v)
    excel_name = cfg["excel_name"]

    # Note BAAC Guyane-only pour onglets FH et FRA
    BAAC_GUYANE_NOTE = (
        "ATTENTION : Les donnees BAAC disponibles pour cette annee ne couvrent que la Guyane (973). "
        "Les chiffres de France Hexagonale et France Entiere sont incomplets. "
        "Source : BAAC local Guyane (baac_guyane/) — le fichier national n'etait pas disponible lors de la generation."
    )

    root_dir = OUTPUT_DIR / f"{theme}_opendata" / str(year)
    if root_dir.exists():
        shutil.rmtree(root_dir)
    root_dir.mkdir(parents=True, exist_ok=True)

    for geo_key, folder_name in GEO_FOLDER_MAPPING.items():
        folder = root_dir / folder_name
        folder.mkdir(exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = geo_key

        headers = [geo_key, "annee"] + variables
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=idx, value=h)
            c.font = Font(bold=True)
            c.fill = ORANGE_FILL

        df = all_levels[geo_key].copy()
        for var in variables:
            if var not in df.columns:
                df[var] = None

        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=str(row.get("codgeo", "")) if pd.notna(row.get("codgeo")) else "")
            ws.cell(row=row_idx, column=2, value=int(row.get("annee", year)) if pd.notna(row.get("annee")) else year)
            for i, var in enumerate(variables):
                val = row.get(var)
                if pd.notna(val):
                    data_cell = ws.cell(row=row_idx, column=3 + i, value=float(val))
                else:
                    data_cell = ws.cell(row=row_idx, column=3 + i, value=None)
                data_cell.fill = ORANGE_FILL

        # CepiDc commune annotation: data is regional-only
        if geo_key == "com" and cfg.get("source_type") == "cepidc":
            _add_note_to_sheet(ws, df,
                "NOTE: La source CepiDc ne fournit pas de donnees communales. "
                "Les effectifs sont vides. Les taux correspondent au taux regional Guyane (proxy).")

        # BAAC Guyane-only annotation in FH and FRA sheets
        if guyane_only and geo_key in ("fh", "fra"):
            _add_note_to_sheet(ws, df, BAAC_GUYANE_NOTE)

        # CepiDc / Odisse rate annotation for DOM/FH/FRA: rates may be NaN (no population weighting)
        if geo_key in ("dom", "fh", "fra") and cfg.get("source_type") in ("cepidc", "odisse_alcool", "odisse_tabac"):
            _add_note_to_sheet(ws, df,
                "NOTE: Les taux agreges (DOM/FH/FRA) ne sont pas disponibles dans la source. "
                "Une moyenne non ponderee serait statistiquement incorrecte (populations regionales differentes). "
                "Les cellules de taux vides correspondent a cette limitation.",
                note_color="0000FF")

        wb.save(folder / f"{excel_name}.xlsx")

    wb_cons = Workbook()
    wb_cons.remove(wb_cons.active)
    for geo_key in GEO_FOLDER_MAPPING:
        ws = wb_cons.create_sheet(geo_key)
        headers = [geo_key, "annee"] + variables
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=idx, value=h)
            c.font = Font(bold=True)
            c.fill = ORANGE_FILL
        df = all_levels[geo_key].copy()
        for var in variables:
            if var not in df.columns:
                df[var] = None
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=str(row.get("codgeo", "")) if pd.notna(row.get("codgeo")) else "")
            ws.cell(row=row_idx, column=2, value=int(row.get("annee", year)) if pd.notna(row.get("annee")) else year)
            for i, var in enumerate(variables):
                val = row.get(var)
                if pd.notna(val):
                    data_cell = ws.cell(row=row_idx, column=3 + i, value=float(val))
                else:
                    data_cell = ws.cell(row=row_idx, column=3 + i, value=None)
                data_cell.fill = ORANGE_FILL

        # CepiDc commune annotation in consolidated
        if geo_key == "com" and cfg.get("source_type") == "cepidc":
            _add_note_to_sheet(ws, df,
                "NOTE: La source CepiDc ne fournit pas de donnees communales. "
                "Les effectifs sont vides. Les taux correspondent au taux regional Guyane (proxy).")

        # BAAC Guyane-only annotation in FH and FRA sheets of consolidated
        if guyane_only and geo_key in ("fh", "fra"):
            _add_note_to_sheet(ws, df, BAAC_GUYANE_NOTE)

        # CepiDc / Odisse rate annotation for DOM/FH/FRA in consolidated
        if geo_key in ("dom", "fh", "fra") and cfg.get("source_type") in ("cepidc", "odisse_alcool", "odisse_tabac"):
            _add_note_to_sheet(ws, df,
                "NOTE: Les taux agreges (DOM/FH/FRA) ne sont pas disponibles dans la source. "
                "Une moyenne non ponderee serait statistiquement incorrecte (populations regionales differentes). "
                "Les cellules de taux vides correspondent a cette limitation.",
                note_color="0000FF")

    wb_cons.save(root_dir / f"{excel_name}_consolidated_{year}.xlsx")
    zip_path = shutil.make_archive(str(OUTPUT_DIR / f"{theme}_opendata_{year}"), "zip", str(OUTPUT_DIR / f"{theme}_opendata"), str(year))
    return root_dir, Path(zip_path)


def generate_theme(theme: str, year: int):
    source_type = THEME_CONFIGS[theme]["source_type"]
    guyane_only = False  # Positionné à True uniquement pour BAAC Guyane-seulement

    if source_type == "educ":
        all_levels = _build_educ_levels(year)
    elif source_type == "couples":
        all_levels = _build_couples_levels(theme, year)
    elif source_type == "caf":
        all_levels = _build_alloc_levels(year)
    elif source_type == "ircom":
        all_levels = _build_revenu_levels(year)
    elif source_type == "pop_legales":
        all_levels = _build_densite_levels(year)
    elif source_type == "baac":
        all_levels, guyane_only = _build_route_levels(year)
    elif source_type == "cepidc":
        all_levels = _build_cepidc_levels(theme, year)
    elif source_type == "odisse_suicide":
        all_levels = _build_odisse_suicide_levels(year)
    elif source_type == "odisse_alcool":
        all_levels = _build_odisse_consommation_levels(year, "alcool")
    elif source_type == "odisse_tabac":
        all_levels = _build_odisse_consommation_levels(year, "tabac")
    elif source_type == "spf_noyades":
        all_levels = _build_spf_noyades_levels(year)
    elif source_type == "drees_eaje":
        all_levels = _build_eaje_levels(year)
    else:
        raise ValueError(f"Source type inconnu: {source_type}")

    root_dir, zip_path = _generate_excel_and_zip(theme, year, all_levels, guyane_only=guyane_only)
    print(f"[OK] {theme} {year}: {zip_path}")
    print("     dossiers:", ", ".join(GEO_FOLDER_MAPPING.values()))
    return root_dir


def main():
    import sys
    if len(sys.argv) == 2 and sys.argv[1].isdigit():
        year = int(sys.argv[1])
        themes = list(THEME_CONFIGS.keys())
    else:
        parser = argparse.ArgumentParser(description="PRISME - Generation Open Data")
        parser.add_argument("--theme", default="all", help=f"Theme: {', '.join(THEME_CONFIGS.keys())} ou all")
        parser.add_argument("--year", type=int, default=2022, help="Annee (defaut: 2022)")
        args = parser.parse_args()
        year = args.year
        if args.theme == "all":
            themes = list(THEME_CONFIGS.keys())
        elif args.theme in THEME_CONFIGS:
            themes = [args.theme]
        else:
            raise SystemExit(f"Theme inconnu: {args.theme}")

    print("=" * 70)
    print(f"Generation Open Data - annee {year}")
    print("=" * 70)
    for theme in themes:
        try:
            generate_theme(theme, year)
        except Exception as exc:
            print(f"[ERROR] {theme}: {exc}")


if __name__ == "__main__":
    main()
