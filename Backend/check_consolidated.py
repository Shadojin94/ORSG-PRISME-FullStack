import openpyxl
from pathlib import Path

# Check consolidated file
cons_file = Path("output/educ_2020_test/Educ/2020/educ_consolidated_2020.xlsx")

if not cons_file.exists():
    print(f"ERROR: Consolidated file not found at {cons_file}")
    exit(1)

print(f"Loading: {cons_file}")
wb = openpyxl.load_workbook(cons_file)
print(f"\nSheets in consolidated file: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet '{sheet_name}': {ws.max_row} rows × {ws.max_column} cols")
    if ws.max_row > 0:
        row1 = [cell.value for cell in ws[1]]
        print(f"  Row 1 (headers): {row1}")
        if ws.max_row > 1:
            row2 = [cell.value for cell in ws[2]]
            print(f"  Row 2 (data): {row2}")
