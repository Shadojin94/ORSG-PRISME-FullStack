"""
generate_mocao_consolidated.py
--------------------------------
Genere un fichier .xlsx consolide au format MOCA-O natif multi-annees,
compatible avec le pipeline client (Fichier Moca O_Comportement_2018_2023.xls).

Structure de sortie (sheets) :
- COM       : toutes communes, toutes annees
- DROM      : DROM toutes annees
- franENT   : France entiere toutes annees
- FranHEX   : France metropolitaine toutes annees
- REG YYYY  : une feuille par annee pour les regions

Usage CLI :
    python generate_mocao_consolidated.py <dataset_id> <yearStart> <yearEnd> [--source moca|opendata]

Exemple :
    python generate_mocao_consolidated.py comp_mortalite 2018 2023
    python generate_mocao_consolidated.py route 2019 2024 --source opendata

Sortie :
    Backend/output/<dataset_id>_mocao_<yearStart>_<yearEnd>.xlsx
"""
import sys
import os
import argparse
import subprocess
import zipfile
import tempfile
import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

BASE = Path(__file__).parent
OUTPUT_DIR = BASE / "output"
PYTHON_EXE = os.environ.get("PYTHON_EXE", sys.executable)

# Mapping des noms de variables internes -> noms client MOCA-O attendus
# Si absent du dict, le nom interne est conserve
VARIABLE_RENAME = {
    # comp_mortalite (Comportements)
    "m_alcool": "alcool",
    "m_tabac": "tabac",
    "m_suicide": "suicide",
    # route (Traumatismes)
    "nb_acci": "nb_accidents",
    "tx_acci": "tx_accidents",
    "nb_blesses": "nb_blesses",
    "nb_morts": "nb_morts",
    "tx_morts": "tx_morts",
    # noyades
    "nb_noyades": "nb_noyades",
    "nb_noyades_deces": "nb_noyades_deces",
    "tx_noyades": "tx_noyades",
    "tx_noyades_deces": "tx_noyades_deces",
}

# Valeur par defaut pour donnees manquantes (-999.0 comme dans le modele client)
MISSING_VALUE = -999.0

# Territoire folder names -> sheet key
TERR_TO_SHEETKEY = {
    "Commune": "com",
    "Commune": "com",
    "Régions": "reg",
    "Région": "reg",
    "Regions": "reg",
    "Region": "reg",
    "DOM": "dom",
    "France entière": "fra",
    "France entiere": "fra",
    "France Hexagonale": "fh",
    "France_Hexagonale": "fh",
}


def _run(cmd, cwd=None, check=True):
    """Run subprocess, print stderr."""
    res = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
    if res.returncode != 0:
        print(f"[ERROR] {' '.join(cmd)}\n{res.stderr}", file=sys.stderr)
        if check:
            raise RuntimeError(f"Command failed: {cmd}")
    return res


def generate_year_zip(dataset_id: str, year: int, source: str = "moca") -> Path:
    """Generate or locate existing ZIP for a given year."""
    if source == "opendata":
        candidate = OUTPUT_DIR / f"{dataset_id}_opendata_{year}.zip"
        if candidate.exists():
            return candidate
        print(f"[GEN] opendata {dataset_id} {year}...", file=sys.stderr)
        code = (
            f"import sys; sys.path.insert(0, r'{BASE}');\n"
            f"from generate_from_opendata import generate_theme;\n"
            f"generate_theme('{dataset_id}', {year})\n"
        )
        _run([PYTHON_EXE, "-c", code])
    else:
        candidate = OUTPUT_DIR / f"{dataset_id}_{year}.zip"
        if candidate.exists():
            return candidate
        print(f"[GEN] moca {dataset_id} {year}...", file=sys.stderr)
        code = (
            f"import sys; sys.path.insert(0, r'{BASE}');\n"
            f"from prisme_engine import generate_prisme_excel;\n"
            f"generate_prisme_excel('{dataset_id}', {year})\n"
        )
        _run([PYTHON_EXE, "-c", code])

    if not candidate.exists():
        raise FileNotFoundError(f"ZIP not generated: {candidate}")
    return candidate


def read_data_from_zip(zip_path: Path, tmpdir: Path) -> dict:
    """
    Extract ZIP and read each territory xlsx.
    Returns {terr_key: [(headers, rows), ...]} where terr_key in {com, reg, dom, fra, fh}
    """
    extract_dir = tmpdir / zip_path.stem
    extract_dir.mkdir(exist_ok=True)
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(extract_dir)

    data = {}
    # Walk the extracted tree looking for territory folders
    for xlsx_path in extract_dir.rglob("*.xlsx"):
        parent = xlsx_path.parent.name  # e.g., "Commune", "Région", "DOM"
        # Skip consolidated files at year-level
        if "consolidated" in xlsx_path.name.lower():
            continue
        terr_key = TERR_TO_SHEETKEY.get(parent)
        if not terr_key:
            continue
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = rows[0]
                body = [r for r in rows[1:] if r and any(c is not None and c != '' for c in r)]
                data.setdefault(terr_key, []).append((headers, body))
            wb.close()
        except Exception as e:
            print(f"[WARN] skip {xlsx_path}: {e}", file=sys.stderr)
    return data


def _rename_headers(headers: tuple) -> list:
    """Apply VARIABLE_RENAME to each header if present."""
    return [VARIABLE_RENAME.get(h, h) if isinstance(h, str) else h for h in headers]


def _norm_value(v):
    """Replace empty/None/0.0 placeholders with MISSING_VALUE.

    -9 (et "-9") = code SPF "donnée non disponible / secret statistique" :
    traité comme manquant pour ne jamais apparaître brut dans l'Excel final.
    """
    if v is None or v == '':
        return MISSING_VALUE
    if v == -9 or v == -9.0 or (isinstance(v, str) and v.strip() == '-9'):
        return MISSING_VALUE
    return v


def build_consolidated_xlsx(dataset_id: str, years: list[int],
                            year_data: dict, out_path: Path):
    """
    Build final consolidated xlsx.
    year_data[year] = {com: [(headers, rows), ...], reg: [...], dom: [...], fra: [...], fh: [...]}
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ========== COM: all communes, all years ==========
    com_headers = None
    com_rows = []
    for y in years:
        blocks = year_data.get(y, {}).get("com", [])
        for headers, body in blocks:
            if com_headers is None:
                com_headers = _rename_headers(headers)
            com_rows.extend(body)
    if com_headers:
        ws = wb.create_sheet("COM")
        ws.append(com_headers)
        for row in com_rows:
            ws.append([_norm_value(c) for c in row])
        for c in ws[1]:
            c.font = Font(bold=True)

    # ========== REG YYYY: one sheet per year ==========
    for y in years:
        blocks = year_data.get(y, {}).get("reg", [])
        if not blocks:
            continue
        headers, body = blocks[0]
        ws = wb.create_sheet(f"REG {y}")
        ws.append(_rename_headers(headers))
        for row in body:
            ws.append([_norm_value(c) for c in row])
        for c in ws[1]:
            c.font = Font(bold=True)

    # ========== DROM: all years ==========
    drom_headers = None
    drom_rows = []
    for y in years:
        blocks = year_data.get(y, {}).get("dom", [])
        for headers, body in blocks:
            if drom_headers is None:
                drom_headers = _rename_headers(headers)
            drom_rows.extend(body)
    if drom_headers:
        ws = wb.create_sheet("DROM")
        ws.append(drom_headers)
        for row in drom_rows:
            ws.append([_norm_value(c) for c in row])
        for c in ws[1]:
            c.font = Font(bold=True)

    # ========== franENT: all years ==========
    fra_headers = None
    fra_rows = []
    for y in years:
        blocks = year_data.get(y, {}).get("fra", [])
        for headers, body in blocks:
            if fra_headers is None:
                fra_headers = _rename_headers(headers)
            fra_rows.extend(body)
    if fra_headers:
        ws = wb.create_sheet("franENT")
        ws.append(fra_headers)
        for row in fra_rows:
            ws.append([_norm_value(c) for c in row])
        for c in ws[1]:
            c.font = Font(bold=True)

    # ========== FranHEX: all years ==========
    fh_headers = None
    fh_rows = []
    for y in years:
        blocks = year_data.get(y, {}).get("fh", [])
        for headers, body in blocks:
            if fh_headers is None:
                fh_headers = _rename_headers(headers)
            fh_rows.extend(body)
    if fh_headers:
        ws = wb.create_sheet("FranHEX")
        ws.append(fh_headers)
        for row in fh_rows:
            ws.append([_norm_value(c) for c in row])
        for c in ws[1]:
            c.font = Font(bold=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dataset_id")
    ap.add_argument("year_start", type=int)
    ap.add_argument("year_end", type=int)
    ap.add_argument("--source", choices=["moca", "opendata"], default="moca")
    args = ap.parse_args()

    years = list(range(args.year_start, args.year_end + 1))
    print(f"[MOCA-CONS] dataset={args.dataset_id} years={years} source={args.source}", file=sys.stderr)

    tmpdir = Path(tempfile.mkdtemp(prefix="mocao_cons_"))
    try:
        year_data = {}
        for y in years:
            try:
                zp = generate_year_zip(args.dataset_id, y, source=args.source)
                year_data[y] = read_data_from_zip(zp, tmpdir)
            except Exception as e:
                print(f"[WARN] year {y} skipped: {e}", file=sys.stderr)
                year_data[y] = {}

        out_path = OUTPUT_DIR / f"{args.dataset_id}_mocao_{args.year_start}_{args.year_end}.xlsx"
        build_consolidated_xlsx(args.dataset_id, years, year_data, out_path)
        print(out_path.name)  # stdout = filename for file_server.js
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    main()
