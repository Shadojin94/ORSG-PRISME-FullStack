import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import shutil

# Mock setup
BASE_DIR = Path("test_output")
if BASE_DIR.exists():
    shutil.rmtree(BASE_DIR)
BASE_DIR.mkdir()

ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

def generate_folder_structure_test(dataset_name, year):
    # Mock Data
    variables = ['var1', 'var2']
    data = {
        'com': [{'com': 97301, 'annee': year, 'var1': 10, 'var2': 20}],
        'reg': [{'reg': 1, 'annee': year, 'var1': 11, 'var2': 21}],
        'dom': [{'dom': 'DOM', 'annee': year, 'var1': 12, 'var2': 22}],
        'fh': [{'fh': 0, 'annee': year, 'var1': 13, 'var2': 23}],
        'fra': [{'fra': 99, 'annee': year, 'var1': 14, 'var2': 24}]
    }

    # Folder Mapping
    # Transcript: "Commune, DOM, Région, France hors Mayotte, France hexagonale"
    # Mapping keys from code to Folder Names
    folder_mapping = {
        'com': 'Communes',
        'dom': 'DOM',
        'reg': 'Regions',
        'fh': 'France_Hexagonale',
        'fra': 'France_Hors_Mayotte' # Assuming 'fra' corresponds to this
    }

    # Create Root Folder: {Theme}/{Year}
    # But wait, looking at transcript: "Education > 2017 > Commune > educ.xlsx"
    # So Root is {Year} inside a Theme folder? 
    # The existing backend typically handles one request at a time. The user downloads "Education 2017".
    # So the zip should probably contain the "2017" folder which contains the subfolders.
    
    root_folder_name = f"{year}"
    root_path = BASE_DIR / root_folder_name
    root_path.mkdir()

    for key, folder_name in folder_mapping.items():
        # Create subfolder
        subfolder = root_path / folder_name
        subfolder.mkdir()
        
        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = key # Naming the sheet as the key (e.g. 'com')
        
        # Add Headers
        headers = [key, 'annee'] + variables
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = ORANGE_FILL
            
        # Add Data
        rows = data.get(key, [])
        for row_idx, row_dict in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_dict.get(key))
            ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
            ws.cell(row=row_idx, column=3, value=10) # Mock
            ws.cell(row=row_idx, column=4, value=20) # Mock

        # Save File
        # Filename must be consistent. e.g. "educ.xlsx"
        filename = f"{dataset_name}.xlsx"
        wb.save(subfolder / filename)
        print(f"Created {subfolder / filename}")

    # Create ZIP
    shutil.make_archive(str(BASE_DIR / f"{dataset_name}_{year}"), 'zip', BASE_DIR)
    print(f"Created Zip: {BASE_DIR / f'{dataset_name}_{year}.zip'}")

if __name__ == "__main__":
    generate_folder_structure_test('educ', 2017)
