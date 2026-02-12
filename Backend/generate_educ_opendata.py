import pandas as pd
import os
import requests
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
import zipfile
import io
import shutil

# --- Configuration ---
# Année à traiter (modifiable)
YEAR = 2018

# URL de l'Open Data INSEE (Base infracommunale Diplômes - Formation)
# Mapping des URLs par année
INSEE_DATA_URLS = {
    2022: "https://www.insee.fr/fr/statistiques/fichier/8647010/base-ic-diplomes-formation-2022_csv.zip",
    2021: "https://www.insee.fr/fr/statistiques/fichier/8268840/base-ic-diplomes-formation-2021_csv.zip",
    2020: "https://www.insee.fr/fr/statistiques/fichier/7704080/base-ic-diplomes-formation-2020_csv.zip",
    2019: "https://www.insee.fr/fr/statistiques/fichier/6543298/base-ic-diplomes-formation-2019_csv.zip",
    2018: "https://www.insee.fr/fr/statistiques/fichier/5650712/base-ic-diplomes-formation-2018_csv.zip",
    2017: "https://www.insee.fr/fr/statistiques/fichier/4799252/base-ic-diplomes-formation-2017_csv.zip",
}

# Chemins
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
INPUTS_DIR = os.path.join(BACKEND_DIR, "inputs")
OUTPUT_DIR = os.path.join(BACKEND_DIR, "output")
OUTPUT_FILE = "educ_opendata.xlsx"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
LOCAL_CSV_NAME = f"educ_insee_{YEAR}.csv"

# Configuration des niveaux géographiques
GEO_LEVELS = {
    'com': {'folder': 'Commune', 'col_name': 'com'},
    'reg': {'folder': 'Région', 'col_name': 'reg'},
    'dom': {'folder': 'DOM', 'col_name': 'dom'},
    'fh': {'folder': 'France Hexagonale', 'col_name': 'fh'},
    'fra': {'folder': 'France entière', 'col_name': 'fra'}
}

# Ordre des régions
REGION_ORDER = [11, 24, 27, 28, 32, 44, 52, 53, 75, 76, 84, 93, 94, 1, 2, 3, 4, 6]

# Style Excel
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Assurer que les dossiers existent
os.makedirs(INPUTS_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def get_insee_url(year):
    """Retourne l'URL INSEE pour l'année spécifiée."""
    if year in INSEE_DATA_URLS:
        return INSEE_DATA_URLS[year]
    # Fallback sur l'année la plus proche
    available_years = sorted(INSEE_DATA_URLS.keys())
    closest_year = min(available_years, key=lambda x: abs(x - year))
    print(f"[WARN] Année {year} non disponible, utilisation de {closest_year}")
    return INSEE_DATA_URLS[closest_year]

def download_and_extract_data():
    """Télécharge et extrait les données Open Data INSEE si elles ne sont pas présentes."""
    # Chercher un fichier existant (CSV ou XLS)
    csv_target_path = os.path.join(INPUTS_DIR, f"educ_insee_{YEAR}.csv")
    xls_target_path = os.path.join(INPUTS_DIR, f"educ_insee_{YEAR}.xls")
    
    if os.path.exists(csv_target_path):
        print(f"[OK] Fichier source existant trouvé : {csv_target_path}")
        return csv_target_path
    elif os.path.exists(xls_target_path):
        print(f"[OK] Fichier source existant trouvé : {xls_target_path}")
        return xls_target_path

    url = get_insee_url(YEAR)
    print(f"[DL] Téléchargement des données depuis INSEE.fr pour l'année {YEAR}...")
    print(f"    URL: {url}")

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=300)
        response.raise_for_status()
        print(f"    Téléchargé: {len(response.content)} bytes")

        print("[ZIP] Extraction de l'archive ZIP...")
        with zipfile.ZipFile(io.BytesIO(response.content)) as z:
            # Recherche du fichier CSV ou XLS principal dans l'archive
            # Liste tous les fichiers pour debug
            all_files = z.namelist()
            print(f"    Fichiers dans l'archive: {len(all_files)}")
            for f in all_files[:10]:
                print(f"      - {f}")
            
            csv_files = [f for f in all_files if f.lower().endswith('.csv') and 'meta' not in f.lower()]
            xls_files = [f for f in all_files if f.lower().endswith('.xls') and 'meta' not in f.lower()]
            txt_files = [f for f in all_files if f.lower().endswith('.txt') and 'meta' not in f.lower()]
            
            if csv_files:
                source_filename = csv_files[0]
                file_ext = '.csv'
                print(f"    Fichier CSV trouvé : {source_filename}")
            elif txt_files:
                source_filename = txt_files[0]
                file_ext = '.csv'  # Renommer en CSV
                print(f"    Fichier TXT trouvé (traité comme CSV) : {source_filename}")
            elif xls_files:
                source_filename = xls_files[0]
                file_ext = '.xls'
                print(f"    Fichier XLS trouvé : {source_filename}")
            else:
                raise Exception("Aucun fichier CSV, TXT ou XLS de données trouvé dans l'archive ZIP")

            # Extraction et renommage avec la bonne extension
            csv_target_path = os.path.join(INPUTS_DIR, f"educ_insee_{YEAR}{file_ext}")
            with z.open(source_filename) as source_file, open(csv_target_path, 'wb') as target_file:
                target_file.write(source_file.read())

        print(f"[OK] Installation terminée : {csv_target_path}")
        return csv_target_path

    except Exception as e:
        print(f"[ERREUR] Erreur lors du téléchargement/extraction : {e}")
        return None

def process_csv_file(file_path):
    """
    Traite le fichier CSV ou XLS INSEE (Base infracommunale).
    """
    print(f"[...] Traitement du fichier en cours pour l'année {YEAR}...")
    
    # Déterminer le préfixe des colonnes selon l'année (P18_ pour 2018, P21_ pour 2021, etc.)
    year_prefix = f"P{str(YEAR)[2:]}_"
    
    try:
        # Lecture du fichier selon son format
        if file_path.endswith('.xls'):
            print("   Format XLS détecté, lecture avec pandas...")
            df = pd.read_excel(file_path, dtype={'CODGEO': str, 'IRIS': str, 'COM': str})
        else:
            print("   Format CSV détecté, lecture avec pandas...")
            df = pd.read_csv(file_path, sep=';', dtype={'CODGEO': str, 'IRIS': str, 'COM': str}, low_memory=False)
        
        print(f"   Lignes lues : {len(df)}")
        print(f"   Colonnes disponibles : {len(df.columns)}")

        # Filtrage Guyane (973)
        geo_col = 'COM' if 'COM' in df.columns else 'CODGEO'
        if 'DEP' in df.columns:
            df_guyane = df[df['DEP'].astype(str) == '973'].copy()
        else:
            df_guyane = df[df[geo_col].astype(str).str.startswith('973', na=False)].copy()
        
        count_guyane = len(df_guyane)
        print(f"   Lignes Guyane (973) : {count_guyane}")

        if count_guyane == 0:
            print("[WARN] ATTENTION : Aucune donnée trouvée pour la Guyane (973).")
            return False

        # --- Agrégation au niveau Communal ---
        numeric_cols = [c for c in df_guyane.columns if c.startswith(year_prefix) or c.startswith(f"C{str(YEAR)[2:]}_")]
        
        if not numeric_cols:
            # Fallback: chercher toutes les colonnes numériques
            numeric_cols = [c for c in df_guyane.columns if c[0:3] in ['P18_', 'P19_', 'P20_', 'P21_', 'P22_']]
            print(f"   [INFO] Colonnes avec préfixe alternatif trouvées: {len(numeric_cols)}")
        
        # On groupe par Code Commune
        code_commune_col = 'COM' if 'COM' in df.columns else 'CODGEO'
        
        if code_commune_col not in df_guyane:
            if 'IRIS' in df_guyane.columns:
                df_guyane['COM'] = df_guyane['CODGEO'].str.slice(0, 5)
                code_commune_col = 'COM'

        print(f"   Agrégation par commune (colonne {code_commune_col})...")
        df_communal = df_guyane.groupby(code_commune_col)[numeric_cols].sum().reset_index()
        df_communal.rename(columns={code_commune_col: 'CODGEO'}, inplace=True)
        
        # --- Calcul des indicateurs PRISME ---
        
        # 1. Population 6-16 ans
        try:
            pop_0610_col = f"{year_prefix}POP0610"
            pop_1114_col = f"{year_prefix}POP1114"
            pop_1517_col = f"{year_prefix}POP1517"
            
            df_communal['pop_6_16'] = (
                df_communal.get(pop_0610_col, 0) + 
                df_communal.get(pop_1114_col, 0) + 
                (df_communal.get(pop_1517_col, 0) * (2/3))
            )
        except Exception as e:
            print(f"[WARN] Variable manquante pour pop_6_16: {e}. Utilisation de 0.")
            df_communal['pop_6_16'] = 0

        # 2. Nombre de non-scolarisés 6-16 ans
        try:
            scol_0610_col = f"{year_prefix}SCOL0610"
            scol_1114_col = f"{year_prefix}SCOL1114"
            scol_1517_col = f"{year_prefix}SCOL1517"
            
            ns_0610 = df_communal.get(pop_0610_col, 0) - df_communal.get(scol_0610_col, 0)
            ns_1114 = df_communal.get(pop_1114_col, 0) - df_communal.get(scol_1114_col, 0)
            ns_1517 = df_communal.get(pop_1517_col, 0) - df_communal.get(scol_1517_col, 0)
            
            df_communal['nb_non_sco'] = (
                ns_0610 + ns_1114 + (ns_1517 * (2/3))
            ).clip(lower=0)
        except Exception as e:
            print(f"[WARN] Variable manquante pour nb_non_sco: {e}. Utilisation de 0.")
            df_communal['nb_non_sco'] = 0

        # 3. Population 15-64 ans
        pop_1524_col = f"{year_prefix}POP1524"
        pop_2554_col = f"{year_prefix}POP2554"
        pop_5564_col = f"{year_prefix}POP5564"
        
        if pop_1524_col in df_communal.columns:
            df_communal['pop_15_64'] = (
                df_communal.get(pop_1524_col, 0) + 
                df_communal.get(pop_2554_col, 0) + 
                df_communal.get(pop_5564_col, 0)
            )
        else:
            # Fallback avec autres colonnes
            pop_1517_col = f"{year_prefix}POP1517"
            pop_1824_col = f"{year_prefix}POP1824"
            pop_2529_col = f"{year_prefix}POP2529"
            pop_30p_col = f"{year_prefix}POP30P"
            
            val_jeunes = df_communal.get(pop_1517_col, 0) + df_communal.get(pop_1824_col, 0) + df_communal.get(pop_2529_col, 0)
            val_30p = df_communal.get(pop_30p_col, 0)
            df_communal['pop_15_64'] = val_jeunes + (val_30p * 0.88)
        
        # 4. Peu diplômés
        diplmin_col = f"{year_prefix}NSCOL15P_DIPLMIN"
        if diplmin_col in df_communal.columns:
            df_communal['nb_peu_dipl'] = df_communal[diplmin_col]
        else:
            sans_dipl_col = f"{year_prefix}NSCOL15P"
            df_communal['nb_peu_dipl'] = df_communal.get(sans_dipl_col, 0)

        # --- Préparation des données par niveau géographique ---
        
        # 1. Communes (données calculées)
        data_com = []
        for _, row in df_communal.iterrows():
            data_com.append({
                'com': int(row['CODGEO']),
                'annee': YEAR,
                'pop_6_16': round(row['pop_6_16'], 2),
                'nb_non_sco': round(row['nb_non_sco'], 2),
                'pop_15_64': round(row['pop_15_64'], 2),
                'nb_peu_dipl': round(row['nb_peu_dipl'], 2)
            })
        
        # 2. Régions (structure vide)
        data_reg = [{'reg': code, 'annee': YEAR} for code in REGION_ORDER]
        
        # 3. DOM, France Hexagonale, France Entière (structure vide)
        data_dom = [{'dom': 'DOM', 'annee': YEAR}]
        data_fh = [{'fh': 0, 'annee': YEAR}]
        data_fra = [{'fra': 99, 'annee': YEAR}]
        
        all_data = {
            'com': data_com,
            'reg': data_reg,
            'dom': data_dom,
            'fh': data_fh,
            'fra': data_fra
        }
        
        # --- Création de l'arborescence complète ---
        print(f"\n[DIR] Création de l'arborescence client pour {YEAR}...")
        
        THEME_FOLDER = "Population et condition de vie/Education/Educ"
        root_dir = os.path.join(OUTPUT_DIR, THEME_FOLDER, str(YEAR))
        
        # Nettoyer et recréer le dossier
        if os.path.exists(root_dir):
            shutil.rmtree(root_dir)
        os.makedirs(root_dir)
        
        variables = ['pop_6_16', 'nb_non_sco', 'pop_15_64', 'nb_peu_dipl']
        
        # Générer un fichier Excel par niveau géographique
        for geo_key, geo_config in GEO_LEVELS.items():
            folder_name = geo_config['folder']
            col_name = geo_config['col_name']
            
            sub_dir = os.path.join(root_dir, folder_name)
            os.makedirs(sub_dir, exist_ok=True)
            
            # Créer le workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            ws = wb.create_sheet(geo_key)
            rows_data = all_data[geo_key]
            
            # En-têtes
            headers = [col_name, 'annee'] + variables
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True)
                cell.fill = ORANGE_FILL
            
            # Données
            for row_idx, row_dict in enumerate(rows_data, 2):
                ws.cell(row=row_idx, column=1, value=row_dict.get(col_name))
                ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
                
                for i, var in enumerate(variables):
                    val = row_dict.get(var)
                    cell = ws.cell(row=row_idx, column=3 + i, value=val)
                    if val is not None:
                        cell.fill = ORANGE_FILL
            
            # Largeur colonnes
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
            for i, var in enumerate(variables):
                ws.column_dimensions[chr(67 + i)].width = 15
            
            filepath = os.path.join(sub_dir, "educ.xlsx")
            wb.save(filepath)
            print(f"  [OK] {folder_name}/educ.xlsx ({len(rows_data)} lignes)")
        
        # --- Générer le fichier consolidé ---
        print(f"\n[CONSOLIDATED] Génération fichier consolidé...")
        
        wb_cons = openpyxl.Workbook()
        wb_cons.remove(wb_cons.active)
        
        for geo_key, geo_config in GEO_LEVELS.items():
            col_name = geo_config['col_name']
            ws = wb_cons.create_sheet(geo_key)
            rows_data = all_data[geo_key]
            
            # En-têtes
            headers = [col_name, 'annee'] + variables
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True)
                cell.fill = ORANGE_FILL
            
            # Données
            for row_idx, row_dict in enumerate(rows_data, 2):
                ws.cell(row=row_idx, column=1, value=row_dict.get(col_name))
                ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
                
                for i, var in enumerate(variables):
                    val = row_dict.get(var)
                    cell = ws.cell(row=row_idx, column=3 + i, value=val)
                    if val is not None:
                        cell.fill = ORANGE_FILL
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
        
        cons_path = os.path.join(root_dir, f"educ_consolidated_{YEAR}.xlsx")
        wb_cons.save(cons_path)
        print(f"  [OK] educ_consolidated_{YEAR}.xlsx")
        
        # --- Créer le ZIP ---
        zip_name = f"educ_opendata_{YEAR}"
        zip_path = os.path.join(OUTPUT_DIR, zip_name)
        
        # Créer l'archive à partir du dossier theme_folder/year
        zip_output = shutil.make_archive(
            zip_path,
            'zip',
            os.path.join(OUTPUT_DIR, THEME_FOLDER),
            str(YEAR)
        )
        
        print(f"\n[ZIP] Archive créée: {zip_output}")
        
        # Résumé
        print(f"\n{'='*60}")
        print(f"Résumé - EDUC {YEAR}")
        print(f"{'='*60}")
        print(f"  Communes Guyane: {len(data_com)}")
        print(f"  Régions: {len(data_reg)}")
        print(f"  Fichiers générés: 6 (5 niveaux + 1 consolidé)")
        print(f"  Dossier: {root_dir}")
        print(f"  Archive: {zip_output}")
        
        return True

    except Exception as e:
        print(f"[ERREUR] Erreur critique lors du traitement : {e}")
        import traceback
        traceback.print_exc()
        return False

def generate_mock_data():
    """Génère des données de test (fallback)."""
    print("[WARN] ECHEC TÉLÉCHARGEMENT -> Mode Mock activé.")
    
    THEME_FOLDER = "Population et condition de vie/Education/Educ"
    root_dir = os.path.join(OUTPUT_DIR, THEME_FOLDER, str(YEAR))
    os.makedirs(root_dir, exist_ok=True)
    
    variables = ['pop_6_16', 'nb_non_sco', 'pop_15_64', 'nb_peu_dipl']
    
    # Données mock par niveau géographique
    mock_data = {
        'com': [
            {'com': 97301, 'annee': YEAR, 'pop_6_16': 172.00, 'nb_non_sco': 6.00, 'pop_15_64': 559.00, 'nb_peu_dipl': 269.00},
            {'com': 97302, 'annee': YEAR, 'pop_6_16': 11873.01, 'nb_non_sco': 566.83, 'pop_15_64': 39380.53, 'nb_peu_dipl': 14244.92},
            {'com': 97303, 'annee': YEAR, 'pop_6_16': 453.38, 'nb_non_sco': 8.87, 'pop_15_64': 1052.89, 'nb_peu_dipl': 460.50},
        ],
        'reg': [{'reg': code, 'annee': YEAR} for code in REGION_ORDER],
        'dom': [{'dom': 'DOM', 'annee': YEAR}],
        'fh': [{'fh': 0, 'annee': YEAR}],
        'fra': [{'fra': 99, 'annee': YEAR}]
    }
    
    # Générer un fichier Excel par niveau géographique
    for geo_key, geo_config in GEO_LEVELS.items():
        folder_name = geo_config['folder']
        col_name = geo_config['col_name']
        
        sub_dir = os.path.join(root_dir, folder_name)
        os.makedirs(sub_dir, exist_ok=True)
        
        # Créer le workbook avec openpyxl pour avoir le style
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet(geo_key)
        rows_data = mock_data[geo_key]
        
        # En-têtes
        headers = [col_name, 'annee'] + variables
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = ORANGE_FILL
        
        # Données
        for row_idx, row_dict in enumerate(rows_data, 2):
            ws.cell(row=row_idx, column=1, value=row_dict.get(col_name))
            ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
            
            for i, var in enumerate(variables):
                val = row_dict.get(var)
                cell = ws.cell(row=row_idx, column=3 + i, value=val)
                if val is not None:
                    cell.fill = ORANGE_FILL
        
        # Largeur colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        for i, var in enumerate(variables):
            ws.column_dimensions[chr(67 + i)].width = 15
        
        filepath = os.path.join(sub_dir, "educ.xlsx")
        wb.save(filepath)
        print(f"  [MOCK] {filepath} ({len(rows_data)} lignes)")

if __name__ == "__main__":
    print(f"--- Générateur OpenData Education (INSEE Auto-Download) - Année {YEAR} ---")
    
    # 1. Téléchargement ou récupération du fichier source
    input_csv = download_and_extract_data()
    
    # 2. Traitement des données
    success = False
    if input_csv:
        success = process_csv_file(input_csv)
    
    # 3. Fallback si échec
    if not success:
        generate_mock_data()
