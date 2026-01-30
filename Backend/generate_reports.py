#!/usr/bin/env python3
"""
PRISME Engine - Génération de rapports Excel pour ORSG
Intégration PocketBase + logique prisme_engine_simple.py
"""

import time
import re
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import warnings
import os

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

POCKETBASE_URL = "http://127.0.0.1:8090"
ADMIN_EMAIL = os.getenv("POCKETBASE_ADMIN_EMAIL", "admin@example.com")
ADMIN_PASSWORD = os.getenv("POCKETBASE_ADMIN_PASSWORD", "ChangeMe123!")

BASE_DIR = Path(__file__).parent
TEMP_DIR = BASE_DIR / "temp"
CSV_SOURCES_DIR = BASE_DIR / "csv_sources"
OUTPUT_DIR = BASE_DIR / "output"

TEMP_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# ============================================================================
# REGION & COMMUNE CONSTANTS
# ============================================================================

REGION_ORDER = [11, 24, 27, 28, 32, 44, 52, 53, 75, 76, 84, 93, 94, 1, 2, 3, 4, 6]

REGION_MAPPING = {
    "Île-de-France": 11, "Ile-de-France": 11, "île-de-France": 11,
    "Centre-Val de Loire": 24,
    "Bourgogne-Franche-Comté": 27, "Bourgogne-Franche-ComtÃ©": 27,
    "Normandie": 28,
    "Hauts-de-France": 32,
    "Grand Est": 44,
    "Pays de la Loire": 52,
    "Bretagne": 53,
    "Nouvelle-Aquitaine": 75,
    "Occitanie": 76,
    "Auvergne-Rhône-Alpes": 84, "Auvergne-RhÃ´ne-Alpes": 84,
    "Provence-Alpes-Côte d'Azur": 93,
    "Corse": 94,
    "Guadeloupe": 1,
    "Martinique": 2,
    "Guyane": 3,
    "La Réunion": 4, "Réunion": 4,
    "Mayotte": 6
}

COMMUNES_GUYANE = [
    97301, 97302, 97303, 97304, 97305, 97306, 97307, 97308, 97309, 97310,
    97311, 97312, 97313, 97314, 97352, 97353, 97356, 97357, 97358, 97360,
    97361, 97362
]

REGIONS_INFO = [
    (1, 'Guadeloupe'), (2, 'Martinique'), (3, 'Guyane'), 
    (4, 'La Réunion'), (6, 'Mayotte'), (11, 'Île-de-France'),
    (24, 'Centre-Val de Loire'), (27, 'Bourgogne-Franche-Comté'),
    (28, 'Normandie'), (32, 'Hauts-de-France'), (44, 'Grand Est'),
    (52, 'Pays de la Loire'), (53, 'Bretagne'), (75, 'Nouvelle-Aquitaine'),
    (76, 'Occitanie'), (84, 'Auvergne-Rhône-Alpes'),
    (93, "Provence-Alpes-Côte d'Azur"), (94, 'Corse')
]

# Dataset configurations
DATASET_CONFIGS = {
    'educ': {
        'name': 'Éducation',
        'variables': ['pop_6_16', 'nb_non_sco', 'pop_15_64', 'nb_peu_dipl'],
        'csv_mapping': {
            'pop_6_16': 'Pop_6-16ans',
            'nb_non_sco': 'nb_non_scol',
            'pop_15_64': 'Pop_15-64ans',
            'nb_peu_dipl': 'Nb_peu_dipl'
        },
        'dic_variables': [
            ("Population 6-16 ans", "educ", "pop_6_16"),
            ("Nb jeunes 6-16 ans non scolarisés", "educ", "nb_non_sco"),
            ("Part jeunes 6-16 ans non scolarisés", "educ", "tx_non_sco"),
            ("Population 15-64 ans", "educ", "pop_15_64"),
            ("Nb personnes peu diplômées", "educ", "nb_peu_dipl"),
            ("Part personnes peu diplômées", "educ", "tx_peu_dipl"),
        ]
    }
}

# ============================================================================
# MOCA-O CSV PARSER
# ============================================================================

def parse_moca_csv(filepath):
    """Parse un fichier CSV au format MOCA-O."""
    result = {'com': [], 'reg': [], 'dom': [], 'fh': [], 'fra': []}
    
    try:
        with open(filepath, 'r', encoding='cp1252', errors='replace') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Erreur lecture {filepath}: {e}")
        return {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) for k in result}

    for line in lines:
        if not line.strip():
            continue
        
        parts = line.strip().split(';')
        if len(parts) < 3:
            continue
        
        # Année (premier champ)
        try:
            annee = int(parts[0])
            if not (2000 <= annee <= 2030):
                continue
        except:
            continue
        
        # Valeur (dernier champ numérique)
        valeur = None
        for p in reversed(parts):
            try:
                valeur = float(p.replace(',', '.'))
                break
            except:
                continue
        
        if valeur is None:
            continue
        
        line_lower = line.lower()
        
        # 1. Communes Guyane (973xx)
        match_com = re.search(r'(973\d{2})', line)
        if match_com:
            code = int(match_com.group(1))
            if code in COMMUNES_GUYANE:
                result['com'].append({'annee': annee, 'codgeo': code, 'valeur': valeur})
            continue
        
        # 2. France entière
        if 'france entiere' in line_lower or 'france (y compris mayotte)' in line_lower or 'france entière' in line_lower:
            result['fra'].append({'annee': annee, 'codgeo': 99, 'valeur': valeur})
            continue
        
        # 3. France hexagonale/métropolitaine
        if 'france metropolitaine' in line_lower or 'france hexagonale' in line_lower or 'france métropolitaine' in line_lower:
            result['fh'].append({'annee': annee, 'codgeo': 0, 'valeur': valeur})
            continue
        
        # 4. DOM
        if "departements d'outre" in line_lower or "départements d'outre" in line_lower:
            result['dom'].append({'annee': annee, 'codgeo': 'DOM', 'valeur': valeur})
            continue
        
        # 5. Régions
        for reg_name, reg_code in REGION_MAPPING.items():
            if reg_name.lower() in line_lower:
                result['reg'].append({'annee': annee, 'codgeo': reg_code, 'valeur': valeur})
                break

    # Convertir en DataFrames
    dfs = {}
    for k, v in result.items():
        if v:
            dfs[k] = pd.DataFrame(v).drop_duplicates(subset=['annee', 'codgeo'])
        else:
            dfs[k] = pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])
    
    return dfs


def find_csv_file(pattern, search_dir=None):
    """Trouve un fichier CSV correspondant au pattern."""
    if search_dir is None:
        search_dir = CSV_SOURCES_DIR
    candidates = list(Path(search_dir).glob(f"*{pattern}*.csv"))
    if candidates:
        return candidates[0]
    return None


# ============================================================================
# EXCEL GENERATOR
# ============================================================================

import shutil

def generate_prisme_excel(dataset_name, year, csv_dir=None):
    """Génère une archive ZIP contenant la structure de fichiers PRISME (Structure SharePoint)."""
    
    if dataset_name not in DATASET_CONFIGS:
        print(f"Dataset inconnu: {dataset_name}")
        return None
    
    config = DATASET_CONFIGS[dataset_name]
    variables = config['variables']
    csv_mapping = config['csv_mapping']
    
    # Charger les données CSV
    csv_data = {}
    for var_name, csv_pattern in csv_mapping.items():
        csv_file = find_csv_file(csv_pattern, csv_dir)
        if csv_file:
            csv_data[var_name] = parse_moca_csv(csv_file)
            print(f"  [OK] {var_name} -> {csv_file.name}")
        else:
            print(f"  [WARN] {var_name} -> Fichier non trouvé (pattern: {csv_pattern})")
            csv_data[var_name] = {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) 
                                  for k in ['com', 'reg', 'dom', 'fh', 'fra']}
    
    # Préparer les structures de données
    data = {
        'com': [{'com': c, 'annee': year} for c in COMMUNES_GUYANE],
        'reg': [{'reg': c, 'annee': year} for c in REGION_ORDER],
        'dom': [{'dom': 'DOM', 'annee': year}],
        'fh': [{'fh': 0, 'annee': year}],
        'fra': [{'fra': 99, 'annee': year}]
    }
    
    # Remplir les données pour chaque variable
    for var_name in variables:
        parsed = csv_data.get(var_name, {})
        
        # Communes
        df = parsed.get('com', pd.DataFrame())
        if not df.empty:
            df_year = df[df['annee'] == year]
            for row in data['com']:
                match = df_year[df_year['codgeo'] == row['com']]
                if not match.empty:
                    row[var_name] = round(match.iloc[0]['valeur'], 2)
        
        # Régions
        df = parsed.get('reg', pd.DataFrame())
        if not df.empty:
            df_year = df[df['annee'] == year]
            for row in data['reg']:
                match = df_year[df_year['codgeo'] == row['reg']]
                if not match.empty:
                    row[var_name] = round(match.iloc[0]['valeur'], 2)
        
        # DOM
        df = parsed.get('dom', pd.DataFrame())
        if not df.empty:
            val = df[df['annee'] == year]['valeur']
            if not val.empty:
                data['dom'][0][var_name] = round(val.values[0], 2)
        
        # France Hexagonale
        df = parsed.get('fh', pd.DataFrame())
        if not df.empty:
            val = df[df['annee'] == year]['valeur']
            if not val.empty:
                data['fh'][0][var_name] = round(val.values[0], 2)
        
        # France entière
        df = parsed.get('fra', pd.DataFrame())
        if not df.empty:
            val = df[df['annee'] == year]['valeur']
            if not val.empty:
                data['fra'][0][var_name] = round(val.values[0], 2)
    
    # --- GÉNÉRATION DE L'ARBORESCENCE (TYPE B: SELECTIVE) ---
    
    # Mapping des dossiers (Nom technique -> Nom Dossier Client)
    folder_mapping = {
        'com': 'Commune',
        'reg': 'Region',
        'dom': 'DOM',
        'fh': 'France_Hexagonale',
        'fra': 'France_Entiere'
    }

    # Création du dossier racine temporaire pour l'année
    root_year_dir = OUTPUT_DIR / f"{year}"
    if root_year_dir.exists():
        shutil.rmtree(root_year_dir)
    root_year_dir.mkdir()

    # Génération des fichiers par niveau géographique
    # Configuration des feuilles (clé_data, nom_colonne_id)
    sheet_configs = {
        'com': 'com',
        'dom': 'dom',
        'reg': 'reg',
        'fh': 'fh',
        'fra': 'fra'
    }

    for key, folder_name in folder_mapping.items():
        # Créer le sous-dossier (ex: output/2017/Commune)
        sub_dir = root_year_dir / folder_name
        sub_dir.mkdir(exist_ok=True)
        
        # Créer le fichier Excel
        wb = Workbook()
        wb.remove(wb.active) # Supprime la feuille par défaut
        
        id_col = sheet_configs[key]
        sheet_name = key # Le nom de l'onglet sera le code technique (com, reg...)
        
        ws = wb.create_sheet(sheet_name)
        rows_data = data[key]
        
        # En-têtes
        headers = [id_col, 'annee'] + variables
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = ORANGE_FILL
        
        # Données
        for row_idx, row_dict in enumerate(rows_data, 2):
            ws.cell(row=row_idx, column=1, value=row_dict.get(id_col))
            ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
            
            for i, var in enumerate(variables):
                val = row_dict.get(var)
                cell = ws.cell(row=row_idx, column=3 + i, value=val)
                if val is not None:
                    cell.fill = ORANGE_FILL
        
        # Largeur colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10

        # Sauvegarder avec le nom générique (ex: educ.xlsx)
        filename = f"{dataset_name}.xlsx"
        wb.save(sub_dir / filename)

    # --- GÉNÉRATION FICHIER CONSOLIDÉ (TYPE A) ---
    print(f"  [INFO] Génération fichier consolidé (Type A)...")
    try:
        wb_cons = Workbook()
        wb_cons.remove(wb_cons.active)
        
        # Pour chaque niveau géographique, créer un onglet dans le même fichier
        for key, folder_name in folder_mapping.items():
            id_col = sheet_configs[key]
            sheet_name = key # com, reg, dom...
            
            ws = wb_cons.create_sheet(sheet_name)
            rows_data = data[key]
            
            # En-têtes
            headers = [id_col, 'annee'] + variables
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True)
                cell.fill = ORANGE_FILL
            
            # Données
            for row_idx, row_dict in enumerate(rows_data, 2):
                ws.cell(row=row_idx, column=1, value=row_dict.get(id_col))
                ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
                
                for i, var in enumerate(variables):
                    val = row_dict.get(var)
                    cell = ws.cell(row=row_idx, column=3 + i, value=val)
                    if val is not None:
                        cell.fill = ORANGE_FILL
            
            # Largeur colonnes
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
            
        # Sauvegarder le fichier consolidé à la racine de l'année
        cons_filename = f"{dataset_name}_consolidated.xlsx"
        wb_cons.save(root_year_dir / cons_filename)
        print(f"  [OK] Fichier consolidé créé: {cons_filename}")
        
    except Exception as e:
        print(f"  [ERROR] Erreur création consolidé: {e}")
        
    # Créer le ZIP final
    zip_base_name = OUTPUT_DIR / f"{dataset_name}_{year}"

    zip_path = shutil.make_archive(str(zip_base_name), 'zip', str(OUTPUT_DIR), str(year)) # Zip uniquement le dossier de l'année
    
    # Nettoyage du dossier temporaire
    # shutil.rmtree(root_year_dir) # MODIFICATION: On garde les fichiers pour permettre le téléchargement sélectif
    
    print(f"[OK] Genere Archive: {zip_path}")
    return Path(zip_path)


# ============================================================================
# POCKETBASE CLIENT
# ============================================================================

class PocketBaseClient:
    def __init__(self, url):
        self.url = url
        self.token = None

    def auth(self, email, password):
        try:
            resp = requests.post(f"{self.url}/api/admins/auth-with-password", json={
                "identity": email, "password": password
            })
            if resp.status_code == 200:
                self.token = resp.json()["token"]
                print("[OK] Authentifie via PocketBase Admin")
                return True
        except Exception as e:
            print(f"Auth Error: {e}")
        return False

    def get_pending_inputs(self):
        headers = {"Authorization": self.token}
        resp = requests.get(
            f"{self.url}/api/collections/inputs/records",
            headers=headers,
            params={"filter": "status='pending'"}
        )
        if resp.status_code == 200:
            return resp.json().get("items", [])
        return []

    def download_file(self, collection, record_id, filename):
        file_url = f"{self.url}/api/files/{collection}/{record_id}/{filename}"
        r = requests.get(file_url, headers={"Authorization": self.token})
        local_path = TEMP_DIR / filename
        with open(local_path, 'wb') as f:
            f.write(r.content)
        return local_path

    def update_input_status(self, record_id, status, logs=""):
        requests.patch(
            f"{self.url}/api/collections/inputs/records/{record_id}",
            headers={"Authorization": self.token, "Content-Type": "application/json"},
            json={"status": status, "logs": logs}
        )

    def create_report(self, year, file_path):
        url = f"{self.url}/api/collections/reports/records"
        
        with open(file_path, 'rb') as f:
            files = {'file': f}
            data = {
                "year": year,
                "status": "completed"
            }
            resp = requests.post(url, headers={"Authorization": self.token}, data=data, files=files)
            if resp.status_code == 200:
                print(f"[OK] Report uploade: {file_path.name}")
                return True
            else:
                print(f"[ERROR] Upload failed: {resp.text}")
                return False


# ============================================================================
# MAIN WATCHDOG LOOP
# ============================================================================

def main():
    print("=" * 60)
    print("PRISME Engine - Watchdog Mode")
    print("=" * 60)
    
    pb = PocketBaseClient(POCKETBASE_URL)
    
    if not pb.auth(ADMIN_EMAIL, ADMIN_PASSWORD):
        print("[ERROR] Impossible de s'authentifier. Arret.")
        return

    print("[WATCH] En attente de nouveaux inputs...")
    
    while True:
        try:
            inputs = pb.get_pending_inputs()
            
            if inputs:
                print(f"\n[INPUT] {len(inputs)} input(s) en attente")
                
                for item in inputs:
                    print(f"[>] Traitement {item['id']}...")
                    
                    try:
                        # Déterminer l'année et le thème
                        year = int(item.get('year', 2022))
                        theme = item.get('theme', 'educ')
                        
                        # Générer le fichier Excel avec les CSV sources locaux
                        output_path = generate_prisme_excel(theme, year, CSV_SOURCES_DIR)
                        
                        if output_path and output_path.exists():
                            # Uploader le rapport
                            pb.create_report(year, output_path)
                            pb.update_input_status(item['id'], "processed", f"Généré: {output_path.name}")
                        else:
                            pb.update_input_status(item['id'], "error", "Génération échouée")
                            
                    except Exception as e:
                        print(f"[ERROR] Erreur: {e}")
                        pb.update_input_status(item['id'], "error", str(e))
            
            time.sleep(5)
            
        except KeyboardInterrupt:
            print("\n[STOP] Arret demande.")
            break
        except Exception as e:
            print(f"Erreur boucle: {e}")
            time.sleep(5)


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        # Test standalone: génère educ_2022.xlsx
        print("Mode test: génération educ_2022.xlsx")
        generate_prisme_excel('educ', 2022)
    else:
        # Mode watchdog
        main()
