#!/usr/bin/env python3
"""
PRISME — Générateur Open Data Unifié
Génère les fichiers Excel PRISME à partir des sources primaires Open Data.

Thèmes supportés:
  - educ : Education (Base Diplômes-Formation)
  - pers_sup65ans_seules : Personnes 65+ seules (Base Couples-Familles-Ménages)
  - familles_mono : Familles monoparentales (Base Couples-Familles-Ménages)
  - pop_inf3ans : Population < 3 ans (Base Couples-Familles-Ménages)
  - pers_menages : Personnes par ménage (Base Couples-Familles-Ménages)
  - types_menages : Types de ménages (Base Couples-Familles-Ménages)

Usage:
  py generate_opendata_all.py --theme pers_sup65ans_seules --year 2022
  py generate_opendata_all.py --theme all --year 2022
  py generate_opendata_all.py --list
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import shutil
import argparse
import sys
import os

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).parent
INPUTS_DIR = BASE_DIR / "inputs" / "opendata"
OUTPUT_DIR = BASE_DIR / "output"
CSV_SOURCES_DIR = BASE_DIR / "csv_sources"

OUTPUT_DIR.mkdir(exist_ok=True)

ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Communes de Guyane
COMMUNES_GUYANE = [
    "97301", "97302", "97303", "97304", "97305", "97306", "97307", "97308",
    "97309", "97310", "97311", "97312", "97313", "97314", "97352", "97353",
    "97356", "97357", "97358", "97360", "97361", "97362"
]

# Ordre des régions pour l'onglet reg
REGION_ORDER = [11, 24, 27, 28, 32, 44, 52, 53, 75, 76, 84, 93, 94, 1, 2, 3, 4, 6]

# Niveaux géographiques
GEO_LEVELS = {
    'com': 'Commune',
    'reg': 'Région',
    'dom': 'DOM',
    'fh': 'France Hexagonale',
    'fra': 'France entière'
}

# Mapping département -> région
DEP_TO_REG = {
    "01": "84", "02": "32", "03": "84", "04": "93", "05": "93", "06": "93",
    "07": "84", "08": "44", "09": "76", "10": "44", "11": "76", "12": "76",
    "13": "93", "14": "28", "15": "84", "16": "75", "17": "75", "18": "24",
    "19": "75", "21": "27", "22": "53", "23": "75", "24": "75", "25": "27",
    "26": "84", "27": "28", "28": "24", "29": "53", "2A": "94", "2B": "94",
    "30": "76", "31": "76", "32": "76", "33": "75", "34": "76", "35": "53",
    "36": "24", "37": "24", "38": "84", "39": "27", "40": "75", "41": "24",
    "42": "84", "43": "84", "44": "52", "45": "24", "46": "76", "47": "75",
    "48": "76", "49": "52", "50": "28", "51": "44", "52": "44", "53": "52",
    "54": "44", "55": "44", "56": "53", "57": "44", "58": "27", "59": "32",
    "60": "32", "61": "28", "62": "32", "63": "84", "64": "75", "65": "76",
    "66": "76", "67": "44", "68": "44", "69": "84", "70": "27", "71": "27",
    "72": "52", "73": "84", "74": "84", "75": "11", "76": "28", "77": "11",
    "78": "11", "79": "75", "80": "32", "81": "76", "82": "76", "83": "93",
    "84": "93", "85": "52", "86": "75", "87": "75", "88": "44", "89": "27",
    "90": "27", "91": "11", "92": "11", "93": "11", "94": "11", "95": "11",
    # DOM
    "971": "01", "972": "02", "973": "03", "974": "04", "976": "06"
}

# DOM codes
DOM_CODES = ["01", "02", "03", "04", "06"]  # Guadeloupe, Martinique, Guyane, Réunion, Mayotte
FH_REGIONS = [str(r) for r in [11, 24, 27, 28, 32, 44, 52, 53, 75, 76, 84, 93, 94]]

# =============================================================================
# CONFIGURATIONS DES THÈMES
# =============================================================================

THEME_CONFIGS = {
    "pers_sup65ans_seules": {
        "name": "Conditions de vie anciens",
        "source_file": "couples_familles_{year}.csv",
        "folder": "Population et condition de vie/Conditions de vie/Condition de vie anciens",
        "excel_name": "pers_sup65ans_seules",
        "variables": ["nb_pop_65ans", "nb_pop_65ans_seule"],
        "description": "Population 65+ et personnes seules 65+"
    },
    "familles_mono": {
        "name": "Familles monoparentales",
        "source_file": "couples_familles_{year}.csv",
        "folder": "Population et condition de vie/Conditions de vie/Familles monoparentales",
        "excel_name": "familles_mono",
        "variables": ["nb_familles_enf", "nb_familles_mono_enf"],
        "description": "Familles avec enfants et familles monoparentales"
    },
    "pop_inf3ans": {
        "name": "Population moins de 3 ans",
        "source_file": "couples_familles_{year}.csv",
        "folder": "Population et condition de vie/Conditions de vie/Population inf 3 ans",
        "excel_name": "pop_inf3ans",
        "variables": ["nb_enfants_0_2", "nb_enfants_0_5"],
        "description": "Nombre d'enfants de 0-2 ans et 0-5 ans"
    },
    "pers_menages": {
        "name": "Personnes par ménage",
        "source_file": "couples_familles_{year}.csv",
        "folder": "Population et condition de vie/Conditions de vie/Personnes et menages",
        "excel_name": "pers_menages",
        "variables": ["nb_menages", "nb_pers_menages"],
        "description": "Nombre de ménages et personnes des ménages"
    },
    "types_menages": {
        "name": "Types de ménages",
        "source_file": "couples_familles_{year}.csv",
        "folder": "Population et condition de vie/Conditions de vie/Types de menages",
        "excel_name": "types_menages",
        "variables": ["nb_men_seul", "nb_men_couple_senf", "nb_men_couple_aenf", "nb_men_fam_mono"],
        "description": "Répartition des ménages par type"
    },
    "alloc": {
        "name": "Allocataires",
        "source_file": "caf_allocataires_2023.csv",
        "folder": "Population et condition de vie/Conditions de vie/Allocataires",
        "excel_name": "alloc",
        "variables": ["nb_alloc", "nb_foyers_rsa", "nb_pers_rsa", "nb_foyers_apl", "nb_pers_apl"],
        "description": "Allocataires CAF : total, RSA, APL",
        "loader": "caf"  # Loader spécifique
    },
}


# =============================================================================
# FONCTIONS DE CHARGEMENT ET FILTRAGE
# =============================================================================

def load_and_filter_csv(source_path: Path, year: int) -> pd.DataFrame:
    """Charge le CSV INSEE et filtre les données Guyane, agrégées par commune."""
    print(f"  [LOAD] Chargement {source_path.name}...")
    
    df = pd.read_csv(source_path, sep=';', dtype={'COM': str, 'IRIS': str}, low_memory=False)
    print(f"  [OK] {len(df)} lignes lues, {len(df.columns)} colonnes")
    
    # Extraire code commune (5 premiers caractères de IRIS ou COM)
    geo_col = 'COM' if 'COM' in df.columns else 'CODGEO'
    df['commune_code'] = df[geo_col].astype(str).str[:5]
    
    # Filtrer Guyane
    df_guyane = df[df['commune_code'].isin(COMMUNES_GUYANE)].copy()
    print(f"  [FILTER] {len(df_guyane)} lignes Guyane ({df_guyane['commune_code'].nunique()} communes)")
    
    if df_guyane.empty:
        print("  [WARN] Aucune donnée pour la Guyane !")
        return pd.DataFrame()
    
    # Agréger par commune (les données sont au niveau IRIS)
    numeric_cols = df_guyane.select_dtypes(include='number').columns.tolist()
    df_communal = df_guyane.groupby('commune_code')[numeric_cols].sum().reset_index()
    print(f"  [AGG] {len(df_communal)} communes après agrégation")
    
    return df_communal


def load_national_data(source_path: Path, year: int) -> pd.DataFrame:
    """Charge et agrège les données nationales pour les niveaux reg/dom/fh/fra."""
    print(f"  [NATIONAL] Chargement données nationales pour agrégation...")
    
    df = pd.read_csv(source_path, sep=';', dtype={'COM': str, 'IRIS': str}, low_memory=False)
    
    geo_col = 'COM' if 'COM' in df.columns else 'CODGEO'
    df['commune_code'] = df[geo_col].astype(str).str[:5]
    
    # Déterminer le département
    df['dep'] = df['commune_code'].apply(lambda x: x[:3] if x.startswith('97') else x[:2])
    
    # Mapper vers la région
    df['reg'] = df['dep'].map(DEP_TO_REG)
    
    # Agréger par commune d'abord (depuis IRIS)
    numeric_cols = df.select_dtypes(include='number').columns.tolist()
    df_com = df.groupby(['commune_code', 'dep', 'reg'])[numeric_cols].sum().reset_index()
    
    # Agréger par région
    df_reg = df_com.groupby('reg')[numeric_cols].sum().reset_index()
    
    # DOM = somme des régions DOM
    df_dom = df_reg[df_reg['reg'].isin(DOM_CODES)][numeric_cols].sum()
    
    # France Hexagonale = somme des régions métropoles
    df_fh = df_reg[df_reg['reg'].isin(FH_REGIONS)][numeric_cols].sum()
    
    # France entière = somme totale
    df_fra = df_reg[numeric_cols].sum()
    
    print(f"  [OK] {len(df_reg)} régions, agrégations DOM/FH/FR calculées")
    
    return {
        'reg': df_reg,
        'dom': df_dom,
        'fh': df_fh,
        'fra': df_fra
    }


# =============================================================================
# FONCTIONS DE CALCUL PAR THÈME
# =============================================================================

def _safe_codgeo(df: pd.DataFrame) -> pd.Series:
    """Extrait codgeo en convertissant en int si possible, sinon garde la valeur."""
    if 'commune_code' in df.columns:
        def _try_int(v):
            try:
                return int(v)
            except (ValueError, TypeError):
                return v
        return df['commune_code'].apply(_try_int)
    return df.get('reg', pd.Series([0]))


def calc_pers_sup65ans_seules(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Calcule les indicateurs pour pers_sup65ans_seules."""
    prefix_p = f"P{str(year)[2:]}_"
    
    result = pd.DataFrame()
    result['codgeo'] = _safe_codgeo(df)
    result['annee'] = year
    
    pop_5579 = df.get(f'{prefix_p}POP5579', 0)
    pop_80p = df.get(f'{prefix_p}POP80P', 0)
    result['nb_pop_65ans'] = ((15/25) * pop_5579 + pop_80p).round(2)
    
    pseul_5579 = df.get(f'{prefix_p}POP5579_PSEUL', 0)
    pseul_80p = df.get(f'{prefix_p}POP80P_PSEUL', 0)
    result['nb_pop_65ans_seule'] = ((15/25) * pseul_5579 + pseul_80p).round(2)
    
    return result


def calc_familles_mono(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Calcule les indicateurs pour familles_mono."""
    prefix_c = f"C{str(year)[2:]}_"
    
    result = pd.DataFrame()
    result['codgeo'] = _safe_codgeo(df)
    result['annee'] = year
    
    # Familles avec enfants = couples avec enfants + monoparentales
    coupaenf = df.get(f'{prefix_c}COUPAENF', 0)
    fammono = df.get(f'{prefix_c}FAMMONO', 0)
    
    # Si colonnes manquantes, essayer les alternatives
    if isinstance(coupaenf, int) and coupaenf == 0:
        coupaenf = df.get(f'{prefix_c}MENCOUPAENF', 0)
    if isinstance(fammono, int) and fammono == 0:
        fammono = df.get(f'{prefix_c}MENFAMMONO', 0)
    
    result['nb_familles_enf'] = (coupaenf + fammono).round(2)
    result['nb_familles_mono_enf'] = fammono if not isinstance(fammono, int) else pd.Series([fammono] * len(df))
    result['nb_familles_mono_enf'] = result['nb_familles_mono_enf'].round(2)
    
    return result


def calc_pop_inf3ans(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Calcule les indicateurs pour pop_inf3ans."""
    prefix_c = f"C{str(year)[2:]}_"
    
    result = pd.DataFrame()
    result['codgeo'] = _safe_codgeo(df)
    result['annee'] = year
    
    # Enfants 0-2 ans : approximation via NE24F0 (familles 0 enfants) etc.
    # En fait, cette base n'a pas directement les enfants < 3 ans
    # On peut estimer via les naissances (NE24Fx) : 
    # NE24F1 = familles 1 enfant < 25 ans, NE24F2 = 2 enfants, etc.
    # Approximation : total enfants < 25 ans * ratio estimé
    
    ne24f1 = df.get(f'{prefix_c}NE24F1', 0)
    ne24f2 = df.get(f'{prefix_c}NE24F2', 0)
    ne24f3 = df.get(f'{prefix_c}NE24F3', 0)
    ne24f4p = df.get(f'{prefix_c}NE24F4P', 0)
    
    # Total enfants < 25 ans dans les familles
    total_enfants = ne24f1 * 1 + ne24f2 * 2 + ne24f3 * 3 + ne24f4p * 4
    
    # Approximation : enfants 0-2 = ~12% du total enfants (3 ans / 25 ans)
    result['nb_enfants_0_2'] = (total_enfants * 0.12).round(2)
    
    # Enfants 0-5 = ~24% du total enfants (6 ans / 25 ans)
    result['nb_enfants_0_5'] = (total_enfants * 0.24).round(2)
    
    return result


def calc_pers_menages(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Calcule les indicateurs pour pers_menages."""
    prefix_c = f"C{str(year)[2:]}_"
    
    result = pd.DataFrame()
    result['codgeo'] = _safe_codgeo(df)
    result['annee'] = year
    
    # Nombre de ménages
    result['nb_menages'] = df.get(f'{prefix_c}MEN', 0)
    if isinstance(result['nb_menages'].iloc[0] if len(result) > 0 else 0, (int, float)):
        result['nb_menages'] = result['nb_menages'].round(2)
    
    # Personnes des ménages
    result['nb_pers_menages'] = df.get(f'{prefix_c}PMEN', 0)
    if isinstance(result['nb_pers_menages'].iloc[0] if len(result) > 0 else 0, (int, float)):
        result['nb_pers_menages'] = result['nb_pers_menages'].round(2)
    
    return result


def calc_types_menages(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Calcule les indicateurs pour types_menages."""
    prefix_c = f"C{str(year)[2:]}_"
    
    result = pd.DataFrame()
    result['codgeo'] = _safe_codgeo(df)
    result['annee'] = year
    
    # Personnes seules
    result['nb_men_seul'] = df.get(f'{prefix_c}MENPSEUL', 0)
    # Couples sans enfant
    result['nb_men_couple_senf'] = df.get(f'{prefix_c}MENCOUPSENF', 0)
    # Couples avec enfant(s)
    result['nb_men_couple_aenf'] = df.get(f'{prefix_c}MENCOUPAENF', 0)
    # Familles monoparentales
    result['nb_men_fam_mono'] = df.get(f'{prefix_c}MENFAMMONO', 0)
    
    for col in ['nb_men_seul', 'nb_men_couple_senf', 'nb_men_couple_aenf', 'nb_men_fam_mono']:
        result[col] = result[col].round(2)
    
    return result


def calc_alloc(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Calcule les indicateurs pour alloc (données CAF)."""
    result = pd.DataFrame()
    result['codgeo'] = _safe_codgeo(df)
    result['annee'] = year
    
    result['nb_alloc'] = df.get('Nombre foyers NDUR', 0)
    result['nb_foyers_rsa'] = df.get('Nombre foyers RSA', 0)
    result['nb_pers_rsa'] = df.get('Nombre personnes RSA', 0)
    result['nb_foyers_apl'] = df.get('Nombre foyers APL', 0)
    result['nb_pers_apl'] = df.get('Nombre personnes APL', 0)
    
    for col in ['nb_alloc', 'nb_foyers_rsa', 'nb_pers_rsa', 'nb_foyers_apl', 'nb_pers_apl']:
        result[col] = pd.to_numeric(result[col], errors='coerce').fillna(0).round(2)
    
    return result


# Dispatch des fonctions de calcul
CALC_FUNCTIONS = {
    "pers_sup65ans_seules": calc_pers_sup65ans_seules,
    "familles_mono": calc_familles_mono,
    "pop_inf3ans": calc_pop_inf3ans,
    "pers_menages": calc_pers_menages,
    "types_menages": calc_types_menages,
    "alloc": calc_alloc,
}


# =============================================================================
# GÉNÉRATION EXCEL PRISME
# =============================================================================

def load_caf_data(source_path: Path, year: int, calc_fn):
    """Loader spécifique pour les données CAF (format différent de l'INSEE)."""
    print(f"  [LOAD-CAF] Chargement {source_path.name}...")
    
    df = pd.read_csv(source_path, sep=';', low_memory=False)
    print(f"  [OK] {len(df)} lignes lues")
    
    # Filtrer par année (format: "2022-12")
    df['year'] = df['Date référence'].str[:4].astype(int)
    df_year = df[df['year'] == year].copy()
    print(f"  [FILTER] {len(df_year)} lignes pour l'année {year}")
    
    if df_year.empty:
        print(f"  [WARN] Aucune donnée pour l'année {year}")
        return None
    
    # Code commune
    df_year['commune_code'] = df_year['Numéro commune'].astype(str).str.zfill(5)
    
    # === Communes Guyane ===
    df_guyane = df_year[df_year['commune_code'].isin(COMMUNES_GUYANE)].copy()
    print(f"  [GUYANE] {len(df_guyane)} communes Guyane")
    data_com = calc_fn(df_guyane, year) if not df_guyane.empty else pd.DataFrame()
    
    # === Agrégation nationale ===
    # Département
    df_year['dep'] = df_year['commune_code'].apply(lambda x: x[:3] if x.startswith('97') else x[:2])
    df_year['reg'] = df_year['dep'].map(DEP_TO_REG)
    
    # Régions : agréger les colonnes numériques
    num_cols = df_year.select_dtypes(include='number').columns.tolist()
    num_cols = [c for c in num_cols if c != 'year']
    
    df_reg = df_year.dropna(subset=['reg']).groupby('reg')[num_cols].sum().reset_index()
    df_reg['commune_code'] = df_reg['reg']
    data_reg = calc_fn(df_reg, year)
    data_reg.rename(columns={'codgeo': 'reg_code'}, inplace=True)
    
    # DOM
    df_dom_data = df_reg[df_reg['reg'].isin(DOM_CODES)][num_cols].sum().to_frame().T
    df_dom_data['commune_code'] = 'DOM'
    data_dom = calc_fn(df_dom_data, year)
    
    # France Hexagonale
    df_fh_data = df_reg[df_reg['reg'].isin(FH_REGIONS)][num_cols].sum().to_frame().T
    df_fh_data['commune_code'] = '0'
    data_fh = calc_fn(df_fh_data, year)
    
    # France entière
    df_fra_data = df_reg[num_cols].sum().to_frame().T
    df_fra_data['commune_code'] = '99'
    data_fra = calc_fn(df_fra_data, year)
    
    print(f"  [OK] {len(df_reg)} régions, agrégations nationales calculées")
    
    return {
        'com': data_com,
        'reg': data_reg,
        'dom': data_dom,
        'fh': data_fh,
        'fra': data_fra
    }


def generate_excel_prisme(theme: str, year: int) -> Path:
    """Génère les fichiers Excel PRISME pour un thème donné depuis Open Data."""
    config = THEME_CONFIGS[theme]
    calc_fn = CALC_FUNCTIONS[theme]
    loader_type = config.get('loader', 'insee')
    
    print(f"\n{'='*60}")
    print(f"  {config['name']} ({theme}) — Année {year}")
    print(f"  Source : Open Data {'CAF' if loader_type == 'caf' else 'INSEE'}")
    print(f"{'='*60}")
    
    # 1. Charger les données source
    source_file = config['source_file'].format(year=year)
    source_path = INPUTS_DIR / source_file
    
    if not source_path.exists():
        print(f"  [ERROR] Fichier source manquant : {source_path}")
        print(f"  [INFO] Lancez : py download_opendata.py --source {'caf' if loader_type == 'caf' else 'couples'} --years {year}")
        return None
    
    # 2. Charger et calculer selon le loader
    if loader_type == 'caf':
        # Loader CAF spécifique
        all_levels = load_caf_data(source_path, year, calc_fn)
        if all_levels is None:
            return None
        data_com = all_levels['com']
        data_reg_calc = all_levels['reg']
        data_dom = all_levels['dom']
        data_fh = all_levels['fh']
        data_fra = all_levels['fra']
    else:
        # Loader INSEE standard
        df_com = load_and_filter_csv(source_path, year)
        if df_com.empty:
            return None
        data_com = calc_fn(df_com, year)
        
        # Données nationales
        nat_data = load_national_data(source_path, year)
        
        df_reg = nat_data['reg'].copy()
        df_reg['commune_code'] = df_reg['reg']
        data_reg_calc = calc_fn(df_reg, year)
        data_reg_calc.rename(columns={'codgeo': 'reg_code'}, inplace=True)
        
        dom_series = nat_data['dom'].to_frame().T.copy()
        dom_series['commune_code'] = 'DOM'
        data_dom = calc_fn(dom_series, year)
        
        fh_series = nat_data['fh'].to_frame().T.copy()
        fh_series['commune_code'] = '0'
        data_fh = calc_fn(fh_series, year)
        
        fra_series = nat_data['fra'].to_frame().T.copy()
        fra_series['commune_code'] = '99'
        data_fra = calc_fn(fra_series, year)
    
    variables = config['variables']
    
    # 4. Créer l'arborescence
    theme_dir = OUTPUT_DIR / config['folder'] / str(year)
    if theme_dir.exists():
        shutil.rmtree(theme_dir)
    theme_dir.mkdir(parents=True)
    
    print(f"\n  [DIR] {theme_dir}")
    
    # Structure all_data pour la génération
    all_data = {
        'com': data_com,
        'reg': data_reg_calc,
        'dom': data_dom,
        'fh': data_fh,
        'fra': data_fra
    }
    
    geo_id_cols = {
        'com': 'codgeo',
        'reg': 'reg_code' if 'reg_code' in data_reg_calc.columns else 'codgeo',
        'dom': 'codgeo',
        'fh': 'codgeo',
        'fra': 'codgeo'
    }
    
    # 5. Fichiers Excel par niveau géographique
    for geo_key, folder_name in GEO_LEVELS.items():
        sub_dir = theme_dir / folder_name
        sub_dir.mkdir(exist_ok=True)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(geo_key)
        
        df_level = all_data[geo_key]
        id_col = geo_id_cols[geo_key]
        
        # En-têtes
        headers = [geo_key, 'annee'] + variables
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = ORANGE_FILL
        
        # Données
        for row_idx, (_, row) in enumerate(df_level.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row.get(id_col, ''))
            ws.cell(row=row_idx, column=2, value=year)
            
            for i, var in enumerate(variables):
                val = row.get(var)
                cell = ws.cell(row=row_idx, column=3 + i)
                if pd.notna(val):
                    cell.value = val
                    cell.fill = ORANGE_FILL
        
        # Largeur colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        for i in range(len(variables)):
            ws.column_dimensions[chr(67 + i)].width = 20
        
        filepath = sub_dir / f"{config['excel_name']}.xlsx"
        wb.save(filepath)
        print(f"  [OK] {folder_name}/{config['excel_name']}.xlsx ({len(df_level)} lignes)")
    
    # 6. Fichier consolidé
    print(f"\n  [CONSOLIDATED] Création fichier consolidé...")
    wb_cons = openpyxl.Workbook()
    wb_cons.remove(wb_cons.active)
    
    for geo_key in GEO_LEVELS.keys():
        ws = wb_cons.create_sheet(geo_key)
        df_level = all_data[geo_key]
        id_col = geo_id_cols[geo_key]
        
        headers = [geo_key, 'annee'] + variables
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = ORANGE_FILL
        
        for row_idx, (_, row) in enumerate(df_level.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row.get(id_col, ''))
            ws.cell(row=row_idx, column=2, value=year)
            for i, var in enumerate(variables):
                val = row.get(var)
                cell = ws.cell(row=row_idx, column=3 + i)
                if pd.notna(val):
                    cell.value = val
                    cell.fill = ORANGE_FILL
    
    cons_path = theme_dir / f"{config['excel_name']}_consolidated_{year}.xlsx"
    wb_cons.save(cons_path)
    print(f"  [OK] {config['excel_name']}_consolidated_{year}.xlsx")
    
    # 7. ZIP
    zip_name = f"{theme}_opendata_{year}"
    zip_path = shutil.make_archive(
        str(OUTPUT_DIR / zip_name),
        'zip',
        str(OUTPUT_DIR / config['folder']),
        str(year)
    )
    print(f"\n  [ZIP] {zip_path}")
    
    # 8. Résumé
    print(f"\n{'='*60}")
    print(f"  Résumé — {config['name']} {year}")
    print(f"{'='*60}")
    print(f"  Communes Guyane  : {len(data_com)}")
    print(f"  Régions          : {len(data_reg_calc)}")
    print(f"  Variables        : {', '.join(variables)}")
    print(f"  Fichiers générés : 6 (5 niveaux + 1 consolidé)")
    print(f"  ZIP              : {zip_name}.zip")
    
    return theme_dir


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="PRISME — Générateur Open Data Unifié")
    parser.add_argument('--theme', type=str, required=True,
                        help=f"Thème à générer: {', '.join(THEME_CONFIGS.keys())}, ou 'all'")
    parser.add_argument('--year', type=int, default=2022, help="Année (défaut: 2022)")
    parser.add_argument('--list', action='store_true', help="Lister les thèmes disponibles")
    
    args = parser.parse_args()
    
    if args.list:
        print("\nThèmes disponibles:")
        for tid, cfg in THEME_CONFIGS.items():
            print(f"  {tid:30s} — {cfg['description']}")
        return
    
    if args.theme == 'all':
        themes = list(THEME_CONFIGS.keys())
    elif args.theme in THEME_CONFIGS:
        themes = [args.theme]
    else:
        print(f"[ERROR] Thème inconnu: {args.theme}")
        print(f"Thèmes valides: {', '.join(THEME_CONFIGS.keys())}, all")
        sys.exit(1)
    
    results = []
    for theme in themes:
        result = generate_excel_prisme(theme, args.year)
        results.append((theme, result))
    
    print(f"\n\n{'='*60}")
    print(f"  RÉCAPITULATIF")
    print(f"{'='*60}")
    for theme, result in results:
        status = "✓" if result else "✗"
        print(f"  [{status}] {theme}")


if __name__ == "__main__":
    main()
