#!/usr/bin/env python3
"""
PRISME Engine - Générateur Miroir SharePoint
Génère l'arborescence complète : Thème > Sous-thème > Année > Niveau Géo > fichier.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import re
import warnings

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = Path(__file__).parent
CSV_SOURCES_DIR = BASE_DIR / "csv_sources"
OUTPUT_DIR = BASE_DIR / "output_sharepoint"

ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Niveaux géographiques (correspondant aux onglets actuels)
GEO_LEVELS = {
    'Commune': 'com',
    'DOM': 'dom',
    'France entière': 'fra',
    'France Hexagonale': 'fh',
    'Région': 'reg'
}

# Mapping des régions
REGION_MAPPING = {
    "Île-de-France": 11, "Ile-de-France": 11,
    "Centre-Val de Loire": 24,
    "Bourgogne-Franche-Comté": 27,
    "Normandie": 28,
    "Hauts-de-France": 32,
    "Grand Est": 44,
    "Pays de la Loire": 52,
    "Bretagne": 53,
    "Nouvelle-Aquitaine": 75,
    "Occitanie": 76,
    "Auvergne-Rhône-Alpes": 84,
    "Provence-Alpes-Côte d'Azur": 93,
    "Corse": 94,
    "Guadeloupe": 1,
    "Martinique": 2,
    "Guyane": 3,
    "La Réunion": 4,
    "Mayotte": 6
}

COMMUNES_GUYANE = [
    97301, 97302, 97303, 97304, 97305, 97306, 97307, 97308, 97309, 97310,
    97311, 97312, 97313, 97314, 97352, 97353, 97356, 97357, 97358, 97360,
    97361, 97362
]

REGION_ORDER = [11, 24, 27, 28, 32, 44, 52, 53, 75, 76, 84, 93, 94, 1, 2, 3, 4, 6]

# Configuration des datasets
DATASET_CONFIGS = {
    'educ': {
        'name': 'Education',
        'theme': 'Population et condition de vie',
        'sub_theme': 'Education',
        'folder_name': 'Educ',
        'variables': ['pop_6_16', 'nb_non_sco', 'pop_15_64', 'nb_peu_dipl'],
        'csv_mapping': {
            'pop_6_16': 'Pop_6-16ans',
            'nb_non_sco': 'nb_non_scol',
            'pop_15_64': 'Pop_15-64ans',
            'nb_peu_dipl': 'Nb_peu_dipl'
        }
    }
}

# ============================================================================
# MOCA-O CSV PARSER
# ============================================================================

def parse_moca_csv(filepath):
    """Parse un fichier CSV au format MOCA-O."""
    result = {'com': [], 'reg': [], 'dom': [], 'fh': [], 'fra': []}
    
    try:
        with open(filepath, 'r', encoding='cp1252', errors='replace') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Erreur lecture {filepath}: {e}")
        return {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) for k in result}

    for line in lines:
        if not line.strip():
            continue
        
        parts = line.strip().split(';')
        if len(parts) < 3:
            continue
        
        try:
            annee = int(parts[0])
            if not (2000 <= annee <= 2030):
                continue
        except:
            continue
        
        valeur = None
        for p in reversed(parts):
            try:
                valeur = float(p.replace(',', '.'))
                break
            except:
                continue
        
        if valeur is None:
            continue
        
        line_lower = line.lower()
        
        # Communes Guyane
        match_com = re.search(r'(973\d{2})', line)
        if match_com:
            code = int(match_com.group(1))
            if code in COMMUNES_GUYANE:
                result['com'].append({'annee': annee, 'codgeo': code, 'valeur': valeur})
            continue
        
        # France entière
        if 'france entiere' in line_lower or 'france (y compris mayotte)' in line_lower or 'france entière' in line_lower:
            result['fra'].append({'annee': annee, 'codgeo': 99, 'valeur': valeur})
            continue
        
        # France hexagonale
        if 'france metropolitaine' in line_lower or 'france hexagonale' in line_lower or 'france métropolitaine' in line_lower:
            result['fh'].append({'annee': annee, 'codgeo': 0, 'valeur': valeur})
            continue
        
        # DOM
        if "departements d'outre" in line_lower or "départements d'outre" in line_lower:
            result['dom'].append({'annee': annee, 'codgeo': 'DOM', 'valeur': valeur})
            continue
        
        # Régions
        for reg_name, reg_code in REGION_MAPPING.items():
            if reg_name.lower() in line_lower:
                result['reg'].append({'annee': annee, 'codgeo': reg_code, 'valeur': valeur})
                break

    dfs = {}
    for k, v in result.items():
        if v:
            dfs[k] = pd.DataFrame(v).drop_duplicates(subset=['annee', 'codgeo'])
        else:
            dfs[k] = pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])
    
    return dfs


def find_csv_file(pattern, search_dir=None):
    """Trouve un fichier CSV correspondant au pattern."""
    if search_dir is None:
        search_dir = CSV_SOURCES_DIR
    candidates = list(Path(search_dir).glob(f"*{pattern}*.csv"))
    if candidates:
        return candidates[0]
    return None


# ============================================================================
# EXCEL GENERATOR PER GEO LEVEL
# ============================================================================

def generate_single_geo_excel(dataset_name, year, geo_level, geo_code, csv_data, variables, output_path):
    """Génère un fichier Excel pour un seul niveau géographique."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = geo_code
    
    # Récupérer les données pour ce niveau géo
    if geo_code == 'com':
        id_col = 'com'
        rows_data = [{'id': c, 'annee': year} for c in COMMUNES_GUYANE]
    elif geo_code == 'reg':
        id_col = 'reg'
        rows_data = [{'id': c, 'annee': year} for c in REGION_ORDER]
    elif geo_code == 'dom':
        id_col = 'dom'
        rows_data = [{'id': 'DOM', 'annee': year}]
    elif geo_code == 'fh':
        id_col = 'fh'
        rows_data = [{'id': 0, 'annee': year}]
    elif geo_code == 'fra':
        id_col = 'fra'
        rows_data = [{'id': 99, 'annee': year}]
    else:
        return None
    
    # Remplir les données
    for var_name in variables:
        parsed = csv_data.get(var_name, {})
        df = parsed.get(geo_code, pd.DataFrame())
        
        if not df.empty:
            df_year = df[df['annee'] == year]
            for row in rows_data:
                match = df_year[df_year['codgeo'] == row['id']]
                if not match.empty:
                    row[var_name] = round(match.iloc[0]['valeur'], 2)
    
    # Écrire les en-têtes
    headers = [id_col, 'annee'] + variables
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = ORANGE_FILL
    
    # Écrire les données
    for row_idx, row_dict in enumerate(rows_data, 2):
        ws.cell(row=row_idx, column=1, value=row_dict.get('id'))
        ws.cell(row=row_idx, column=2, value=row_dict.get('annee'))
        
        for i, var in enumerate(variables):
            val = row_dict.get(var)
            cell = ws.cell(row=row_idx, column=3 + i, value=val)
            if val is not None:
                cell.fill = ORANGE_FILL
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    
    # Sauvegarder
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================================
# MAIN SHAREPOINT MIRROR GENERATOR
# ============================================================================

def generate_sharepoint_mirror(dataset_name, years=None):
    """Génère l'arborescence complète miroir SharePoint."""
    
    if dataset_name not in DATASET_CONFIGS:
        print(f"Dataset inconnu: {dataset_name}")
        return
    
    config = DATASET_CONFIGS[dataset_name]
    variables = config['variables']
    csv_mapping = config['csv_mapping']
    
    theme = config['theme']
    sub_theme = config['sub_theme']
    folder_name = config['folder_name']
    
    print(f"\n{'='*60}")
    print(f"PRISME - Génération Miroir SharePoint")
    print(f"Dataset: {config['name']}")
    print(f"{'='*60}")
    
    # Charger les données CSV
    print("\n[1/3] Chargement des sources CSV...")
    csv_data = {}
    for var_name, csv_pattern in csv_mapping.items():
        csv_file = find_csv_file(csv_pattern)
        if csv_file:
            csv_data[var_name] = parse_moca_csv(csv_file)
            print(f"  [OK] {var_name} -> {csv_file.name}")
        else:
            print(f"  [WARN] {var_name} -> Fichier non trouvé")
            csv_data[var_name] = {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) 
                                  for k in ['com', 'reg', 'dom', 'fh', 'fra']}
    
    # Détecter les années disponibles
    if years is None:
        all_years = set()
        for var_data in csv_data.values():
            for geo_df in var_data.values():
                if not geo_df.empty and 'annee' in geo_df.columns:
                    all_years.update(geo_df['annee'].unique())
        years = sorted(all_years)
    
    print(f"\n[2/3] Années détectées: {years}")
    
    # Générer l'arborescence
    print(f"\n[3/3] Génération de l'arborescence...")
    
    generated_files = []
    
    for year in years:
        print(f"\n  Année {year}:")
        
        for geo_level_name, geo_code in GEO_LEVELS.items():
            # Construire le chemin: Theme/SubTheme/FolderName/Year/GeoLevel/file.xlsx
            output_path = OUTPUT_DIR / theme / sub_theme / folder_name / str(year) / geo_level_name / f"{dataset_name}.xlsx"
            
            result = generate_single_geo_excel(
                dataset_name, year, geo_level_name, geo_code,
                csv_data, variables, output_path
            )
            
            if result:
                print(f"    [OK] {geo_level_name}/{dataset_name}.xlsx")
                generated_files.append(result)
    
    print(f"\n{'='*60}")
    print(f"Génération terminée!")
    print(f"Fichiers générés: {len(generated_files)}")
    print(f"Dossier de sortie: {OUTPUT_DIR}")
    print(f"{'='*60}")
    
    return generated_files


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import sys
    
    # Test avec Education pour les années 2020-2022
    years_to_generate = [2020, 2021, 2022]
    
    if len(sys.argv) > 1:
        years_to_generate = [int(y) for y in sys.argv[1].split(',')]
    
    generate_sharepoint_mirror('educ', years_to_generate)
