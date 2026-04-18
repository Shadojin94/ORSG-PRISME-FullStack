"""
convert_baac_client_xlsx.py
---------------------------
Hyoga-BAAC - Extraction des indicateurs BAAC depuis le bilan annuel ONISR
(fichier Excel pre-formate envoye par le client ORSG).

Contexte
========
Le client ORSG nous a envoye 2 fichiers Excel :
  - livrables_17avril/Retour client/Accidents de la circulation_2023.xlsx
  - livrables_17avril/Retour client/Accidents de la circulation_2024.xlsx
Ces fichiers sont le *bilan statistique annuel ONISR* (94+ feuilles mise-en-page
rapport) et non pas un export brut BAAC. Ils contiennent la meme donnee que les
fichiers nationaux publies sur data.gouv, mais en format tableaux de synthese.

Verification faite (avril 2026) : les totaux Guyane du bilan client sont
strictement egaux a ceux calcules depuis les CSV deja presents dans
Backend/inputs/opendata/baac_guyane/annees_YYYY/ :

                 2023 client | CSV existant    | 2024 client | CSV existant
  Accidents      600         | 600             | 631         | 631
  Tues           34          | 34              | 34          | 34
  Blesses hospi. 191         | 191             | 125         | 125
  Blesses legers 620         | 620             | 720         | 720

=> Aucun remplacement de CSV necessaire. Les CSV existants sont deja la source
officielle BAAC/ONISR conforme au bilan client.

Usage
=====
Le script ci-dessous extrait, depuis le xlsx client, les indicateurs agreges
Guyane (dep 973) pour verification/archivage, et ecrit un CSV de synthese :
  Backend/inputs/opendata/baac_guyane/verification_bilan_onisr.csv

Il peut aussi re-executer le recalcul depuis les CSV existants pour confirmer
la concordance.

  python Backend/scripts/convert_baac_client_xlsx.py --verify
  python Backend/scripts/convert_baac_client_xlsx.py --extract-bilan
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
BACKEND = HERE.parent
REPO = BACKEND.parent

CLIENT_DIR = REPO / "livrables_17avril" / "Retour client"
BAAC_DIR = BACKEND / "inputs" / "opendata" / "baac_guyane"


def _find_outre_mer_sheet(wb: openpyxl.Workbook) -> str:
    """Le bilan ONISR place le tableau 'Outre-Mer' en Page 12."""
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True, max_row=5):
            for cell in row:
                if cell and isinstance(cell, str) and "Outre-Mer" in cell:
                    return name
    raise RuntimeError("Feuille Outre-Mer introuvable dans le bilan client")


def extract_bilan(year: int) -> dict:
    """Extrait la ligne Guyane du tableau Outre-Mer Page 12."""
    path = CLIENT_DIR / f"Accidents de la circulation_{year}.xlsx"
    if not path.exists():
        raise FileNotFoundError(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = _find_outre_mer_sheet(wb)
    ws = wb[sheet]
    # Layout observe (2023/2024) : premiere colonne vide, donnees a partir de col B
    #   col 1..3 = [dep_code, '-', nom]
    #   col 4..10 = [Accidents, Dont mortels, Dont graves, Tues,
    #                Blesses hospi, Blesses legers, Total blesses]
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        # Trouve l'offset du code departement '973'
        offset = None
        for idx, v in enumerate(vals):
            if v is not None and str(v).strip() == "973":
                offset = idx
                break
        if offset is None:
            continue
        def _i(x):
            return int(x) if x is not None and x != "" else None
        out = {
            "annee": year,
            "dep": "973",
            "nom": str(vals[offset + 2]).strip() if vals[offset + 2] else "Guyane",
            "nb_acci": _i(vals[offset + 3]),
            "dont_mortels": _i(vals[offset + 4]),
            "dont_graves": _i(vals[offset + 5]),
            "nb_morts": _i(vals[offset + 6]),
            "nb_blesses_hospi": _i(vals[offset + 7]),
            "nb_blesses_legers": _i(vals[offset + 8]),
            "nb_blesses_total": _i(vals[offset + 9]),
        }
        wb.close()
        return out
    wb.close()
    raise RuntimeError(f"Ligne Guyane (973) introuvable pour {year}")


def recompute_from_csv(year: int) -> dict:
    """Recalcule les indicateurs depuis les CSV BAAC existants (Guyane)."""
    base = BAAC_DIR / f"annees_{year}"
    caract_path = base / f"caract_{year}.csv"
    usagers_path = base / f"usagers_{year}.csv"
    caract = pd.read_csv(caract_path, sep=";", dtype=str)
    usagers = pd.read_csv(usagers_path, sep=";", dtype=str)
    acc_c = "Num_Acc" if "Num_Acc" in caract.columns else "Accident_Id"
    acc_u = "Num_Acc" if "Num_Acc" in usagers.columns else "Accident_Id"
    usagers["grav"] = usagers["grav"].astype(str).str.strip()
    merged = usagers.merge(
        caract[[acc_c]].drop_duplicates().rename(columns={acc_c: "acc"}),
        left_on=acc_u, right_on="acc", how="inner",
    )
    return {
        "annee": year,
        "dep": "973",
        "nom": "Guyane",
        "nb_acci": caract[acc_c].nunique(),
        "nb_morts": int((merged["grav"] == "2").sum()),
        "nb_blesses_hospi": int((merged["grav"] == "3").sum()),
        "nb_blesses_legers": int((merged["grav"] == "4").sum()),
        "nb_blesses_total": int(merged["grav"].isin(["3", "4"]).sum()),
    }


def verify(years=(2023, 2024)) -> int:
    """Compare les indicateurs client vs CSV existants. Exit code != 0 si divergence."""
    rc = 0
    rows = []
    for year in years:
        bilan = extract_bilan(year)
        csv_stats = recompute_from_csv(year)
        keys = ("nb_acci", "nb_morts", "nb_blesses_hospi", "nb_blesses_legers", "nb_blesses_total")
        ok = all(bilan.get(k) == csv_stats.get(k) for k in keys)
        rc = rc or (0 if ok else 1)
        print(f"=== {year} Guyane ===")
        print(f"  {'':20s}  client  | csv      | match")
        for k in keys:
            b = bilan.get(k); c = csv_stats.get(k)
            mark = "OK" if b == c else "DIFF"
            print(f"  {k:20s}  {str(b):6s}  | {str(c):6s}   | {mark}")
        rows.append({"source": "bilan_client", **bilan})
        rows.append({"source": "csv_existant", **csv_stats})
    # Ecrit un CSV de synthese pour tracabilite
    out_path = BAAC_DIR / "verification_bilan_onisr.csv"
    if rows:
        fieldnames = [
            "source", "annee", "dep", "nom", "nb_acci", "nb_morts",
            "nb_blesses_hospi", "nb_blesses_legers", "nb_blesses_total",
            "dont_mortels", "dont_graves",
        ]
        with out_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
            w.writeheader()
            for r in rows:
                w.writerow({k: r.get(k, "") for k in fieldnames})
        print(f"\n[OK] Synthese ecrite : {out_path}")
    return rc


def main() -> int:
    parser = argparse.ArgumentParser(description="BAAC client xlsx -> indicateurs Guyane")
    parser.add_argument("--years", nargs="+", type=int, default=[2023, 2024])
    parser.add_argument("--verify", action="store_true", help="Compare bilan client vs CSV existants")
    parser.add_argument("--extract-bilan", action="store_true", help="Affiche juste les valeurs du bilan client")
    args = parser.parse_args()

    if args.extract_bilan:
        for y in args.years:
            print(extract_bilan(y))
        return 0
    # Par defaut : verification complete
    return verify(tuple(args.years))


if __name__ == "__main__":
    sys.exit(main())
