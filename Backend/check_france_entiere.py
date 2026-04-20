import openpyxl
from pathlib import Path

# Find the France entière folder
base = Path("output/educ_2020_test/Educ/2020")
folders = list(base.iterdir())
print(f"Found folders: {[f.name for f in folders if f.is_dir()]}")

france_folder = None
for f in folders:
    if "France" in f.name and "Hexagonale" not in f.name and f.is_dir():
        france_folder = f
        break

if not france_folder:
    print("ERROR: France entière folder not found!")
    exit(1)

print(f"\nChecking: {france_folder}")

excel_file = france_folder / "educ.xlsx"
if not excel_file.exists():
    print(f"ERROR: Excel file not found at {excel_file}")
    exit(1)

print(f"Loading: {excel_file}")
wb = openpyxl.load_workbook(excel_file)
print(f"Sheets: {wb.sheetnames}")

ws = wb['fra']
print(f"\nSheet 'fra': {ws.max_row} rows × {ws.max_column} cols")

# Print first 3 rows
for i in range(min(3, ws.max_row)):
    row_values = [cell.value for cell in ws[i+1]]
    print(f"Row {i+1}: {row_values}")
