import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from Backend.qa_excel_check import check_generated_artifact


def build_workbook(path, fill=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "com"
    sheet.append(["code", "valeur"])
    sheet.append(["97301", 42])
    if fill is not None:
        sheet["B2"].fill = fill
    workbook.save(path)


class ExcelQaGuardTest(unittest.TestCase):
    def test_xlsx_without_colored_fill_is_accepted(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "clean.xlsx"
            build_workbook(path, PatternFill(fill_type="solid", fgColor="FFFFFFFF"))

            result = check_generated_artifact(path)

            self.assertTrue(result["ok"], result["issues"])

    def test_colored_fill_is_rejected_with_cell_location(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "colored.xlsx"
            build_workbook(path, PatternFill(fill_type="solid", fgColor="FFFF0000"))

            result = check_generated_artifact(path)

            self.assertFalse(result["ok"])
            message = " ".join(result["issues"])
            self.assertIn("Remplissage colore PRESENT", message)
            self.assertIn("B2=rgb:FFFF0000", message)

    def test_theme_zero_fill_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "theme-zero.xlsx"
            build_workbook(
                path,
                PatternFill(fill_type="solid", fgColor=Color(theme=0)),
            )

            result = check_generated_artifact(path)

            self.assertFalse(result["ok"])
            self.assertIn("B2=theme:0", " ".join(result["issues"]))

    def test_theme_one_without_tint_is_accepted_as_white(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "theme-one.xlsx"
            build_workbook(
                path,
                PatternFill(fill_type="solid", fgColor=Color(theme=1)),
            )

            result = check_generated_artifact(path)

            self.assertTrue(result["ok"], result["issues"])

    def test_theme_one_with_tint_is_rejected_fail_closed(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "theme-tinted.xlsx"
            build_workbook(
                path,
                PatternFill(fill_type="solid", fgColor=Color(theme=1, tint=0.4)),
            )

            result = check_generated_artifact(path)

            self.assertFalse(result["ok"])
            self.assertIn("tint=0.4", " ".join(result["issues"]))

    def test_zip_rejects_a_colored_workbook(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            clean_path = tmp_path / "clean.xlsx"
            colored_path = tmp_path / "colored.xlsx"
            zip_path = tmp_path / "pack.zip"
            build_workbook(clean_path)
            build_workbook(
                colored_path,
                PatternFill(fill_type="solid", fgColor="FF00FF00"),
            )
            with zipfile.ZipFile(zip_path, "w") as archive:
                archive.write(clean_path, "Commune/clean.xlsx")
                archive.write(colored_path, "Region/colored.xlsx")

            result = check_generated_artifact(zip_path)

            self.assertFalse(result["ok"])
            self.assertIn("Region/colored.xlsx", " ".join(result["issues"]))


if __name__ == "__main__":
    unittest.main()
