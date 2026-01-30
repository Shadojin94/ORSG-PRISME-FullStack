#!/usr/bin/env python3
"""
PRISME Engine - Moteur de génération Excel pour ORSG
Version: 4.0 - Architecture config-driven
Utilise themes_config.json pour la configuration des datasets
"""

import json
import re
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import warnings
import tempfile

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION & PATHS
# ============================================================================

BASE_DIR = Path(__file__).parent.parent
CSV_SOURCES_DIR = BASE_DIR / "csv_sources"
OUTPUT_DIR = BASE_DIR / "output"
CONFIG_FILE = BASE_DIR / "themes_config.json"

OUTPUT_DIR.mkdir(exist_ok=True)

ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# ============================================================================
# CHARGER LA CONFIGURATION DEPUIS themes_config.json
# ============================================================================

def load_themes_config():
    """Charge la configuration des themes depuis themes_config.json."""
    if not CONFIG_FILE.exists():
        print(f"[WARN] Config file not found: {CONFIG_FILE}")
        return {"datasets": {}, "themeTree": [], "geoLevels": {}}
    
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"[ERROR] Failed to load config: {e}")
        return {"datasets": {}, "themeTree": [], "geoLevels": {}}

# Charger la config au démarrage
THEMES_CONFIG = load_themes_config()

def reload_config():
    """Recharge la configuration (utile après mise à jour)."""
    global THEMES_CONFIG
    THEMES_CONFIG = load_themes_config()
    return THEMES_CONFIG

def get_dataset_config(dataset_id):
    """Récupère la configuration d'un dataset depuis themes_config.json."""
    datasets = THEMES_CONFIG.get('datasets', {})
    return datasets.get(dataset_id)

def get_available_datasets():
    """Retourne la liste des datasets disponibles."""
    return list(THEMES_CONFIG.get('datasets', {}).keys())

def get_theme_tree():
    """Retourne l'arborescence des thèmes."""
    return THEMES_CONFIG.get('themeTree', [])

# ============================================================================
# REGION & COMMUNE CONSTANTS
# ============================================================================

REGION_ORDER = [11, 24, 27, 28, 32, 44, 52, 53, 75, 76, 84, 93, 94, 1, 2, 3, 4, 6]

REGION_MAPPING = {
    "Île-de-France": 11, "Ile-de-France": 11, "île-de-France": 11,
    "Centre-Val de Loire": 24,
    "Bourgogne-Franche-Comté": 27, "Bourgogne-Franche-ComtÃ©": 27,
    "Normandie": 28,
    "Hauts-de-France": 32,
    "Grand Est": 44,
    "Pays de la Loire": 52,
    "Bretagne": 53,
    "Nouvelle-Aquitaine": 75,
    "Occitanie": 76,
    "Auvergne-Rhône-Alpes": 84, "Auvergne-RhÃ´ne-Alpes": 84,
    "Provence-Alpes-Côte d'Azur": 93,
    "Corse": 94,
    "Guadeloupe": 1,
    "Martinique": 2,
    "Guyane": 3,
    "La Réunion": 4, "Réunion": 4,
    "Mayotte": 6
}

COMMUNES_GUYANE = [
    97301, 97302, 97303, 97304, 97305, 97306, 97307, 97308, 97309, 97310,
    97311, 97312, 97313, 97314, 97352, 97353, 97356, 97357, 97358, 97360,
    97361, 97362
]

# Mapping des niveaux géo vers les noms de dossiers
GEO_FOLDER_MAPPING = {
    'com': 'Commune',
    'reg': 'Région',
    'dom': 'DOM',
    'fh': 'France Hexagonale',
    'fra': 'France entière'
}

GEO_ID_COLS = {
    'com': 'com',
    'reg': 'reg',
    'dom': 'dom',
    'fh': 'fh',
    'fra': 'fra'
}

# ============================================================================
# MOCA-O CSV PARSER (Legacy Matrix Format)
# ============================================================================

def parse_moca_csv(filepath):
    """Parse un fichier CSV au format MOCA-O classique (matrice)."""
    result = {'com': [], 'reg': [], 'dom': [], 'fh': [], 'fra': []}
    
    try:
        with open(filepath, 'r', encoding='cp1252', errors='replace') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Erreur lecture {filepath}: {e}")
        return {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) for k in result}

    for line in lines:
        if not line.strip():
            continue
        
        parts = line.strip().split(';')
        if len(parts) < 3:
            continue
        
        try:
            annee = int(parts[0])
            if not (2000 <= annee <= 2030):
                continue
        except:
            continue
        
        valeur = None
        for p in reversed(parts):
            try:
                valeur = float(p.replace(',', '.'))
                break
            except:
                continue
        
        if valeur is None:
            continue
        
        line_lower = line.lower()
        
        match_com = re.search(r'(973\d{2})', line)
        if match_com:
            code = int(match_com.group(1))
            if code in COMMUNES_GUYANE:
                result['com'].append({'annee': annee, 'codgeo': code, 'valeur': valeur})
            continue
        
        if 'france entiere' in line_lower or 'france (y compris mayotte)' in line_lower or 'france entière' in line_lower:
            result['fra'].append({'annee': annee, 'codgeo': 99, 'valeur': valeur})
            continue
        
        if 'france metropolitaine' in line_lower or 'france hexagonale' in line_lower or 'france métropolitaine' in line_lower:
            result['fh'].append({'annee': annee, 'codgeo': 0, 'valeur': valeur})
            continue
        
        if "departements d'outre" in line_lower or "départements d'outre" in line_lower:
            result['dom'].append({'annee': annee, 'codgeo': 'DOM', 'valeur': valeur})
            continue
        
        for reg_name, reg_code in REGION_MAPPING.items():
            if reg_name.lower() in line_lower:
                result['reg'].append({'annee': annee, 'codgeo': reg_code, 'valeur': valeur})
                break

    dfs = {}
    for k, v in result.items():
        if v:
            dfs[k] = pd.DataFrame(v).drop_duplicates(subset=['annee', 'codgeo'])
        else:
            dfs[k] = pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])
    
    return dfs

# ============================================================================
# LONG FORMAT PARSER (Transposed)
# ============================================================================

def parse_long_format_csv(filepath):
    """Parse un fichier CSV au format Long (une ligne = une entité)."""
    result = {'com': [], 'reg': [], 'dom': [], 'fh': [], 'fra': []}
    
    try:
        with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Erreur lecture {filepath}: {e}")
        return {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) for k in result}

    for line in lines:
        if not line.strip():
            continue
        
        parts = line.strip().split(';')
        if len(parts) < 3:
            continue
            
        try:
            annee = int(parts[0])
            if not (2000 <= annee <= 2030):
                continue
        except:
            continue

        try:
            valeur = float(parts[-1].replace(',', '.'))
        except:
            continue
            
        geo_str = parts[-2].strip()
        geo_lower = geo_str.lower()
        
        match_com = re.search(r'(973\d{2})', geo_str)
        if match_com:
            code = int(match_com.group(1))
            if code in COMMUNES_GUYANE:
                result['com'].append({'annee': annee, 'codgeo': code, 'valeur': valeur})
            continue
            
        if 'france entiere' in geo_lower or 'france entière' in geo_lower:
            result['fra'].append({'annee': annee, 'codgeo': 99, 'valeur': valeur})
            continue

        if 'france metropolitaine' in geo_lower or 'france hexagonale' in geo_lower or 'france métropolitaine' in geo_lower:
            result['fh'].append({'annee': annee, 'codgeo': 0, 'valeur': valeur})
            continue
        
        if "departements d'outre" in geo_lower or "départements d'outre" in geo_lower:
            result['dom'].append({'annee': annee, 'codgeo': 'DOM', 'valeur': valeur})
            continue
            
        for reg_name, reg_code in REGION_MAPPING.items():
            if reg_name.lower() in geo_lower:
                result['reg'].append({'annee': annee, 'codgeo': reg_code, 'valeur': valeur})
                break
                
    dfs = {}
    for k, v in result.items():
        if v:
            dfs[k] = pd.DataFrame(v).drop_duplicates(subset=['annee', 'codgeo'])
        else:
            dfs[k] = pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])
            
    return dfs


# ============================================================================
# TABULAR FORMAT PARSER (OpenData style)
# ============================================================================

def parse_tabular_csv(filepath, value_column=2, year_column=0, geo_column=1):
    """Parse un fichier CSV au format tabulaire (OpenData).

    Format configurable:
    - year_column: colonne contenant 'Année#YYYY' (défaut: 0)
    - geo_column: colonne contenant '97XXX - NomCommune' (défaut: 1)
    - value_column: colonne contenant la valeur (défaut: 2)
    """
    result = {'com': [], 'reg': [], 'dom': [], 'fh': [], 'fra': []}

    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Erreur lecture {filepath}: {e}")
        return {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) for k in result}

    max_col = max(value_column, year_column, geo_column)

    for line in lines[1:]:  # Skip header
        if not line.strip():
            continue

        parts = line.strip().split(';')
        if len(parts) <= max_col:
            continue

        # Extract year from specified column
        year_match = re.search(r'(\d{4})', parts[year_column])
        if not year_match:
            continue
        annee = int(year_match.group(1))
        if not (2000 <= annee <= 2030):
            continue

        # Extract value
        try:
            valeur = float(parts[value_column].replace(',', '.'))
        except:
            continue

        geo_str = parts[geo_column] if len(parts) > geo_column else ''
        geo_lower = geo_str.lower()

        # Check for commune (973XX)
        match_com = re.search(r'(973\d{2})', geo_str)
        if match_com:
            code = int(match_com.group(1))
            if code in COMMUNES_GUYANE:
                result['com'].append({'annee': annee, 'codgeo': code, 'valeur': valeur})
            continue

        # France entière
        if 'france entiere' in geo_lower or 'france entière' in geo_lower or 'france (y compris' in geo_lower:
            result['fra'].append({'annee': annee, 'codgeo': 99, 'valeur': valeur})
            continue

        # France hexagonale
        if 'france metropolitaine' in geo_lower or 'france hexagonale' in geo_lower or 'france métropolitaine' in geo_lower:
            result['fh'].append({'annee': annee, 'codgeo': 0, 'valeur': valeur})
            continue

        # DOM
        if "departements d'outre" in geo_lower or "départements d'outre" in geo_lower or "dom -" in geo_lower:
            result['dom'].append({'annee': annee, 'codgeo': 'DOM', 'valeur': valeur})
            continue

        # Régions
        for reg_name, reg_code in REGION_MAPPING.items():
            if reg_name.lower() in geo_lower:
                result['reg'].append({'annee': annee, 'codgeo': reg_code, 'valeur': valeur})
                break

    dfs = {}
    for k, v in result.items():
        if v:
            dfs[k] = pd.DataFrame(v).drop_duplicates(subset=['annee', 'codgeo'])
        else:
            dfs[k] = pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])

    return dfs


# ============================================================================
# MOCA FILTER PARSER (MOCA with column filter)
# ============================================================================

def parse_moca_filter_csv(filepath, filter_column, filter_value, year_column=3, geo_column=5, value_column=6):
    """Parse un fichier MOCA avec filtre sur une colonne spécifique.

    Format configurable (colonnes séparées par ;):
    - year_column: colonne année (défaut: 3)
    - filter_column: colonne à filtrer
    - geo_column: colonne code géo (défaut: 5)
    - value_column: colonne valeur (défaut: 6)
    """
    result = {'com': [], 'reg': [], 'dom': [], 'fh': [], 'fra': []}

    try:
        with open(filepath, 'r', encoding='cp1252', errors='replace') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Erreur lecture {filepath}: {e}")
        return {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur']) for k in result}

    max_col = max(filter_column, year_column, geo_column, value_column)

    for line in lines:
        if not line.strip():
            continue

        parts = line.strip().split(';')
        if len(parts) <= max_col:
            continue

        # Check filter match
        filter_cell = parts[filter_column].strip()
        if filter_value.lower() not in filter_cell.lower():
            continue

        # Extract year
        try:
            annee = int(parts[year_column])
            if not (2000 <= annee <= 2030):
                continue
        except:
            continue

        # Extract value
        try:
            valeur = float(parts[value_column].replace(',', '.'))
        except:
            continue

        geo_str = parts[geo_column] if len(parts) > geo_column else ''
        geo_lower = geo_str.lower()

        # Check for commune (973XX)
        match_com = re.search(r'(973\d{2})', geo_str)
        if match_com:
            code = int(match_com.group(1))
            if code in COMMUNES_GUYANE:
                result['com'].append({'annee': annee, 'codgeo': code, 'valeur': valeur})
            continue

        # France entière
        if 'france entiere' in geo_lower or 'france entière' in geo_lower or 'france (y compris' in geo_lower:
            result['fra'].append({'annee': annee, 'codgeo': 99, 'valeur': valeur})
            continue

        # France hexagonale
        if 'france metropolitaine' in geo_lower or 'france hexagonale' in geo_lower or 'france métropolitaine' in geo_lower:
            result['fh'].append({'annee': annee, 'codgeo': 0, 'valeur': valeur})
            continue

        # DOM
        if "departements d'outre" in geo_lower or "départements d'outre" in geo_lower:
            result['dom'].append({'annee': annee, 'codgeo': 'DOM', 'valeur': valeur})
            continue

        # Régions
        for reg_name, reg_code in REGION_MAPPING.items():
            if reg_name.lower() in geo_lower:
                result['reg'].append({'annee': annee, 'codgeo': reg_code, 'valeur': valeur})
                break

    dfs = {}
    for k, v in result.items():
        if v:
            dfs[k] = pd.DataFrame(v).drop_duplicates(subset=['annee', 'codgeo'])
        else:
            dfs[k] = pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])

    return dfs


# ============================================================================
# CSV FILE FINDER
# ============================================================================

def find_csv_file(pattern, search_dir=None):
    """Trouve un fichier CSV correspondant au pattern."""
    if search_dir is None:
        search_dir = CSV_SOURCES_DIR
    candidates = list(Path(search_dir).glob(f"*{pattern}*.csv"))
    if not candidates:
        return None
    # Prioriser les fichiers commençant par le pattern (évite les faux positifs)
    for c in candidates:
        if c.name.startswith(pattern):
            return c
    return candidates[0]


# ============================================================================
# YEAR DETECTION FROM CSV FILES
# ============================================================================

def detect_available_years(dataset_id):
    """Détecte les années disponibles dans les fichiers CSV pour un dataset."""
    config = get_dataset_config(dataset_id)
    if not config:
        return []
    
    years = set()
    columns = config.get('columns', [])
    
    for col in columns:
        if col.get('type') != 'variable':
            continue
        
        csv_pattern = col.get('csvPattern')
        if not csv_pattern:
            continue
            
        csv_file = find_csv_file(csv_pattern)
        if not csv_file:
            continue
        
        parser_type = col.get('parser', 'moca')
        try:
            if parser_type == 'long':
                parsed = parse_long_format_csv(csv_file)
            elif parser_type == 'tabular':
                value_col = col.get('column', 2)
                year_col = col.get('yearColumn', 0)
                geo_col = col.get('geoColumn', 1)
                parsed = parse_tabular_csv(csv_file, value_column=value_col, year_column=year_col, geo_column=geo_col)
            elif parser_type == 'moca_filter':
                filter_col = col.get('filterColumn', 4)
                filter_val = col.get('filterValue', '')
                year_col = col.get('yearColumn', 3)
                geo_col = col.get('geoColumn', 5)
                value_col = col.get('valueColumn', 6)
                parsed = parse_moca_filter_csv(csv_file, filter_col, filter_val, year_col, geo_col, value_col)
            elif parser_type == 'external':
                # Skip external sources (DREES, etc.)
                continue
            else:
                parsed = parse_moca_csv(csv_file)

            # Extraire les années depuis tous les niveaux geo
            for level_df in parsed.values():
                if not level_df.empty and 'annee' in level_df.columns:
                    years.update(level_df['annee'].unique())
        except Exception as e:
            print(f"[WARN] Erreur parsing {csv_file}: {e}")
            continue
    
    return sorted(list(years))


# ============================================================================
# GEO ENTITIES
# ============================================================================

GEO_ENTITIES = {
    'com': COMMUNES_GUYANE,
    'reg': REGION_ORDER,
    'dom': ['DOM'],
    'fh': [0],
    'fra': [99]
}


# ============================================================================
# EXCEL SHEET WRITER HELPER
# ============================================================================

def _write_sheet(ws, headers, rows_data, col_keys):
    """Écrit les en-têtes et les données dans une feuille Excel.

    Args:
        ws: Worksheet openpyxl
        headers: Liste des noms de colonnes
        rows_data: Liste de dicts contenant les données
        col_keys: Liste des clés à extraire de chaque row_dict (même ordre que headers)
    """
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = ORANGE_FILL

    for row_idx, row_dict in enumerate(rows_data, 2):
        for col_idx, key in enumerate(col_keys, 1):
            val = row_dict.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if val is not None and col_idx > 1:
                cell.fill = ORANGE_FILL

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10


# ============================================================================
# FILL VARIABLE DATA FROM CSV
# ============================================================================

def _fill_variable_data(data, var_id, parsed, year, time_col_id):
    """Remplit les données d'une variable dans les structures de données.

    Args:
        data: Dict {geo_key: [row_dicts]}
        var_id: ID de la variable
        parsed: Dict {geo_key: DataFrame} du CSV parsé
        year: Année ou période à filtrer
        time_col_id: Nom de la colonne temps (annee ou periode)
    """
    for geo_key in ['com', 'reg', 'dom', 'fh', 'fra']:
        df = parsed.get(geo_key, pd.DataFrame())
        if df.empty:
            continue

        # Filtrer par année/période
        if 'annee' in df.columns:
            df_filtered = df[df['annee'] == year]
        else:
            continue

        id_col = GEO_ID_COLS[geo_key]

        for row in data[geo_key]:
            match = df_filtered[df_filtered['codgeo'] == row[id_col]]
            if not match.empty:
                row[var_id] = round(match.iloc[0]['valeur'], 2)


# ============================================================================
# EXCEL GENERATOR - VERSION CONFIG-DRIVEN (SIMPLE + MULTI-DIMENSION)
# ============================================================================

def generate_prisme_excel(dataset_id, year):
    """Génère une archive ZIP contenant les fichiers Excel PRISME.

    Supporte 3 types de datasets :
    - Simple : [geo, annee, var1, var2, ...]
    - Multi-dimension : [geo, annee, dim, var1, var2, ...]
      (une ligne par combinaison geo × dimension_value)
    - Période : [geo, periode, var1, ...] (au lieu d'annee)

    Args:
        dataset_id: Identifiant du dataset (ex: 'educ', 'pers_sup65ans_seules')
        year: Année ou période à générer (ex: 2021 ou '2015-2020')

    Returns:
        Path du fichier ZIP généré ou None en cas d'erreur
    """

    # Récupérer la config du dataset
    config = get_dataset_config(dataset_id)
    if not config:
        print(f"[ERROR] Dataset inconnu: {dataset_id}")
        print(f"[INFO] Datasets disponibles: {get_available_datasets()}")
        return None

    print(f"[ENGINE] Generation {dataset_id} ({config.get('name', dataset_id)}) pour {year}...")

    # ---- Analyser les colonnes depuis la config ----
    columns = config.get('columns', [])

    # Colonnes par type
    geo_col = next((c for c in columns if c['type'] == 'geo_id'), None)
    time_col = next((c for c in columns if c['type'] in ('year', 'period')), None)
    dim_cols = [c for c in columns if c['type'] == 'dimension']
    var_cols = [c for c in columns if c['type'] == 'variable']

    time_col_id = time_col['id'] if time_col else 'annee'
    variable_ids = [c['id'] for c in var_cols]

    # Multi-dimension ?
    multi_row_dim = config.get('multiRowDimension')
    dim_values = []
    if multi_row_dim:
        dim_config = next((c for c in dim_cols if c['id'] == multi_row_dim), None)
        if dim_config:
            dim_values = dim_config.get('values', [])
            if dim_values:
                print(f"  [DIM] {multi_row_dim} : {len(dim_values)} valeurs ({', '.join(str(v) for v in dim_values[:5])}{'...' if len(dim_values) > 5 else ''})")
            else:
                print(f"  [WARN] Dimension {multi_row_dim} sans valeurs definies")

    # ---- Charger les données CSV pour chaque variable ----
    empty_dfs = lambda: {k: pd.DataFrame(columns=['annee', 'codgeo', 'valeur'])
                         for k in ['com', 'reg', 'dom', 'fh', 'fra']}
    csv_data = {}

    for col in var_cols:
        var_id = col['id']
        csv_pattern = col.get('csvPattern')
        parser_type = col.get('parser', 'moca')

        if not csv_pattern:
            print(f"  [WARN] {var_id} -> Pas de pattern CSV defini")
            csv_data[var_id] = empty_dfs()
            continue

        # Skip external data sources (not CSV-based)
        if parser_type == 'external':
            print(f"  [INFO] {var_id} -> Source externe ({col.get('source', '?')}), pas de CSV")
            csv_data[var_id] = empty_dfs()
            continue

        csv_file = find_csv_file(csv_pattern, CSV_SOURCES_DIR)
        if csv_file:
            if parser_type == 'long':
                csv_data[var_id] = parse_long_format_csv(csv_file)
            elif parser_type == 'tabular':
                value_col = col.get('column', 2)
                year_col = col.get('yearColumn', 0)
                geo_col = col.get('geoColumn', 1)
                csv_data[var_id] = parse_tabular_csv(csv_file, value_column=value_col, year_column=year_col, geo_column=geo_col)
            elif parser_type == 'moca_filter':
                filter_col = col.get('filterColumn', 4)
                filter_val = col.get('filterValue', '')
                year_col = col.get('yearColumn', 3)
                geo_col = col.get('geoColumn', 5)
                value_col = col.get('valueColumn', 6)
                csv_data[var_id] = parse_moca_filter_csv(csv_file, filter_col, filter_val, year_col, geo_col, value_col)
            else:
                csv_data[var_id] = parse_moca_csv(csv_file)
            print(f"  [OK] {var_id} -> {csv_file.name} (parser: {parser_type})")
        else:
            print(f"  [WARN] {var_id} -> Fichier non trouve (pattern: {csv_pattern})")
            csv_data[var_id] = empty_dfs()

    # ---- Construire les structures de données par niveau géo ----
    data = {}
    for geo_key, entities in GEO_ENTITIES.items():
        rows = []
        for entity in entities:
            if dim_values:
                # Multi-dimension : une ligne par entité × valeur de dimension
                for dv in dim_values:
                    row = {GEO_ID_COLS[geo_key]: entity, time_col_id: year, multi_row_dim: dv}
                    rows.append(row)
            else:
                row = {GEO_ID_COLS[geo_key]: entity, time_col_id: year}
                # Ajouter les colonnes dimension non-multi-row avec valeur par défaut
                for dc in dim_cols:
                    if dc.get('values'):
                        row[dc['id']] = dc['values'][0]
                rows.append(row)
        data[geo_key] = rows

    # ---- Remplir les données pour chaque variable ----
    for var_id in variable_ids:
        parsed = csv_data.get(var_id, {})
        _fill_variable_data(data, var_id, parsed, year, time_col_id)

    # ---- Construire l'ordre des colonnes (headers) ----
    # Suit l'ordre exact du config : geo, time, dimensions, variables
    col_keys_template = []  # clés pour extraire les données (sans geo qui varie)
    header_suffixes = []     # noms de colonnes (sans geo qui varie)

    for c in columns:
        if c['type'] == 'geo_id':
            continue  # géré séparément (varie par niveau)
        elif c['type'] in ('year', 'period'):
            col_keys_template.append(c['id'])
            header_suffixes.append(c['id'])
        elif c['type'] == 'dimension':
            col_keys_template.append(c['id'])
            header_suffixes.append(c['id'])
        elif c['type'] == 'variable':
            col_keys_template.append(c['id'])
            header_suffixes.append(c['id'])

    # ---- Noms fichier/dossier ----
    file_name = config.get('fileName', dataset_id)
    folder_path = config.get('folderPath', dataset_id.capitalize())
    theme_folder_name = folder_path.split('/')[-1] if '/' in folder_path else folder_path

    # ---- Dossier temporaire ----
    temp_base = Path(tempfile.mkdtemp())
    root_theme_dir = temp_base / theme_folder_name / str(year)
    root_theme_dir.mkdir(parents=True)

    # ---- Générer un fichier Excel par niveau géographique ----
    for geo_key, folder_name in GEO_FOLDER_MAPPING.items():
        sub_dir = root_theme_dir / folder_name
        sub_dir.mkdir(exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        id_col = GEO_ID_COLS[geo_key]
        ws = wb.create_sheet(geo_key)

        headers = [id_col] + header_suffixes
        col_keys = [id_col] + col_keys_template

        _write_sheet(ws, headers, data[geo_key], col_keys)

        wb.save(sub_dir / f"{file_name}.xlsx")

    # ---- Fichier Consolidé ----
    print(f"  [INFO] Generation fichier consolide...")
    wb_cons = Workbook()
    wb_cons.remove(wb_cons.active)

    for geo_key in GEO_FOLDER_MAPPING:
        id_col = GEO_ID_COLS[geo_key]
        ws = wb_cons.create_sheet(geo_key)

        headers = [id_col] + header_suffixes
        col_keys = [id_col] + col_keys_template

        _write_sheet(ws, headers, data[geo_key], col_keys)

    cons_filename = f"{file_name}_consolidated_{year}.xlsx"
    wb_cons.save(root_theme_dir / cons_filename)
    print(f"  [OK] Fichier consolide cree: {cons_filename}")

    # ---- ZIP final ----
    zip_filename = f"{file_name}_{year}.zip"
    zip_path = OUTPUT_DIR / zip_filename

    if zip_path.exists():
        zip_path.unlink()

    shutil.make_archive(
        str(OUTPUT_DIR / f"{file_name}_{year}"),
        'zip',
        str(temp_base),
        theme_folder_name
    )

    shutil.rmtree(temp_base)

    print(f"[OK] Archive generee: {zip_path}")
    return zip_path


# ============================================================================
# API FUNCTIONS FOR SERVER
# ============================================================================

def get_datasets_info():
    """Retourne les infos de tous les datasets pour l'API."""
    datasets = THEMES_CONFIG.get('datasets', {})
    result = {}
    for ds_id, ds_config in datasets.items():
        result[ds_id] = {
            'id': ds_id,
            'name': ds_config.get('name', ds_id),
            'folderPath': ds_config.get('folderPath', ''),
            'fileName': ds_config.get('fileName', ds_id),
            'sheets': ds_config.get('sheets', ['com', 'reg', 'dom', 'fh', 'fra']),
            'variables': [c['id'] for c in ds_config.get('columns', []) if c.get('type') == 'variable']
        }
    return result

def check_dataset_csv_availability(dataset_id):
    """Vérifie quels fichiers CSV sont disponibles pour un dataset."""
    config = get_dataset_config(dataset_id)
    if not config:
        return {'available': False, 'missing': [], 'found': []}
    
    found = []
    missing = []
    
    for col in config.get('columns', []):
        if col.get('type') != 'variable':
            continue
        
        csv_pattern = col.get('csvPattern')
        if not csv_pattern:
            continue
        
        csv_file = find_csv_file(csv_pattern)
        if csv_file:
            found.append({
                'variable': col['id'],
                'pattern': csv_pattern,
                'file': csv_file.name
            })
        else:
            missing.append({
                'variable': col['id'],
                'pattern': csv_pattern
            })
    
    return {
        'available': len(missing) == 0 and len(found) > 0,
        'found': found,
        'missing': missing
    }


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("PRISME Engine v4.0 - Config-Driven Architecture")
    print("=" * 60)
    
    # Afficher les datasets disponibles
    print(f"\n[INFO] Datasets disponibles: {get_available_datasets()}")
    
    # Test 1: educ (parser moca)
    print("\n[TEST 1] Génération educ_2022.zip")
    result = generate_prisme_excel('educ', 2022)
    if result:
        print(f"  -> Fichier généré: {result}")
    else:
        print("  -> Erreur de génération")
    
    # Test 2: pers_sup65ans_seules (parser long)
    print("\n[TEST 2] Génération pers_sup65ans_seules_2021.zip")
    result = generate_prisme_excel('pers_sup65ans_seules', 2021)
    if result:
        print(f"  -> Fichier généré: {result}")
    else:
        print("  -> Erreur de génération")
    
    # Afficher les années disponibles pour educ
    print("\n[INFO] Années disponibles pour 'educ':", detect_available_years('educ'))
    print("[INFO] Années disponibles pour 'pers_sup65ans_seules':", detect_available_years('pers_sup65ans_seules'))
