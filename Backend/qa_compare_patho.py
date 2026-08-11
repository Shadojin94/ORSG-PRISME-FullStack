#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare deux classeurs pathologies (genere vs cible fournie par le client).

Verifie : memes onglets, memes colonnes dans le meme ordre, memes lignes
(cle = 3 premieres colonnes : geo + sexe + annee, dans l'ordre du fichier),
memes valeurs comparees en float avec une tolerance de 1e-6. Les cibles
drom / fh / Re stockent parfois les cles et/ou les valeurs en texte : tout est
converti avant comparaison. Une cellule vide (secret statistique) ne peut
correspondre qu'a une cellule vide.

Usage :
    python qa_compare_patho.py <genere.xlsx> <cible.xlsx> [--tolerance 1e-6]
    python qa_compare_patho.py --sheet com <multi_onglets.xlsx> <cible.xlsx>

Avec --sheet, seul l'onglet nomme du classeur genere est compare (utile pour
verifier un classeur multi-onglets contre les fichiers cibles mono-onglet) ;
cote cible on prend l'onglet de meme nom, ou son unique onglet.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

TOLERANCE = 1e-6
NB_CLES = 3


def _colonnes_utiles(entete):
    """Retire les colonnes fantomes de fin (en-tete vide)."""
    derniere = 0
    for i, nom in enumerate(entete):
        if nom is not None and str(nom).strip() != "":
            derniere = i + 1
    return derniere


def _cle(valeur):
    """Normalise une cellule de cle : 97301 et '97301' sont equivalents, '000' non."""
    if valeur is None:
        return ""
    if isinstance(valeur, float) and valeur.is_integer():
        return str(int(valeur))
    if isinstance(valeur, int):
        return str(valeur)
    return str(valeur).strip()


def _nombre(valeur):
    if valeur is None:
        return None
    if isinstance(valeur, str):
        valeur = valeur.strip()
        if valeur == "":
            return None
        valeur = valeur.replace(",", ".")
    try:
        return float(valeur)
    except (TypeError, ValueError):
        return None


def _lire(chemin: Path):
    classeur = load_workbook(str(chemin), data_only=True)
    contenu = {}
    for feuille in classeur.worksheets:
        lignes = list(feuille.iter_rows(values_only=True))
        if not lignes:
            contenu[feuille.title] = ([], [])
            continue
        nb = _colonnes_utiles(lignes[0])
        entete = [str(c).strip() for c in lignes[0][:nb]]
        donnees = []
        for ligne in lignes[1:]:
            ligne = list(ligne[:nb]) + [None] * max(0, nb - len(ligne))
            if all(c is None or str(c).strip() == "" for c in ligne):
                continue
            donnees.append(ligne)
        contenu[feuille.title] = (entete, donnees)
    return contenu


def comparer(genere: Path, cible: Path, tolerance=TOLERANCE, onglet_cible=None):
    ecarts = []
    a = _lire(genere)
    b = _lire(cible)

    if onglet_cible is not None:
        if onglet_cible not in a:
            return [f"onglet {onglet_cible!r} absent du genere : {list(a)}"]
        if onglet_cible in b:
            reference = b[onglet_cible]
        elif len(b) == 1:
            reference = next(iter(b.values()))
        else:
            return [f"onglet {onglet_cible!r} absent de la cible : {list(b)}"]
        a = {onglet_cible: a[onglet_cible]}
        b = {onglet_cible: reference}

    if list(a) != list(b):
        ecarts.append(f"onglets differents : genere={list(a)} cible={list(b)}")
        return ecarts

    for onglet in b:
        entete_a, lignes_a = a[onglet]
        entete_b, lignes_b = b[onglet]
        if entete_a != entete_b:
            ecarts.append(
                f"[{onglet}] colonnes differentes\n"
                f"    genere : {entete_a}\n"
                f"    cible  : {entete_b}"
            )
            continue

        index_a = {}
        for i, ligne in enumerate(lignes_a):
            index_a[tuple(_cle(c) for c in ligne[:NB_CLES])] = (i, ligne)
        index_b = {}
        for i, ligne in enumerate(lignes_b):
            index_b[tuple(_cle(c) for c in ligne[:NB_CLES])] = (i, ligne)

        manquantes = [k for k in index_b if k not in index_a]
        en_trop = [k for k in index_a if k not in index_b]
        for k in manquantes:
            ecarts.append(f"[{onglet}] ligne absente du genere : {k}")
        for k in en_trop:
            ecarts.append(f"[{onglet}] ligne en trop dans le genere : {k}")

        for cle, (rang_b, ligne_b) in index_b.items():
            if cle not in index_a:
                continue
            rang_a, ligne_a = index_a[cle]
            if rang_a != rang_b:
                ecarts.append(
                    f"[{onglet}] ligne {cle} a l'ordre {rang_a + 2} "
                    f"au lieu de {rang_b + 2}"
                )
            for j in range(NB_CLES, len(entete_b)):
                va, vb = _nombre(ligne_a[j]), _nombre(ligne_b[j])
                if va is None and vb is None:
                    continue
                if va is None or vb is None:
                    ecarts.append(
                        f"[{onglet}] {cle} / {entete_b[j]} : "
                        f"genere={ligne_a[j]!r} cible={ligne_b[j]!r}"
                    )
                elif abs(va - vb) > tolerance:
                    ecarts.append(
                        f"[{onglet}] {cle} / {entete_b[j]} : "
                        f"genere={va} cible={vb} (ecart {abs(va - vb):.3g})"
                    )
    return ecarts


def main(argv=None):
    parseur = argparse.ArgumentParser(
        description="Compare un classeur pathologies genere a la cible client."
    )
    parseur.add_argument("genere", type=Path)
    parseur.add_argument("cible", type=Path)
    parseur.add_argument("--tolerance", type=float, default=TOLERANCE)
    parseur.add_argument("--max", type=int, default=30,
                         help="nombre maximum d'ecarts affiches")
    parseur.add_argument("--sheet", default=None,
                         help="ne comparer que cet onglet du classeur genere "
                              "(ex. com, dom, fra, fh, reg)")
    args = parseur.parse_args(argv)

    for chemin in (args.genere, args.cible):
        if not chemin.is_file():
            raise SystemExit(f"fichier introuvable : {chemin}")

    ecarts = comparer(args.genere, args.cible, args.tolerance, args.sheet)
    suffixe = f" [onglet {args.sheet}]" if args.sheet else ""
    if not ecarts:
        print(f"OK 0 ecart{suffixe}  ({args.genere.name} == {args.cible.name})")
        return 0
    print(f"KO {len(ecarts)} ecart(s){suffixe}  ({args.genere.name} vs {args.cible.name})")
    for message in ecarts[:args.max]:
        print("  -", message)
    if len(ecarts) > args.max:
        print(f"  ... {len(ecarts) - args.max} ecart(s) supplementaire(s)")
    return 1


if __name__ == "__main__":
    sys.exit(main())
