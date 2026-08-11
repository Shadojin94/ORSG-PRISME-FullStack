#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reagencement du volet "pathologie" PRISME.

Lit l'export MOCA-O `mortalite_patho_<debut>_<fin>.xls` (format OLE2, lu par xlrd)
et produit, pour une annee donnee, les 5 classeurs attendus par le client :

    mortalite_com.xlsx   onglet "com"   codes communes 973xx      (22 x 3 sexes)
    mortalite_Re.xlsx    onglet "reg"   codes regions              (17 x 3 sexes)
    mortalite_drom.xlsx  onglet "dom"   code 999  (DROM)           (1 x 3 sexes)
    mortalite_fe.xlsx    onglet "fra"   code 99   (France entiere) (1 x 3 sexes)
    mortalite_fh.xlsx    onglet "fh"    code 000  (France hexag.)  (1 x 3 sexes)

Avec --single-file, les memes 5 feuilles sont ecrites dans UN seul classeur
`mortalite_patho_<annee>.xlsx` (onglets com, dom, fra, fh, reg dans cet ordre),
chaque onglet strictement identique a celui du fichier separe correspondant.

Usage :
    python generate_patho_reorganisation.py <source.xls> <annee> [--outdir DIR]
                                            [--no-fill] [--single-file]

Structure du fichier source (cf. README de synthese) :
  - 237 colonnes, mais SEULES les colonnes A-G portent des donnees. Les 230 autres
    ne contiennent que la legende des modalites en ligne 1 (blocs fantomes MOCA-O).
  - colonnes A-G : annee | lieu | sexe | age | population_reference | cause | valeur
  - le fichier est la concatenation de 4 sections (communes / regions / France+DROM /
    France avec Mayotte), chacune precedee de 2 lignes d'en-tete.
  - sexe : "Sexe" = ensemble, "masculin" = H, "feminin" = F.
  - valeur -999 = secret statistique -> cellule vide dans les sorties.
  - 3 causes ont perdu leur libelle en colonne F (la cellule vaut
    "Causes_de_deces_alphabetique" sans "#libelle") : BPCO, Chap 5 troubles mentaux
    et Insuffisance cardiaque. Elles sont retrouvees par POSITION : chaque triplet
    (lieu, sexe, annee) apparait exactement 17 fois, toujours dans le meme ordre,
    celui des blocs de la ligne 1. Les 14 causes libellees servent de controle.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from decimal import Decimal, ROUND_DOWN
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# --------------------------------------------------------------------------
# Causes : ordre d'apparition dans le fichier source (= ordre des blocs ligne 1)
# --------------------------------------------------------------------------
# (slug, fragment de libelle attendu ; None = cause dont le libelle est absent)
CAUSES_SOURCE_ORDER = [
    ("alzheimer",        "Alzheimer"),
    ("asthme",           "Asthme"),
    ("bpco",             None),                      # Broncho-pneumopathie chronique obstructive
    ("kc_prostate",      "Cancer de la prostate"),
    ("kc_poumon",        "Cancer de la trach"),      # trachee, bronches et poumon
    ("kc_colon",         "Cancer du c"),             # colon-rectum
    ("kc_sein",          "Cancer du sein"),
    ("cardiopath_isch",  "Cardiopathies isch"),
    ("trouble_ment",     None),                      # Chap 5 : Troubles mentaux
    ("maladies_respi",   "Chap 10"),                 # Maladies de l'appareil respiratoire
    ("diabete",          "Diab"),
    ("insuff_cardiaque", None),                      # Insuffisance cardiaque
    ("foie",             "Maladie chronique du foie"),
    ("parkinson",        "Maladie de Parkinson"),
    ("hypertensive",     "Maladie hypertensive"),
    ("sida",             "Sida"),
    ("avc",              "Maladies vasculaires"),
]

# Ordre des colonnes indicateurs dans les fichiers cibles (avc AVANT sida)
INDICATEURS = [
    "alzheimer", "asthme", "bpco", "kc_prostate", "kc_poumon", "kc_colon",
    "kc_sein", "cardiopath_isch", "trouble_ment", "maladies_respi", "diabete",
    "insuff_cardiaque", "foie", "parkinson", "hypertensive", "avc", "sida",
]
COLONNES_INDICATEURS = ["m_" + s for s in INDICATEURS]

SECRET_STAT = -999.0

# sexe cible -> sexe source
SEXES = {"Ens": "Sexe", "H": "masculin", "F": "feminin"}

# code region -> nom normalise (ASCII) tel que present dans le source
REGIONS = [
    ("11", "le de france"),                 # Ile-de-France
    ("24", "centre val de loire"),
    ("27", "bourgogne franche comt"),
    ("28", "normandie"),
    ("32", "hauts de france"),
    ("44", "grand est"),
    ("52", "pays de la loire"),
    ("53", "bretagne"),
    ("75", "nouvelle aquitaine"),
    ("76", "occitanie"),
    ("84", "auvergne rhne alpes"),
    ("93", "provence alpes cte d azur"),
    ("94", "corse"),
    ("01", "guadeloupe"),
    ("02", "martinique"),
    ("03", "guyane"),
    ("04", "runion"),
]
GEO_DROM = "departements d outre mer"
GEO_FRANCE_HEXA = "france mtropolitaine"
GEO_FRANCE_ENTIERE = "france (y compris mayotte)"

# Mise en forme reprise des fichiers exemples du client
HEADER_FILL_RGB = "FFABBBDB"
FREEZE_PANES = "D2"


def tronquer15(valeur: float) -> float:
    """Tronque a 15 chiffres significatifs, comme le stockage Excel des cibles."""
    if valeur == 0:
        return 0.0
    nombre = Decimal(repr(valeur))
    exposant = Decimal(1).scaleb(nombre.adjusted() - 14)
    return float(nombre.quantize(exposant, rounding=ROUND_DOWN))


def normalise(libelle: str) -> str:
    """Squelette ASCII minuscule d'un libelle geographique.

    Les libelles du source ont un encodage casse (accents remplaces par U+FFFD) :
    on retire purement et simplement tout caractere non-ASCII.
    """
    ascii_only = "".join(ch for ch in libelle if ord(ch) < 128)
    return " ".join(ascii_only.lower().replace("-", " ").replace("'", " ").split())


# --------------------------------------------------------------------------
# Lecture du source
# --------------------------------------------------------------------------
def lire_source(chemin: Path):
    """Retourne (valeurs, communes) ou

    valeurs  : dict[(geo_normalise, sexe_cible, annee, slug_cause)] -> float | None
               (None = secret statistique -999)
    communes : liste ordonnee des (code_insee, geo_normalise)
    """
    classeur = xlrd.open_workbook(str(chemin))
    feuille = classeur.sheet_by_index(0)

    sexe_inverse = {v: k for k, v in SEXES.items()}
    compteur = Counter()
    valeurs = {}
    communes = []
    communes_vues = set()
    incoherences = []

    for idx in range(feuille.nrows):
        ligne = feuille.row_values(idx)
        annee_brute = ligne[0]
        if not isinstance(annee_brute, float):
            continue  # lignes d'en-tete des 4 sections
        sexe_source = str(ligne[2]).strip()
        if sexe_source not in sexe_inverse:
            continue

        annee = int(annee_brute)
        geo_brut = str(ligne[1])
        geo = normalise(geo_brut)
        sexe = sexe_inverse[sexe_source]

        cle_triplet = (geo, sexe, annee)
        rang = compteur[cle_triplet]
        compteur[cle_triplet] += 1
        if rang >= len(CAUSES_SOURCE_ORDER):
            raise ValueError(
                f"ligne {idx + 1} : plus de {len(CAUSES_SOURCE_ORDER)} causes pour "
                f"{cle_triplet} - structure du fichier source inattendue"
            )
        slug, fragment = CAUSES_SOURCE_ORDER[rang]

        # controle : quand le libelle est present il doit coller a la position
        libelle = str(ligne[5])
        if "#" in libelle:
            queue = libelle.split("#", 1)[1]
            if fragment is None or not queue.startswith(fragment):
                incoherences.append((idx + 1, rang, libelle))

        try:
            valeur = float(ligne[6])
        except (TypeError, ValueError):
            valeur = None
        if valeur == SECRET_STAT:
            valeur = None

        valeurs[(geo, sexe, annee, slug)] = valeur

        code = geo_brut.split(" - ", 1)[0].strip()
        if code.isdigit() and len(code) == 5 and geo not in communes_vues:
            communes_vues.add(geo)
            communes.append((int(code), geo))

    if incoherences:
        ligne, rang, libelle = incoherences[0]
        raise ValueError(
            f"{len(incoherences)} libelle(s) de cause incoherent(s) avec la position "
            f"(ex. ligne {ligne}, position {rang} : {libelle!r}). "
            "L'ordre des causes du fichier source a change."
        )

    mauvais = [c for c, n in compteur.items() if n != len(CAUSES_SOURCE_ORDER)]
    if mauvais:
        raise ValueError(
            f"{len(mauvais)} triplet(s) (lieu, sexe, annee) n'ont pas "
            f"{len(CAUSES_SOURCE_ORDER)} causes, ex. {mauvais[0]}"
        )

    return valeurs, communes


# --------------------------------------------------------------------------
# Ecriture des classeurs
# --------------------------------------------------------------------------
def ecrire_feuille(feuille, onglet, entetes, lignes, formats, avec_fill=True):
    """Remplit une feuille existante (mise en forme unique pour les deux modes).

    formats : slug_de_rendu -> applique aux cellules de valeur
      "brut"    valeur float tronquee a 15 chiffres significatifs (precision de
                stockage d'Excel, celle des fichiers exemples), affichage 0.00
      "arrondi" valeur float arrondie a 2 decimales, format General
      "texte"   valeur ecrite en texte (str(float)), format General
    """
    feuille.title = onglet

    gras = Font(bold=True)
    fond = PatternFill("solid", fgColor=HEADER_FILL_RGB) if avec_fill else None
    for col, nom in enumerate(entetes, start=1):
        cellule = feuille.cell(row=1, column=col, value=nom)
        cellule.font = gras
        cellule.number_format = "@"
        if fond is not None:
            cellule.fill = fond

    for i, (cles, valeurs) in enumerate(lignes, start=2):
        for col, cle in enumerate(cles, start=1):
            cellule = feuille.cell(row=i, column=col, value=cle)
            if isinstance(cle, str):
                cellule.number_format = "@"
        for j, valeur in enumerate(valeurs):
            col = len(cles) + 1 + j
            if valeur is None:
                cellule = feuille.cell(row=i, column=col)
            elif formats == "arrondi":
                cellule = feuille.cell(row=i, column=col, value=round(valeur, 2))
            elif formats == "texte":
                cellule = feuille.cell(row=i, column=col, value=str(valeur))
            else:  # brut
                cellule = feuille.cell(row=i, column=col, value=tronquer15(valeur))
            if formats == "brut":
                cellule.number_format = "0.00"

    derniere = len(lignes) + 1
    feuille.auto_filter.ref = f"A1:T{derniere}"
    feuille.freeze_panes = FREEZE_PANES
    return derniere - 1


def serie(valeurs, geo, sexe, annee):
    """Les 17 indicateurs, dans l'ordre des colonnes cibles."""
    return [valeurs.get((geo, sexe, annee, slug)) for slug in INDICATEURS]


def construire_feuilles(valeurs, communes, annee):
    """Definition des 5 feuilles, dans l'ordre du classeur unique.

    Retourne une liste de dicts : onglet, fichier (mode 5 fichiers), entetes,
    lignes, formats. C'est la seule source de verite pour les deux modes.
    """
    return [
        {
            "onglet": "com",
            "fichier": "mortalite_com.xlsx",
            "entetes": ["com", "sexe", "annee"] + COLONNES_INDICATEURS,
            "formats": "brut",
            "lignes": [
                ((code, sexe, annee), serie(valeurs, geo, sexe, annee))
                for sexe in ("Ens", "H", "F")
                for code, geo in communes
            ],
        },
        {
            "onglet": "dom",
            "fichier": "mortalite_drom.xlsx",
            "entetes": ["dom", "sexe", "annee"] + COLONNES_INDICATEURS,
            "formats": "texte",
            "lignes": [
                (("999", sexe, str(annee)), serie(valeurs, GEO_DROM, sexe, annee))
                for sexe in ("Ens", "H", "F")
            ],
        },
        {
            # France entiere : colonnes fra / annee / sexe, ordre Ens, F, H
            "onglet": "fra",
            "fichier": "mortalite_fe.xlsx",
            "entetes": ["fra", "annee", "sexe"] + COLONNES_INDICATEURS,
            "formats": "arrondi",
            "lignes": [
                (("99", str(annee), sexe), serie(valeurs, GEO_FRANCE_ENTIERE, sexe, annee))
                for sexe in ("Ens", "F", "H")
            ],
        },
        {
            "onglet": "fh",
            "fichier": "mortalite_fh.xlsx",
            "entetes": ["fh", "sexe", "annee"] + COLONNES_INDICATEURS,
            "formats": "arrondi",
            "lignes": [
                (("000", sexe, str(annee)), serie(valeurs, GEO_FRANCE_HEXA, sexe, annee))
                for sexe in ("Ens", "H", "F")
            ],
        },
        {
            "onglet": "reg",
            "fichier": "mortalite_Re.xlsx",
            "entetes": ["reg", "sexe", "annee"] + COLONNES_INDICATEURS,
            "formats": "arrondi",
            "lignes": [
                ((code, sexe, str(annee)), serie(valeurs, geo, sexe, annee))
                for sexe in ("Ens", "H", "F")
                for code, geo in REGIONS
            ],
        },
    ]


def generer(source: Path, annee: int, outdir: Path, avec_fill=True, single_file=False):
    valeurs, communes = lire_source(source)

    annees = sorted({cle[2] for cle in valeurs})
    if annee not in annees:
        raise SystemExit(
            f"annee {annee} absente du fichier source (disponibles : "
            f"{', '.join(str(a) for a in annees)})"
        )

    geos = {cle[0] for cle in valeurs}
    for attendu in (GEO_DROM, GEO_FRANCE_HEXA, GEO_FRANCE_ENTIERE):
        if attendu not in geos:
            raise SystemExit(f"niveau geographique absent du source : {attendu!r}")
    for code, nom in REGIONS:
        if nom not in geos:
            raise SystemExit(f"region absente du source : {code} / {nom!r}")

    outdir.mkdir(parents=True, exist_ok=True)
    feuilles = construire_feuilles(valeurs, communes, annee)

    if single_file:
        classeur = Workbook()
        total = 0
        for i, spec in enumerate(feuilles):
            feuille = classeur.active if i == 0 else classeur.create_sheet()
            total += ecrire_feuille(feuille, spec["onglet"], spec["entetes"],
                                    spec["lignes"], spec["formats"], avec_fill)
        chemin = outdir / f"mortalite_patho_{annee}.xlsx"
        classeur.save(str(chemin))
        return [(chemin, total)]

    produits = []
    for spec in feuilles:
        classeur = Workbook()
        n = ecrire_feuille(classeur.active, spec["onglet"], spec["entetes"],
                           spec["lignes"], spec["formats"], avec_fill)
        chemin = outdir / spec["fichier"]
        classeur.save(str(chemin))
        produits.append((chemin, n))
    return produits


def main(argv=None):
    parseur = argparse.ArgumentParser(
        description="Genere les 5 classeurs pathologies PRISME a partir de l'export MOCA-O."
    )
    parseur.add_argument("source", type=Path, help="fichier .xls MOCA-O")
    parseur.add_argument("annee", type=int, help="annee a extraire (ex. 2018)")
    parseur.add_argument("--outdir", type=Path, default=Path("."),
                         help="repertoire de sortie (defaut : repertoire courant)")
    parseur.add_argument("--no-fill", action="store_true",
                         help="ne pas colorer la ligne d'en-tete (les fichiers "
                              "exemples du client sont en ABBBDB)")
    parseur.add_argument("--single-file", action="store_true",
                         help="produire un unique classeur mortalite_patho_<annee>.xlsx "
                              "a 5 onglets (com, dom, fra, fh, reg) au lieu des "
                              "5 fichiers separes")
    args = parseur.parse_args(argv)

    if not args.source.is_file():
        raise SystemExit(f"fichier source introuvable : {args.source}")

    produits = generer(args.source, args.annee, args.outdir,
                       avec_fill=not args.no_fill, single_file=args.single_file)
    for chemin, nb in produits:
        print(f"OK  {chemin}  ({nb} lignes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
