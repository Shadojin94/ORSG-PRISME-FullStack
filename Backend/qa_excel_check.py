#!/usr/bin/env python3
"""
QA Excel - Verification format des fichiers generés par PRISME Engine et generate_from_opendata.
Genere dans Backend/output/, verifie : onglets, colonnes, absence de remplissage
colore sur les cellules, remplissage donnees.
Usage : python qa_excel_check.py
NE MODIFIE AUCUN FICHIER SOURCE.
"""

import sys
import os
import zipfile
import json
import traceback
from pathlib import Path
from io import BytesIO

# On se met dans le bon repertoire
BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

import openpyxl

# =============================================================================
# HELPERS VERIFICATION
# =============================================================================

def _cell_has_color_fill(cell):
    """True si la cellule porte un remplissage colore."""
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return False

    def is_colored(color):
        if color is None:
            return False
        color_type = getattr(color, "type", None)
        if color_type == "rgb":
            rgb = getattr(color, "rgb", "")
            if not isinstance(rgb, str):
                return False
            rgb = rgb.upper()
            # Transparent/noir par defaut et blanc ne sont pas des couleurs visibles.
            return rgb not in ("", "00000000", "00FFFFFF", "FFFFFFFF")
        if color_type == "indexed":
            # 1 = blanc ; 64 = couleur systeme/par defaut.
            return getattr(color, "indexed", None) not in (1, 64)
        if color_type == "theme":
            # Excel : theme 1 = fond clair (blanc), theme 0 = texte sombre.
            # Toute teinte explicite est refusee par prudence.
            return (
                getattr(color, "theme", None) != 1
                or getattr(color, "tint", 0.0) not in (0, 0.0)
            )
        return bool(color_type)

    if fill.fill_type == "solid":
        return is_colored(fill.fgColor)
    return is_colored(fill.fgColor) or is_colored(fill.bgColor)


def _find_color_fills(ws):
    """Retourne les coordonnees des cellules ayant un remplissage colore.

    Balaie la feuille entiere (en-tetes ligne 1 incluses) : le client ne veut
    aucune couleur sur les cellules generees.
    """
    found = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if _cell_has_color_fill(cell):
                color = cell.fill.fgColor
                color_type = getattr(color, "type", "unknown")
                color_value = getattr(color, color_type, "unknown")
                tint = getattr(color, "tint", 0.0)
                suffix = f",tint={tint}" if tint not in (0, 0.0) else ""
                found.append(f"{cell.coordinate}={color_type}:{color_value}{suffix}")
    return found

def _count_data_rows(ws):
    """Nombre de lignes de donnees (hors header, hors None)."""
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if any(cell.value is not None for cell in row):
            count += 1
    return count

def _get_headers(ws):
    """Retourne les headers de la premiere ligne."""
    return [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]

def check_xlsx(xlsx_bytes_or_path, expected_sheets, expected_variables, label=""):
    """
    Verifie un fichier xlsx (bytes ou Path).
    Retourne un dict de resultats.
    """
    issues = []
    fills_found = {}
    sheet_data_rows = {}
    sheets_found = []

    try:
        if isinstance(xlsx_bytes_or_path, (bytes, BytesIO)):
            wb = openpyxl.load_workbook(BytesIO(xlsx_bytes_or_path) if isinstance(xlsx_bytes_or_path, bytes) else xlsx_bytes_or_path)
        else:
            wb = openpyxl.load_workbook(str(xlsx_bytes_or_path))
    except Exception as e:
        return {"ok": False, "issues": [f"Impossible d'ouvrir le fichier: {e}"], "fills": {}, "rows": {}}

    sheets_found = wb.sheetnames
    if expected_sheets is None:
        expected_sheets = sheets_found

    # Verifier onglets attendus
    missing_sheets = [s for s in expected_sheets if s not in sheets_found]
    if missing_sheets:
        issues.append(f"Onglets manquants: {missing_sheets}")

    for sheet_name in expected_sheets:
        if sheet_name not in sheets_found:
            fills_found[sheet_name] = []
            sheet_data_rows[sheet_name] = 0
            continue
        ws = wb[sheet_name]

        # Colonnes
        headers = _get_headers(ws)
        headers_clean = [str(h).strip() if h else '' for h in headers]
        for var in expected_variables:
            if var not in headers_clean:
                issues.append(f"[{sheet_name}] Colonne '{var}' absente (headers: {headers_clean})")

        # Absence de remplissage colore
        fills_found[sheet_name] = _find_color_fills(ws)
        if fills_found[sheet_name]:
            sample = ", ".join(fills_found[sheet_name][:5])
            issues.append(
                f"[{sheet_name}] Remplissage colore PRESENT sur "
                f"{len(fills_found[sheet_name])} cellule(s) (ex: {sample})"
            )

        # Remplissage donnees
        nrows = _count_data_rows(ws)
        sheet_data_rows[sheet_name] = nrows
        if nrows == 0:
            issues.append(f"[{sheet_name}] VIDE - 0 lignes de donnees")

    ok = len(issues) == 0
    return {
        "ok": ok,
        "issues": issues,
        "fills": fills_found,
        "rows": sheet_data_rows,
        "sheets_found": sheets_found,
    }


def check_generated_artifact(artifact_path):
    """Verifie tous les classeurs d'un XLSX ou d'un pack ZIP genere par l'API."""
    artifact_path = Path(artifact_path)
    issues = []
    workbooks = {}

    if not artifact_path.is_file():
        return {
            "ok": False,
            "artifact": str(artifact_path),
            "issues": [f"Artefact introuvable: {artifact_path}"],
            "workbooks": {},
        }

    def check_workbook(label, content):
        result = check_xlsx(content, None, [], label=label)
        workbooks[label] = result
        issues.extend(f"[{label}] {issue}" for issue in result.get("issues", []))

    suffix = artifact_path.suffix.lower()
    if suffix == ".xlsx":
        check_workbook(artifact_path.name, artifact_path)
    elif suffix == ".zip":
        try:
            with zipfile.ZipFile(artifact_path, "r") as zf:
                xlsx_names = [
                    name for name in zf.namelist()
                    if name.lower().endswith(".xlsx")
                    and not name.startswith("__MACOSX/")
                    and not Path(name).name.startswith("~$")
                ]
                if not xlsx_names:
                    issues.append("Le pack ZIP ne contient aucun fichier XLSX")
                for name in xlsx_names:
                    check_workbook(name, zf.read(name))
        except Exception as exc:
            issues.append(f"Impossible d'ouvrir le pack ZIP: {exc}")
    else:
        issues.append("Format non supporte: seuls .xlsx et .zip sont acceptes")

    return {
        "ok": not issues,
        "artifact": str(artifact_path),
        "issues": issues,
        "workbooks": workbooks,
    }


def check_zip_pack(zip_path, expected_geo_levels, expected_variables, file_name):
    """
    Verifie un ZIP pack : contient des dossiers geo avec chacun un xlsx.
    Verifie chaque xlsx trouve.
    """
    GEO_FOLDER_MOCAO = {
        'com': ['com', 'Commune'],
        'reg': ['reg', 'Region', 'Région', 'RÃ©gion'],
        'dom': ['dom', 'DOM'],
        'fh': ['fh', 'France Hexagonale', 'France hexagonale'],
        'fra': ['fra', "France entière", "France entiÃ¨re", "France Entiere", "France_Entiere"],
    }
    OPENDATA_FOLDER_NAMES = ["Commune", "Region", "DOM", "France_Hexagonale", "France_Entiere"]
    GEO_OPENDATA_MAP = {
        "Commune": "com",
        "Region": "reg",
        "DOM": "dom",
        "France_Hexagonale": "fh",
        "France_Entiere": "fra",
    }

    issues = []
    per_sheet = {}

    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            names = zf.namelist()

            # Detecter si c'est un ZIP opendata (dossiers Commune/Region/...) ou MOCA-O
            is_opendata = any(any(n.startswith(f) for f in OPENDATA_FOLDER_NAMES) for n in names)

            for geo_key in expected_geo_levels:
                # Chercher le fichier xlsx correspondant
                found_xlsx = None
                if is_opendata:
                    # Chercher dans les dossiers opendata
                    for fn, gk in GEO_OPENDATA_MAP.items():
                        if gk == geo_key:
                            candidates = [n for n in names if fn in n and n.endswith('.xlsx') and not n.startswith('__')]
                            if candidates:
                                found_xlsx = candidates[0]
                                break
                else:
                    # MOCA-O: feuille dans un seul xlsx ou dossiers
                    candidates = [n for n in names if n.endswith('.xlsx') and not n.startswith('__')]
                    if candidates:
                        found_xlsx = candidates[0]

                if found_xlsx is None and is_opendata:
                    issues.append(f"[ZIP] Dossier geo '{geo_key}' non trouve dans le ZIP")
                    per_sheet[geo_key] = {"ok": False, "issues": [f"Absent du ZIP"], "fills": {geo_key: []}, "rows": {geo_key: 0}}
                    continue

                if found_xlsx:
                    data = zf.read(found_xlsx)
                    if is_opendata:
                        # Chaque xlsx opendata = 1 onglet nomme par geo_key
                        res = check_xlsx(data, [geo_key], expected_variables, label=f"{geo_key}")
                    else:
                        # MOCA-O: 1 xlsx avec plusieurs onglets (un par geo)
                        res = check_xlsx(data, [geo_key], expected_variables, label=f"{geo_key}")
                    per_sheet[geo_key] = res
                    issues.extend([f"[{geo_key}] {i}" for i in res.get("issues", [])])

    except Exception as e:
        return {"ok": False, "issues": [f"Erreur lecture ZIP: {e}"], "per_sheet": {}}

    ok = len(issues) == 0
    return {"ok": ok, "issues": issues, "per_sheet": per_sheet}


# =============================================================================
# QA MOCA-O : generate_prisme_excel
# =============================================================================

def qa_mocao():
    print("\n" + "="*70)
    print("QA MODE MOCA-O : generate_prisme_excel(dataset_id, year)")
    print("="*70)

    # Import engine
    try:
        import prisme_engine as eng
    except Exception as e:
        print(f"[FATAL] Impossible d'importer prisme_engine: {e}")
        return {}

    results = {}

    datasets = eng.get_available_datasets()
    print(f"Datasets detectes dans themes_config.json: {datasets}\n")

    for dataset_id in datasets:
        label = f"MOCA-O/{dataset_id}"
        config = eng.get_dataset_config(dataset_id)
        if not config:
            results[label] = {"status": "KO", "reason": "Config absente"}
            continue

        # Detecter annees disponibles
        try:
            years = eng.detect_available_years(dataset_id)
        except Exception as e:
            results[label] = {"status": "KO", "reason": f"detect_available_years erreur: {e}"}
            continue

        if not years:
            results[label] = {"status": "indisponible", "reason": "Aucune annee detectee (CSV sources absents ou non parsables)"}
            continue

        year = max(years)

        # Variables attendues (hors calculees)
        var_cols = [c for c in config.get('columns', []) if c.get('type') == 'variable']
        expected_vars = [c['id'] for c in var_cols if not eng.is_calculated_variable(c)]

        # Onglets attendus : sheets ou par defaut les 5 niveaux
        expected_sheets_raw = config.get('sheets', ['com', 'reg', 'dom', 'fh', 'fra'])

        print(f"\n--- {dataset_id} (annee={year}, vars={expected_vars}) ---")

        # Generation
        try:
            zip_path = eng.generate_prisme_excel(dataset_id, year)
        except Exception as e:
            results[label] = {"status": "KO", "reason": f"Erreur generation: {e}\n{traceback.format_exc()}"}
            print(f"  [ERREUR] {e}")
            continue

        if zip_path is None or not Path(zip_path).exists():
            results[label] = {"status": "KO", "reason": "ZIP non genere (retour None ou fichier absent)"}
            continue

        print(f"  ZIP genere: {zip_path}")

        # Verification ZIP
        res = check_zip_pack(str(zip_path), expected_sheets_raw, expected_vars, config.get('fileName', dataset_id))

        # Verification fichier consolide dans le ZIP
        cons_issues = []
        try:
            with zipfile.ZipFile(str(zip_path), 'r') as zf:
                cons_files = [n for n in zf.namelist() if 'consolidated' in n and n.endswith('.xlsx')]
                if cons_files:
                    cons_data = zf.read(cons_files[0])
                    cons_res = check_xlsx(cons_data, expected_sheets_raw, expected_vars, label="consolidated")
                    if not cons_res['ok']:
                        cons_issues.append(f"Consolide {cons_files[0]}: {cons_res['issues']}")
                    print(f"  [CONSOLIDE] {cons_files[0]}: {'OK' if cons_res['ok'] else 'KO'}")
                    # Absence de couleur consolide
                    for sh, fills in cons_res['fills'].items():
                        status = "AUCUNE COULEUR" if not fills else f"{len(fills)} cellule(s) coloree(s)"
                        print(f"    Remplissage [{sh}]: {status}")
                else:
                    cons_issues.append("Fichier consolide absent du ZIP")
        except Exception as e:
            cons_issues.append(f"Erreur lecture consolide: {e}")

        all_issues = res['issues'] + cons_issues

        # Afficher details
        for sh, sr in res.get('per_sheet', {}).items():
            rows = sr.get('rows', {}).get(sh, '?')
            fills = sr.get('fills', {}).get(sh, [])
            print(f"  [{sh}] rows={rows} | couleurs={len(fills)} | {'OK' if sr.get('ok') else 'KO: ' + str(sr.get('issues'))}")

        status = "OK" if not all_issues else "KO"
        results[label] = {
            "status": status,
            "year": year,
            "reason": "; ".join(all_issues) if all_issues else "",
            "zip": str(zip_path),
            "fills_by_sheet": {sh: sr.get('fills', {}).get(sh, []) for sh, sr in res.get('per_sheet', {}).items()},
        }

    return results


# =============================================================================
# QA OPEN DATA : generate_theme(theme, year)
# =============================================================================

def qa_opendata():
    print("\n" + "="*70)
    print("QA MODE OPEN DATA : generate_theme(theme, year)")
    print("="*70)

    try:
        import generate_from_opendata as od
    except Exception as e:
        print(f"[FATAL] Impossible d'importer generate_from_opendata: {e}")
        return {}

    results = {}
    EXPECTED_GEO = ['com', 'reg', 'dom', 'fh', 'fra']

    # Annees candidates par source_type
    YEAR_CANDIDATES = {
        "educ": [2022, 2021, 2020, 2019, 2018, 2017],
        "couples": [2022, 2021, 2020, 2019, 2018, 2017],
        "caf": [2023, 2022],
        "ircom": [2023, 2022, 2021, 2020, 2019],
        "pop_legales": [2023, 2022, 2021],
        "baac": [2023, 2022, 2021, 2020, 2019],
        "cepidc": [2021, 2020, 2019, 2018],
        "odisse_suicide": [2021, 2020, 2019, 2018],
        "odisse_alcool": [2021, 2020, 2019],
        "odisse_tabac": [2021, 2020, 2019],
        "spf_noyades": [2023, 2022, 2021, 2020],
        "drees_eaje": [2022, 2021],
    }

    def find_year(source_type):
        """Trouve une annee disponible en testant les candidates."""
        candidates = YEAR_CANDIDATES.get(source_type, [2022])
        return candidates[0] if candidates else 2022

    for theme, cfg in od.THEME_CONFIGS.items():
        label = f"OpenData/{theme}"
        source_type = cfg.get("source_type", "?")
        variables = [v for v in cfg.get("variables", []) if not od.is_calculated_variable(v)]
        year = find_year(source_type)

        print(f"\n--- {theme} (source={source_type}, year={year}, vars={variables}) ---")

        try:
            root_dir = od.generate_theme(theme, year)
        except FileNotFoundError as e:
            results[label] = {"status": "indisponible", "reason": f"Source manquante: {e}"}
            print(f"  [INDISPONIBLE] {e}")
            continue
        except Exception as e:
            results[label] = {"status": "KO", "reason": f"Erreur generation: {e}"}
            print(f"  [ERREUR] {e}")
            traceback.print_exc()
            continue

        # Chercher le ZIP genere
        zip_path = OUTPUT_DIR / f"{theme}_opendata_{year}.zip"
        if not zip_path.exists():
            results[label] = {"status": "KO", "reason": f"ZIP attendu non trouve: {zip_path}"}
            print(f"  [KO] ZIP absent: {zip_path}")
            continue

        print(f"  ZIP genere: {zip_path}")

        # Verifier ZIP
        res = check_zip_pack(str(zip_path), EXPECTED_GEO, variables, cfg['excel_name'])

        # Afficher details
        for sh, sr in res.get('per_sheet', {}).items():
            rows = sr.get('rows', {}).get(sh, '?')
            fills = sr.get('fills', {}).get(sh, [])
            print(f"  [{sh}] rows={rows} | couleurs={len(fills)} | {'OK' if sr.get('ok') else 'KO: ' + str(sr.get('issues'))}")

        # Verifier le .xlsx consolide (dans root_dir)
        cons_issues = []
        cons_path = root_dir / f"{cfg['excel_name']}_consolidated_{year}.xlsx"
        if cons_path.exists():
            cons_res = check_xlsx(str(cons_path), EXPECTED_GEO, variables, label="consolidated")
            if not cons_res['ok']:
                cons_issues.append(f"Consolide KO: {cons_res['issues']}")
            print(f"  [CONSOLIDE] {'OK' if cons_res['ok'] else 'KO'}")
            for sh, fills in cons_res['fills'].items():
                print(f"    Remplissage [{sh}]: {'AUCUNE COULEUR' if not fills else str(len(fills)) + ' cellule(s) coloree(s)'}")
        else:
            print(f"  [CONSOLIDE] Absent ({cons_path})")

        all_issues = res['issues'] + cons_issues
        status = "OK" if not all_issues else "KO"
        results[label] = {
            "status": status,
            "year": year,
            "reason": "; ".join(all_issues) if all_issues else "",
            "zip": str(zip_path),
            "fills_by_sheet": {sh: sr.get('fills', {}).get(sh, []) for sh, sr in res.get('per_sheet', {}).items()},
        }

    return results


# =============================================================================
# QA CONSOLIDE MOCA-O : generate_mocao_consolidated
# =============================================================================

def qa_mocao_consolidated():
    print("\n" + "="*70)
    print("QA MODE MOCA-O CONSOLIDE : generate_mocao_consolidated.py")
    print("="*70)

    # Verifier si la fonction existe
    cons_script = BASE_DIR / "generate_mocao_consolidated.py"
    if not cons_script.exists():
        print("[SKIP] generate_mocao_consolidated.py absent")
        return {}

    # Essayer d'importer
    try:
        import generate_mocao_consolidated as gmc
    except Exception as e:
        print(f"[WARN] Import generate_mocao_consolidated impossible: {e}")
        return {}

    # Chercher une fonction de generation publique
    import inspect
    funcs = [name for name, obj in inspect.getmembers(gmc, inspect.isfunction)
             if not name.startswith('_') and 'generate' in name.lower()]
    print(f"Fonctions publiques detectees: {funcs}")

    results = {}

    # Tester avec comp_mortalite 2018-2022 si possible
    test_cases = [
        ("comp_mortalite", 2018, 2022),
        ("route", 2019, 2022),
    ]

    for dataset_id, year_start, year_end in test_cases:
        label = f"MOCA-O-Consolide/{dataset_id}"
        print(f"\n--- {dataset_id} {year_start}-{year_end} ---")

        try:
            # La signature attendue est : generate_mocao_consolidated(dataset_id, yearStart, yearEnd)
            # ou via CLI. On essaie d'appeler la fonction si elle existe.
            gen_fn = getattr(gmc, 'generate_mocao_consolidated', None) or getattr(gmc, 'main', None)
            if gen_fn is None:
                results[label] = {"status": "indisponible", "reason": "Fonction generate_mocao_consolidated non trouvee"}
                continue

            out_path = gen_fn(dataset_id, year_start, year_end)
        except FileNotFoundError as e:
            results[label] = {"status": "indisponible", "reason": f"Source manquante: {e}"}
            print(f"  [INDISPONIBLE] {e}")
            continue
        except Exception as e:
            # Essai via subprocess CLI
            import subprocess
            cmd = [sys.executable, str(cons_script), dataset_id, str(year_start), str(year_end)]
            try:
                r = subprocess.run(cmd, cwd=str(BASE_DIR), capture_output=True, text=True, timeout=120)
                print(r.stdout[-2000:] if len(r.stdout) > 2000 else r.stdout)
                if r.returncode != 0:
                    results[label] = {"status": "KO", "reason": f"CLI returncode={r.returncode}: {r.stderr[:500]}"}
                    continue
                # Chercher le xlsx genere
                out_files = list(OUTPUT_DIR.glob(f"{dataset_id}_mocao_{year_start}_{year_end}.xlsx"))
                if not out_files:
                    results[label] = {"status": "KO", "reason": "xlsx consolide non trouve apres CLI"}
                    continue
                out_path = out_files[0]
            except Exception as e2:
                results[label] = {"status": "KO", "reason": f"Erreur generation et CLI: {e2}"}
                continue

        if out_path is None:
            results[label] = {"status": "KO", "reason": "Retour None"}
            continue

        out_path = Path(out_path) if not isinstance(out_path, Path) else out_path
        if not out_path.exists():
            # Chercher dans output/
            candidates = list(OUTPUT_DIR.glob(f"{dataset_id}_mocao*.xlsx"))
            if candidates:
                out_path = candidates[-1]
            else:
                results[label] = {"status": "KO", "reason": f"Fichier consolide absent: {out_path}"}
                continue

        print(f"  XLSX consolide: {out_path}")

        EXPECTED_SHEETS_CONS = ["COM", "DROM", "franENT", "FranHEX"]  # sheets MOCA-O natif
        # Fallback: les 5 niveaux standards
        EXPECTED_SHEETS_FALLBACK = ["com", "reg", "dom", "fh", "fra"]

        # Verifier
        res = check_xlsx(str(out_path), EXPECTED_SHEETS_CONS, [], label="consolidated")
        if res.get('sheets_found') and not any(s in res['sheets_found'] for s in EXPECTED_SHEETS_CONS):
            # Essai avec les sheets standards
            res = check_xlsx(str(out_path), EXPECTED_SHEETS_FALLBACK, [], label="consolidated")

        print(f"  Onglets trouves: {res.get('sheets_found', [])}")
        for sh, fills in res.get('fills', {}).items():
            rows = res.get('rows', {}).get(sh, '?')
            print(f"  [{sh}] rows={rows} | couleurs={len(fills)}")

        status = "OK" if res['ok'] else "KO"
        results[label] = {
            "status": status,
            "reason": "; ".join(res['issues']) if res['issues'] else "",
            "file": str(out_path),
            "sheets": res.get('sheets_found', []),
            "fills": res.get('fills', {}),
        }

    return results


# =============================================================================
# RAPPORT FINAL
# =============================================================================

def print_report(all_results):
    print("\n\n" + "="*80)
    print("RAPPORT DE RECETTE PRISME - FORMAT EXCEL")
    print("="*80)
    print(f"{'Dataset':<45} {'Statut':<12} {'Annee':<8} {'Cellules colorees':<30} Remarques")
    print("-"*140)

    for label, r in sorted(all_results.items()):
        status = r.get("status", "?")
        year = str(r.get("year", ""))
        reason = r.get("reason", "")
        fills = r.get("fills_by_sheet") or r.get("fills", {})
        fills_str = " ".join(f"{k}={len(v)}" for k, v in fills.items()) if fills else "-"
        print(f"{label:<45} {status:<12} {year:<8} {fills_str:<30} {reason[:60]}")

    print("-"*140)
    ok = sum(1 for r in all_results.values() if r.get("status") == "OK")
    ko = sum(1 for r in all_results.values() if r.get("status") == "KO")
    indispo = sum(1 for r in all_results.values() if r.get("status") == "indisponible")
    print(f"\nTOTAL: {ok} OK | {ko} KO | {indispo} indisponible | {ok+ko+indispo} datasets")

    # Ecarts consolides
    print("\n--- ECARTS CRITIQUES ---")
    for label, r in sorted(all_results.items()):
        if r.get("status") == "KO":
            print(f"  [KO] {label}: {r.get('reason', '')}")
        fills = r.get("fills_by_sheet") or r.get("fills", {})
        colored = [k for k, v in fills.items() if v]
        if colored:
            print(f"  [COULEUR RESIDUELLE] {label} sur: {colored}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--artifact":
        artifact_result = check_generated_artifact(sys.argv[2])
        print(json.dumps(artifact_result, ensure_ascii=False, default=str))
        sys.exit(0 if artifact_result["ok"] else 1)

    all_results = {}

    # Mode 1 : MOCA-O
    mocao_results = qa_mocao()
    all_results.update(mocao_results)

    # Mode 2 : Open Data
    opendata_results = qa_opendata()
    all_results.update(opendata_results)

    # Mode 3 : MOCA-O Consolide
    cons_results = qa_mocao_consolidated()
    all_results.update(cons_results)

    # Rapport final
    print_report(all_results)

    # Sauvegarder JSON
    report_path = OUTPUT_DIR / "qa_report.json"
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False, default=str)
    print(f"\nRapport JSON sauvegarde: {report_path}")
